#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import sys
import time
import openpyxl
import subprocess

# SCRIPT set up with incrontab to run every time the XLSX is modified and closed. ## NE MARCHE PAS AVEC NTFS, utiliser crontab plutot
# 		/media/n06lbth/sauvegardes_pgm/checkMut_requests.xlsx IN_CLOSE_WRITE python .../scripts/checkMut_requests.py
# 		sudo service incron restart

pipeline_folder = os.environ['NGS_PIPELINE_BX_DIR']

# IF CHECKMUT-REQUESTS ALREADY RUNNING, QUIT
p = subprocess.Popen(['pgrep', '-f', 'your_script.py'], stdout=subprocess.PIPE)
out, err = p.communicate()
# print len(out.strip())
# print out
# print err

if len(out.strip()) > 1: # if another instance of this script is running, value would be 2 (other instance + this instance)
	print "checkMut requests is already running"
	exit()

# START
wb = openpyxl.load_workbook('/media/n06lbth/sauvegardes_pgm/checkMut_requests.xlsx')
ws = wb['checkMut']
checkMut_path = '%s/checkMut/checkMut.py' % pipeline_folder

print "parsing file..."
for row_idx in range(10, ws.max_row+1):
	run_path = ws.cell(row=row_idx,column=1).value
	gene = ws.cell(row=row_idx,column=2).value.upper().replace(' ','')
	cpos = ws.cell(row=row_idx,column=3).value.replace(' ','')
	run_type = ws.cell(row=row_idx,column=4).value.replace(' ','')
	min_sample = ws.cell(row=row_idx,column=5).value
	use_processed = ws.cell(row=row_idx,column=6).value
	if min_sample != None:
		min_sample = str(int(min_sample))
	state = ws.cell(row=row_idx,column=7).value

	if (run_path and gene and cpos and run_type) and (state is None): # (state != 'OK')
		run_folder = run_path.replace('\\','/') # ex avant : \\ZISILON01\N06lbth\sauvegardes_pgm\SBT\Run_500-599\Auto_user_S5...
		run_folder = run_folder.split('sauvegardes_pgm/')[-1]
		run_folder = '/media/n06lbth/sauvegardes_pgm/%s' % run_folder

		print "Running checkMut with param:"
		print " - run_folder : %s" % run_folder
		print " - gene : %s" % gene
		print " - cpos : %s" % cpos
		print " - run_type : %s" % run_type
		print " - min_sample : %s" % min_sample


		cmd_args = ['python',checkMut_path,'--run-folder',run_folder,'--gene',gene,'--cpos',cpos,'--run-type',run_type]
		if min_sample != None:
			cmd_args.append('--min-sample')
			cmd_args.append(str(min_sample))
		if use_processed != None:
			if use_processed.lower() == 'true':
				cmd_args.append('--use-processed')
		print ' '.join(cmd_args)
		cmd = subprocess.Popen(cmd_args,stdout=subprocess.PIPE, stderr=subprocess.PIPE)
		stdout, stderr = cmd.communicate()

		fig_path = '%s/_checkMut/%s_%s.png' % (run_folder,gene,cpos.replace('>','_'))
		if os.path.exists(fig_path):
			state = 'OK'
			link = fig_path.replace('/media/n06lbth/sauvegardes_pgm/','').replace('/','\\')
			ws.cell(row=row_idx,column=8).hyperlink = link
			ws.cell(row=row_idx,column=8).value = 'résultat'
			ws.cell(row=row_idx,column=8).style = 'Hyperlink'
		else:
			state = 'error'
			ws.cell(row=row_idx,column=9).value = stdout
			ws.cell(row=row_idx,column=10).value = stderr
		ws.cell(row=row_idx,column=7).value = state

		write_success = False
		for i in range(6):
			try:
				wb.save('/media/n06lbth/sauvegardes_pgm/checkMut_requests.xlsx')
				write_success = True
				break
			except Exception as e:
				print e.message
				print "- retrying in 10sec..."
				time.sleep(5)
		if not write_success :
			print "Cannot write file after 6 try in 30 sec. Aborting."
