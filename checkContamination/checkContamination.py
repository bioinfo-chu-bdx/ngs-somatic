#!/usr/bin/env python
from optparse import OptionParser
import subprocess
import openpyxl
import sqlite3
import pysam
import time
import json
import copy
import csv
import sys
import os

""" Script to analyse control contamination """


def dict_factory(cursor, row):
	d = {}
	for idx, col in enumerate(cursor.description):
		d[col[0]] = row[idx]
	return d

def representsInt(s):
	try: 
		s = int(s)
		return s
	except ValueError:
		return s

def get_column2index(sheet): # renvoie dict col2index
	col2index = {}
	for i in range(1,sheet.max_column+1):
		col2index[sheet.cell(row=1,column=i).value] = i
	return col2index

def check_stderr(stderr_path,indent=0):
	with open(stderr_path,'r') as stderr:
		for line in stderr.readlines():
			line = line.replace('\n','')
			for i in range(indent):
				line = '\t%s' % line
			print line

############################################################################################
parser = OptionParser()
parser.add_option('-i', '--run-folder', help="Run folder", dest='run_folder')
(options, args) = parser.parse_args()

control_names = ['H2O','H20','NTC'] # liste des noms possibles pour les temoins negatifs

pipeline_folder = os.environ['NGS_PIPELINE_BX_DIR']
with open('%s/global_parameters.json' % pipeline_folder, 'r') as g:
	global_param = json.loads(g.read().replace('$NGS_PIPELINE_BX_DIR',os.environ['NGS_PIPELINE_BX_DIR']))

if os.path.isfile('%s/barcodes.json' % options.run_folder):
	with open(options.run_folder+'/barcodes.json', 'r') as g:
		barcodes_json = json.load(g)
else:
	print "error : barcodes.json not found in run folder"
	exit()

db_path = global_param['VariantBase']
db_con = sqlite3.connect(db_path)
db_con.row_factory = dict_factory
db_cur = db_con.cursor()

checkconta_folder = '%s/_checkContamination' % options.run_folder
if not os.path.isdir(checkconta_folder):
	subprocess.call(['mkdir', checkconta_folder])

if options.run_folder.endswith('/'):
	run_id = os.path.basename(os.path.dirname(options.run_folder))
else:
	run_id = os.path.basename(options.run_folder)
db_cur.execute("SELECT platform FROM Run WHERE runID = '%s'" % run_id)
db_run = db_cur.fetchone()
platform = db_run['platform']

ok_barcodes = []
for barcode in barcodes_json:
	db_cur.execute("SELECT analysisID,isControl FROM Analysis INNER JOIN Sample ON Sample.sampleID = Analysis.sample WHERE Sample='%s' AND Run='%s'" % (barcodes_json[barcode]['sample_id'],run_id)) # get analysis ID from DB is safer than from barcodes.json
	db_analysis = db_cur.fetchone()
	barcodes_json[barcode]['analysis_id'] = db_analysis['analysisID']
	barcodes_json[barcode]['sample_folder'] = '%s/%s' % (options.run_folder,barcodes_json[barcode]['sample'])
	barcodes_json[barcode]['checkConta_read_len'] = global_param['panel'][barcodes_json[barcode]['panel']]['checkContamination_read_len']
	barcodes_json[barcode]['bam'] = '%s/%s/%s_%s.bam' % (options.run_folder,barcodes_json[barcode]['sample'],barcodes_json[barcode]['sample'],barcode)
	barcodes_json[barcode]['finalreport'] = '%s/%s/%s_%s.finalReport.xlsx' % (options.run_folder,barcodes_json[barcode]['sample'],barcodes_json[barcode]['sample'],barcode)
	barcodes_json[barcode]['is_control'] = db_analysis['isControl']
	if '-checkContamination' in barcodes_json[barcode]['sample']:
		print "- note : sample %s is already a filtered one and will not be used" % barcodes_json[barcode]['sample']
		continue
	if not os.path.exists(barcodes_json[barcode]['bam']):
		print "(warning : %s not found)" % barcodes_json[barcode]['bam']
		continue
	ok_barcodes.append(barcode)

###  __   __   __   __   ___  __   __      ___       __           __   __       ___  __   __       
### |__) |__) /  \ /  ` |__  /__` /__`    |__   /\  /  ` |__|    /  ` /  \ |\ |  |  |__) /  \ |    
### |    |  \ \__/ \__, |___ .__/ .__/    |___ /~~\ \__, |  |    \__, \__/ | \|  |  |  \ \__/ |___ 

for control_barcode in ok_barcodes:
	if barcodes_json[control_barcode]['is_control'] == 0:
		continue
	is_H2O_NTC = False
	for control_name in control_names:
		if control_name in barcodes_json[control_barcode]['sample']:
			is_H2O_NTC = True
	if not is_H2O_NTC:
		continue
	print "- Analysing contamination of %s :" % barcodes_json[control_barcode]['sample']

	###################
	### TARGET DATA ### 
	###################

	db_cur.execute("SELECT targetedRegionName,TargetedRegion.chromosome,start,stop,transcript,gene,details FROM TargetedRegion INNER JOIN Transcript ON Transcript.transcriptID = TargetedRegion.transcript WHERE panel = '%s'" % barcodes_json[control_barcode]['panel'])
	db_regions = db_cur.fetchall()
	region_coverage = {}
	for db_region in db_regions:
		region_coverage[db_region['targetedRegionName']] = {}
		region_coverage[db_region['targetedRegionName']]['chr'] = db_region['chromosome']
		region_coverage[db_region['targetedRegionName']]['start'] = db_region['start']
		region_coverage[db_region['targetedRegionName']]['stop'] = db_region['stop']
		region_coverage[db_region['targetedRegionName']]['transcript'] = db_region['transcript']
		region_coverage[db_region['targetedRegionName']]['gene_details'] = '%s_%s' % (db_region['gene'],db_region['details'])
		region_coverage[db_region['targetedRegionName']]['contamination count'] = 0

	######################
	### BAM FILTERING  ###
	######################

	# CREATING FILTERED CONTROL ("checkConta") FOLDER, BARCODE, NAME, ID, ...
	checkConta_bam_folder = '%s-checkContamination' % barcodes_json[control_barcode]['sample_folder']
	checkConta_barcode = '%s-checkContamination' % control_barcode
	checkConta_sample = '%s-checkContamination' % barcodes_json[control_barcode]['sample']
	checkConta_sample_id = '%s-checkContamination' % barcodes_json[control_barcode]['sample_id']
	checkConta_analysis_id = '%s-checkContamination' % barcodes_json[control_barcode]['analysis_id']
	checkConta_bam = '%s/%s_%s.bam' % (checkConta_bam_folder,checkConta_sample,checkConta_barcode)

	if not os.path.isdir(checkConta_bam_folder):
		print "\t- creating filtered bam folder..."
		subprocess.call(['mkdir',checkConta_bam_folder])

	# ADDING FILTERED CONTROL ENTRY IN BARCODES JSON
	if checkConta_barcode not in barcodes_json.keys():
		print "\t- adding filtered bam in barcodes_json..."
		barcodes_json[checkConta_barcode] = {}
		for key in barcodes_json[barcode].keys(): # ON FAIT D'ABORD UNE COPIE DES DONNEES DU BARCODE CONTROL ORIGINEL
			barcodes_json[checkConta_barcode][key] = barcodes_json[control_barcode][key]
		barcodes_json[checkConta_barcode]['sample'] = checkConta_sample
		barcodes_json[checkConta_barcode]['sample_id'] = checkConta_sample_id
		barcodes_json[checkConta_barcode]['analysis_id'] = checkConta_analysis_id

		print "- Re-writing barcodes JSON..."
		with open('%s/barcodes.json' % options.run_folder, 'r') as g:
			mod_json = json.load(g)
		mod_json[checkConta_barcode] = barcodes_json[checkConta_barcode]
		json_text = json.dumps(mod_json, indent=4, sort_keys=True)
		bc_json = open('%s/barcodes.json' % options.run_folder,'w')
		bc_json.write(json_text)
		bc_json.close()

		db_cur.execute("INSERT INTO Sample (sampleID, sampleName, isControl) VALUES ('%s', '%s', 1)" % (checkConta_sample_id,checkConta_sample))
		db_cur.execute("INSERT INTO Analysis (analysisID, sample, barcode, run, panel, bamPath, analysisDate) VALUES ('%s', '%s', '%s', '%s', '%s', '%s', '%s')" % (checkConta_analysis_id, checkConta_sample_id, checkConta_barcode, run_id, barcodes_json[control_barcode]['panel'], checkConta_bam, time.strftime("%Y%m%d")))
		db_con.commit()

	bamfile = pysam.Samfile(barcodes_json[control_barcode]['bam'],'rb')
	checkConta_bamfile = pysam.Samfile(checkConta_bam, 'wb', template=bamfile)

	### CREATE NEW BAM WITH READS > XX bases
	print "\t- bam filtering..."
	for read in bamfile.fetch():
		if (len(read.query) >= int(barcodes_json[control_barcode]['checkConta_read_len'])): # aligned portion of the read, exclude soft-clipped bases
			checkConta_bamfile.write(read)
	bamfile.close()
	checkConta_bamfile.close()

	# SORT NEW BAM AND CREATE BAI
	print "\t- samtools sort..."
	subprocess.call(['samtools','sort',checkConta_bam,'-o',checkConta_bam])
	print "\t- samtools index..."
	subprocess.call(['samtools','index',checkConta_bam])

	########################
	### NEW BAM ANALYSIS ###
	########################

	print "\t- filtered bam complete analysis..."
	cmd = subprocess.call(['python','%s/run_analysis.py' % pipeline_folder,'--sample',checkConta_bam_folder,'--skip-pre-processing'],stdout=open('%s/checkConta_bam_analysis.stdout.txt' % checkConta_bam_folder,'w'),stderr=open('%s/checkConta_bam_analysis.stderr.txt' % checkConta_bam_folder,'w'))
	check_stderr('%s/checkConta_bam_analysis.stderr.txt' % checkConta_bam_folder, indent=1)

	##################################################################
	### COMPARAISON AVEC COUVERTURE / VARIANTS DES AUTRES PATIENTS ###
	##################################################################

	print "\t- comparing with other samples..."
	# creation dic avec variants de l'annotation du check conta
	variants_checkConta = {}
	checkConta_finalreport = openpyxl.load_workbook('%s/%s_%s.finalReport.xlsx' % (checkConta_bam_folder,checkConta_sample,checkConta_barcode))
	if platform == 'ion torrent':
		checkConta_covSheet = checkConta_finalreport.get_sheet_by_name('Amplicon Coverage')
	else:
		checkConta_covSheet = checkConta_finalreport.get_sheet_by_name('Target Coverage')
	checkConta_annoSheet = checkConta_finalreport.get_sheet_by_name('Annotation')
	checkConta_covcol = get_column2index(checkConta_covSheet)
	checkConta_anncol = get_column2index(checkConta_annoSheet)
	# PARSING TARGET COVERAGE
	for i in range(2,checkConta_covSheet.max_row+1):
		if platform == 'ion torrent':
			region_coverage[checkConta_covSheet.cell(row=i,column=checkConta_covcol['region_id']).value]['contamination count'] = int(checkConta_covSheet.cell(row=i,column=checkConta_covcol['total_reads']).value)
		else:
			region_coverage[checkConta_covSheet.cell(row=i,column=checkConta_covcol['region_id']).value]['contamination count'] = int(checkConta_covSheet.cell(row=i,column=checkConta_covcol['ave_basereads']).value)
	# PARSING ANNOTATION SHEET
	for i in range(2,checkConta_annoSheet.max_row+1):
		if checkConta_annoSheet.cell(row=i,column=checkConta_anncol['Transcript']).value:
			transcript = checkConta_annoSheet.cell(row=i,column=checkConta_anncol['Transcript']).value.split('.')[0]
			c_nomen = checkConta_annoSheet.cell(row=i,column=checkConta_anncol['c.']).value
			variants_checkConta[(transcript,c_nomen)] = []

	# recuperation de la liste des patients a comparer
	sample2compare = []
	for barcode in barcodes_json:
		if (barcodes_json[barcode]['is_control'] == 0) and (barcodes_json[barcode]['library'] == barcodes_json[control_barcode]['library']) and (barcodes_json[barcode]['panel'] == barcodes_json[control_barcode]['panel']):
			sample2compare.append((barcodes_json[barcode]['sample'],barcode))
			if os.path.isfile(barcodes_json[barcode]['finalreport']):
				finalReport = openpyxl.load_workbook(barcodes_json[barcode]['finalreport'])
				if platform == 'ion torrent':
					covSheet = finalReport.get_sheet_by_name('Amplicon Coverage')
				else:
					covSheet = finalReport.get_sheet_by_name('Target Coverage')
				annoSheet = finalReport.get_sheet_by_name('Annotation')
				covcol = get_column2index(covSheet)
				anncol = get_column2index(annoSheet)
				for i in range(2,covSheet.max_row+1):
					region = covSheet.cell(row=i,column=covcol['region_id']).value
					if platform == 'ion torrent':
						totalReads = covSheet.cell(row=i,column=covcol['total_reads']).value
					else:
						totalReads = covSheet.cell(row=i,column=covcol['ave_basereads']).value
					region_coverage[region][barcodes_json[barcode]['sample']] = int(round(float(totalReads)))
				for i in range(2,annoSheet.max_row+1):
					if annoSheet.cell(row=i,column=anncol['Transcript']).value:
						transcript = annoSheet.cell(row=i,column=anncol['Transcript']).value.split('.')[0]
						c_nomen = annoSheet.cell(row=i,column=anncol['c.']).value
						if (transcript,c_nomen) in variants_checkConta.keys():
							variants_checkConta[(transcript,c_nomen)].append('%s_%s' % (barcodes_json[barcode]['sample'],barcode))

	###############################
	#### ecriture des resultats ###
	###############################

	print "\t- writing results..."
	conta_report = openpyxl.Workbook()
	regionSheet = conta_report.create_sheet(title='Regions Contamination')
	annotationSheet = conta_report.create_sheet(title='Annotation')
	try:
		del conta_report['Sheet']
	except:
		pass

	# REGION SHEET
	header = ['Region','Chr','Start','End','Gene_details','Transcript','%s reads > %s b' % (barcodes_json[control_barcode]['sample'],barcodes_json[control_barcode]['checkConta_read_len'])]
	for s in sample2compare:
		regular_sample = s[0]
		header.append(regular_sample)
	for i in range(len(header)):
		regionSheet.cell(row=1,column=i+1).value = header[i]
		regionSheet.cell(row=1,column=i+1).font = openpyxl.styles.Font(name='Calibri', size=11, bold=True)
		regionSheet.cell(row=1,column=i+1).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'),bottom=openpyxl.styles.Side(style='thin'))

	to_write = []
	for region in region_coverage.keys():
		line = [region,region_coverage[region]['chr'],region_coverage[region]['start'],region_coverage[region]['stop'],region_coverage[region]['gene_details'],region_coverage[region]['transcript'],region_coverage[region]['contamination count']]
		for s in sample2compare:
			try:
				line.append(region_coverage[region][s[0]])
			except:
				line.append('?')
		to_write.append(line)
	to_write.sort(key=lambda x: x[6]) # classement par ordre decroissant de nombre de reads > read_len
	to_write.reverse()

	for i in range(len(to_write)):
		for j in range(len(to_write[i])):
			regionSheet.cell(row=i+2,column=j+1).font = openpyxl.styles.Font(name='Calibri', size=11)
			if j <= 6:
				regionSheet.cell(row=i+2,column=j+1).value = representsInt(to_write[i][j])
				if j == 6 and int(to_write[i][6]) >= 100:
					regionSheet.cell(row=i+2,column=j+1).font = openpyxl.styles.Font(name='Calibri', size=11, color='ff0000')
			elif j > 6:
				try:
					if int(to_write[i][j]) < 5*(int(to_write[i][6])):
						regionSheet.cell(row=i+2,column=j+1).value = representsInt(to_write[i][j])
						regionSheet.cell(row=i+2,column=j+1).fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='d28e8e')
					else:
						regionSheet.cell(row=i+2,column=j+1).value = representsInt(to_write[i][j])
				except:
					regionSheet.cell(row=i+2,column=j+1).fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='d28e8e')

	# ANNOTATION SHEET
	annotationSheet.cell(row=1,column=1).value = 'Library Search'
	annotationSheet.cell(row=1,column=1).font = openpyxl.styles.Font(name='Calibri', size=11, bold=True)
	annotationSheet.cell(row=1,column=1).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'),bottom=openpyxl.styles.Side(style='thin'))
	for j in range(1,checkConta_annoSheet.max_column+1):
		annotationSheet.cell(row=1,column=j+1).value = checkConta_annoSheet.cell(row=1,column=j).value
		annotationSheet.cell(row=1,column=j+1).font = openpyxl.styles.Font(name='Calibri', size=11, bold=True)
		annotationSheet.cell(row=1,column=j+1).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'),bottom=openpyxl.styles.Side(style='thin'))

	for i in range(2,checkConta_annoSheet.max_row+1):
		if checkConta_annoSheet.cell(row=i,column=checkConta_anncol['Transcript']).value:
			if not checkConta_annoSheet.cell(row=i,column=checkConta_anncol['Commentaire']).value:
				annotationSheet.cell(row=i,column=checkConta_anncol['Commentaire']+1).value = '.'
			transcript = checkConta_annoSheet.cell(row=i,column=checkConta_anncol['Transcript']).value.split('.')[0]
			c_nomen = checkConta_annoSheet.cell(row=i,column=checkConta_anncol['c.']).value
			if variants_checkConta[(transcript,c_nomen)]: # list not empty = variant found somewhere
				foundstring = ','.join(variants_checkConta[(transcript,c_nomen)])
				annotationSheet.cell(row=i,column=1).value = 'Variant found in ' + foundstring
				annotationSheet.cell(row=i,column=1).font = openpyxl.styles.Font(name='Calibri', size=11, color='ff0000')
			else:
				annotationSheet.cell(row=i,column=1).value = 'not found'
				annotationSheet.cell(row=i,column=1).font = openpyxl.styles.Font(name='Calibri', size=11, color='00b050')
		for j in range(1,checkConta_annoSheet.max_column+1):
			if checkConta_annoSheet.cell(row=i,column=j).value:
				annotationSheet.cell(row=i,column=j+1).value = representsInt(checkConta_annoSheet.cell(row=i,column=j).value)
				annotationSheet.cell(row=i,column=j+1).font = openpyxl.styles.Font(name='Calibri', size=11)
				annotationSheet.cell(row=i,column=j+1).fill = copy.copy(checkConta_annoSheet.cell(row=i,column=j).fill)

	# coloration rouge cosmic
	for i in range(2,checkConta_annoSheet.max_row+1):
		if annotationSheet.cell(row=i,column=checkConta_anncol['COSMIC']+1).value and annotationSheet.cell(row=i,column=checkConta_anncol['COSMIC']+1).value != '.':
			annotationSheet.cell(row=i,column=checkConta_anncol['COSMIC']+1).font = openpyxl.styles.Font(name='Calibri', size=11, color='ff0000')

	## LAYOUT
	ws = conta_report['Regions Contamination']
	ws.column_dimensions[ws['A1'].column].width = 18
	ws.column_dimensions[ws['B1'].column].width = 8
	ws.column_dimensions[ws['C1'].column].width = 12
	ws.column_dimensions[ws['D1'].column].width = 12
	ws.column_dimensions[ws['E1'].column].width = 15
	ws.column_dimensions[ws['F1'].column].width = 15
	ws = conta_report['Annotation']
	maxsize = 15
	dims = {}
	for row in ws.rows:
		for cell in row:
			if cell.value:
				try:
					text_size = len(str(cell.value)) + 2
					dims[cell.column] = min(max(dims.get(cell.column, 0), text_size),maxsize)
				except:
					pass
	for col, value in dims.items():
		ws.column_dimensions[col].width = value
	ws.column_dimensions[ws['J1'].column].width = 8

	conta_report.save(checkconta_folder+'/checkContamination_%s.xlsx' % barcodes_json[control_barcode]['sample'])
	subprocess.call(['mv',checkConta_bam_folder,checkconta_folder]) # OU RM ??
