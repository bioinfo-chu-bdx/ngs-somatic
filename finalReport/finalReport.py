#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import sys
import json
import csv
import urllib2
import openpyxl
from PIL import Image as PILimage
from openpyxl.drawing.image import Image
import copy
import glob
import ast
import time
import zipfile
from subprocess import *
from operator import itemgetter 
from optparse import OptionParser
import sqlite3

def dict_factory(cursor, row):
	d = {}
	for idx, col in enumerate(cursor.description):
		d[col[0]] = row[idx]
	return d

def alamut_variants_vbscript(intermediate_folder, variant_list):
	request = ','.join(variant_list)
	vbs = open("%s/ALAMUT_load_variants.vbs" % intermediate_folder, 'w')
	vbs.write('Set objHTTP = CreateObject("MSXML2.ServerXMLHTTP")\r\n')
	vbs.write('URL = "http://localhost:10000/show?request=%s"\r\n' % request)
	vbs.write('objHTTP.Open "GET", URL, False\r\n')
	vbs.write('objHTTP.send ("")\r\n')
	vbs.close()

def alamut_bam_vbscript(intermediate_folder, sample, barcode, processed):
	if processed:
		vbs = open("%s/ALAMUT_load_processed_bam.vbs" % intermediate_folder, 'w')
	else:
		vbs = open("%s/ALAMUT_load_bam.vbs" % intermediate_folder, 'w')
	vbs.write('Set objHTTP = CreateObject("MSXML2.ServerXMLHTTP")\r\n')
	vbs.write('dim fso: set fso = CreateObject("Scripting.FileSystemObject")\r\n')
	vbs.write('dim CurrentDirectory\r\n')
	vbs.write('CurrentDirectory = fso.GetAbsolutePathName(".")\r\n')
	vbs.write('dim BamPath\r\n')
	if processed:
		vbs.write('BamPath= fso.BuildPath(CurrentDirectory, "%s_%s.processed.bam")\r\n' % (sample,barcode))
	else:
		vbs.write('BamPath= fso.BuildPath(CurrentDirectory, "%s_%s.bam")\r\n' % (sample,barcode))
	vbs.write('URL = "http://localhost:10000/show?request=BAM<" & BamPath\r\n')
	vbs.write('objHTTP.Open "GET", URL, False\r\n')
	vbs.write('objHTTP.send ("")\r\n')
	vbs.close()

def print_vbscript(intermediate_folder, sample, barcode, project):
	vbs = open("%s/PRINT_finalReport.vbs" % intermediate_folder, 'w')
	vbs.write('Dim iAnswer\r\n')
	vbs.write('iAnswer = MsgBox("Voulez-vous lancer l\'impression?", vbOKCancel + vbQuestion, "Continue")\r\n')
	vbs.write('if iAnswer = vbCancel Then\r\n')
	vbs.write('WScript.quit()\r\n')
	vbs.write('End If\r\n')
	vbs.write('Set exl = WScript.CreateObject("Excel.Application")\r\n')
	vbs.write('exl.Visible = True \'False\r\n')
	vbs.write('Set sh = WScript.CreateObject("WScript.Shell")\r\n')
	vbs.write('on error resume next\r\n')
	vbs.write('sh.RegWrite "HKEY_CURRENT_USER\Software\Microsoft\Office\\11.0\Excel\Security\\accessVBOM",1,"REG_DWORD"\r\n')
	vbs.write('on error goto 0\r\n')
	vbs.write('dim fso: set fso = CreateObject("Scripting.FileSystemObject")\r\n')
	vbs.write('dim CurrentDirectory\r\n')
	vbs.write('CurrentDirectory = fso.GetAbsolutePathName(".")\r\n')
	vbs.write('dim FinalReportlist\r\n')
	vbs.write('Set FinalReportlist = WScript.CreateObject("System.Collections.ArrayList")\r\n')
	vbs.write('FinalReportlist.Add CurrentDirectory & "\\%s_%s.finalReport.xlsx"\r\n' % (sample,barcode))
	vbs.write('For Each FinalReport in FinalReportlist\r\n')
	vbs.write('set fichxl = exl.workbooks.add(FinalReport)\r\n')
	vbs.write('Set mdle = fichxl.VBProject.VBComponents.Add(1)\r\n')
	vbs.write('num=0\r\n')
	vbs.write('num=num+1:mdle.CodeModule.InsertLines num, "Sub PrintPage1()"\r\n')
	vbs.write('num=num+1:mdle.CodeModule.InsertLines num, "With Sheets(""Annotation"").PageSetup"\r\n')
	vbs.write('num=num+1:mdle.CodeModule.InsertLines num, ".Orientation = xlLandscape"\r\n')
	vbs.write('num=num+1:mdle.CodeModule.InsertLines num, ".CenterHeader = ""Page &A"""\r\n')
	vbs.write('num=num+1:mdle.CodeModule.InsertLines num, ".RightHeader = ""Printed &D & &T"""\r\n')
	vbs.write('num=num+1:mdle.CodeModule.InsertLines num, ".CenterFooter = ""&F"""\r\n')
	vbs.write('num=num+1:mdle.CodeModule.InsertLines num, "End With"\r\n')
	vbs.write('num=num+1:mdle.CodeModule.InsertLines num, "Application.PrintCommunication = False"\r\n')
	vbs.write('num=num+1:mdle.CodeModule.InsertLines num, "With Sheets(""Annotation"").PageSetup"\r\n')
	vbs.write('num=num+1:mdle.CodeModule.InsertLines num, ".FitToPagesWide = 1"\r\n')
	vbs.write('num=num+1:mdle.CodeModule.InsertLines num, ".FitToPagesTall = False"\r\n')
	vbs.write('num=num+1:mdle.CodeModule.InsertLines num, ".LeftMargin = Application.InchesToPoints(0.1)"\r\n')
	vbs.write('num=num+1:mdle.CodeModule.InsertLines num, ".RightMargin = Application.InchesToPoints(0.1)"\r\n')
	vbs.write('num=num+1:mdle.CodeModule.InsertLines num, "End With"\r\n')
	vbs.write('num=num+1:mdle.CodeModule.InsertLines num, "Application.PrintCommunication = True"\r\n')
	vbs.write('num=num+1:mdle.CodeModule.InsertLines num, "NbLignes = Sheets(""Annotation"").UsedRange.Rows.Count"\r\n')
	vbs.write('num=num+1:mdle.CodeModule.InsertLines num, "Sheets(""Annotation"").Range(""A1:AZ1"").Interior.Color = vbWhite"\r\n')
	if project == 'SBT': # Tout prendre jusqu'a dbSNP, et enlever colonne "c.p.f"
		vbs.write('num=num+1:mdle.CodeModule.InsertLines num, "Sheets(""Annotation"").Range(""K:K"").EntireColumn.Hidden = True"\r\n')
		vbs.write('num=num+1:mdle.CodeModule.InsertLines num, "Sheets(""Annotation"").Range(""A1:U"" & NbLignes).printOut"\r\n')
	elif project in ['LAM','Leuc','FLT3','ABL1']:
		# ne pas prendre "Position", "Ref", "Alt", "Var.Cov." Prendre jusqu'a "dbSNP". Rajouter en fin "Class." et "c. Annovar"
		vbs.write('num=num+1:mdle.CodeModule.InsertLines num, "Sheets(""Annotation"").Range(""F:H,N:N,T:AF,AH:AI"").EntireColumn.Hidden = True"\r\n')
		vbs.write('num=num+1:mdle.CodeModule.InsertLines num, "Sheets(""Annotation"").Range(""A1:AJ"" & NbLignes).printOut"\r\n')
	elif project == 'TP53': # comme LAM mais decalage du aux colonnes "Patho UMD" et "Comment UMD"
		vbs.write('num=num+1:mdle.CodeModule.InsertLines num, "Sheets(""Annotation"").Range(""F:H,N:N,V:AH,AJ:AK"").EntireColumn.Hidden = True"\r\n')
		vbs.write('num=num+1:mdle.CodeModule.InsertLines num, "Sheets(""Annotation"").Range(""A1:AL"" & NbLignes).printOut"\r\n')
	else: # finalReport sans colonnes supplementaires (sensitivity, TP53-truc...)
		vbs.write('num=num+1:mdle.CodeModule.InsertLines num, "Sheets(""Annotation"").Range(""A1:S"" & NbLignes).printOut"\r\n')
	vbs.write('num=num+1:mdle.CodeModule.InsertLines num, "End Sub"\r\n')
	vbs.write('exl.Application.Run "PrintPage1"\r\n')
	vbs.write('exl.DisplayAlerts = False\r\n')
	vbs.write('exl.ActiveWorkbook.Close False\r\n')
	vbs.write('Next\r\n')
	vbs.close()

def representsInt(s): # pour eviter avertissement "nombre ecrit en texte" sous excel
	try: 
		s = int(s)
		return s
	except:
		return s

def cell_format(cell, font=None, alignment=None, color=None, format=None, border=None):
	# BORDER
	if border == 'thin':
		cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'),bottom=openpyxl.styles.Side(style='thin'))
	elif border == 'hair':
		cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='hair'),right=openpyxl.styles.Side(style='hair'), top=openpyxl.styles.Side(style='hair'),bottom=openpyxl.styles.Side(style='hair'))
	elif border == 'top_medium':
		cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='medium'))
	# FONT
	if font == 'bold':
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, bold=True)
	elif font == 'bigBold':
		cell.font = openpyxl.styles.Font(name='Calibri', size=13, bold=True)
	elif font == 'BlueBold':
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, color='004c99', bold=True)
	elif font == 'red':
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, color='f44242')
	elif font == 'boldDarkRed':
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, color='9e0000', bold=True)
	elif font == 'boldDarkGreen':
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, color='007332', bold=True)
	elif font == 'boldDarkOrange':
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, color='974706', bold=True)
	elif font == 'DarkRed':
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, color='9e0000')
	elif font == 'DarkGreen':
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, color='007332')
	elif font == 'DarkOrange':
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, color='974706')
	else:
		cell.font = openpyxl.styles.Font(name='Calibri', size=11)
	# ALIGNMENT
	if alignment == 'center':
		cell.alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
	elif alignment == 'left':
		cell.alignment = openpyxl.styles.Alignment(horizontal='left')
	elif alignment == 'right':
		cell.alignment = openpyxl.styles.Alignment(horizontal='right')
	elif alignment == 'wrap':
		cell.alignment = openpyxl.styles.Alignment(vertical='top',wrap_text=True)
	# CELL COLOR
	if color == 'LightGreen':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='D8E4BC')
	if color == 'DarkGreen':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='C3D696')
	elif color == 'LightRed':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='E6B8B7')
	elif color == 'DarkRed':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='e89e9b')
	elif color == 'LightOrange':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='ffe7e0')
	elif color == 'LightBlue':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='add8e6')
	elif color == 'LightPink':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='ffc9f7')
	elif color == 'LightPurple':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='ebe7f1')
	elif color == 'LightGrey':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='D9D9D9')
	elif color == 'DarkGrey':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='bfbfbf')
	elif color == 'Yellow':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='feffa3')
	elif color == 'Blue':
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, color='004c99')
	#FORMAT
	if format == 'Percent':
		cell.number_format = '0.0%'


### GATHERING PARAMETERS ############################################################

parser = OptionParser()
parser.add_option('-a', '--analysis', 	help="DB AnalysisID",				dest='analysis')
parser.add_option('-x', '--xmin',		help="Min X for amplicon coverage",	dest='xmin', default=300)
(options, args) = parser.parse_args()

analysis_id = options.analysis
minX = float(options.xmin)

pipeline_folder = os.environ['NGS_PIPELINE_BX_DIR']
with open('%s/global_parameters.json' % pipeline_folder, 'r') as g:
	global_param = json.loads(g.read().replace('$NGS_PIPELINE_BX_DIR',os.environ['NGS_PIPELINE_BX_DIR']))

db_path = global_param['VariantBase']
db_con = sqlite3.connect(db_path)
db_con.row_factory = dict_factory
db_cur = db_con.cursor()

db_cur.execute("SELECT * FROM Analysis WHERE analysisID='%s'"% analysis_id)
db_analysis = db_cur.fetchone()
if db_analysis is None:
	sys.exit("Error : analysisID %s not found in DB" % analysis_id)
barcode = db_analysis['barcode']
sample_id = db_analysis['sample']
run_id = db_analysis['run']
panel = db_analysis['panel']
db_cur.execute("SELECT panelProject,panelSubProject FROM Panel WHERE panelID='%s'"% panel)
db_panel = db_cur.fetchone()
project = db_panel['panelProject']
subproject = db_panel['panelSubProject']
bam_path = db_analysis['bamPath']
if os.path.exists(bam_path):
	sample_folder = os.path.dirname(bam_path)
else:
	db_cur.execute("SELECT * FROM Sample WHERE sampleID='%s'"% sample_id)
	db_sample = db_cur.fetchone()
	if subproject:
		sample_folder = '%s/%s/%s/%s/%s' % (global_param['ngs_results_folder'],project,subproject,run_id,db_sample['sampleName'])
	else:
		sample_folder = '%s/%s/%s/%s' % (global_param['ngs_results_folder'],project,run_id,db_sample['sampleName'])
if not os.path.exists(sample_folder):
	sys.exit("Error : sample folder not found (%s)" % sample_folder)

sample = sample_folder.split('/')[-1]
run_folder = os.path.dirname(sample_folder)
intermediate_folder = sample_folder+'/intermediate_files'

db_cur.execute("""SELECT TargetedRegion.chromosome,start,stop,targetedRegionName,transcript,gene,details FROM TargetedRegion 
INNER JOIN Transcript ON Transcript.transcriptID = TargetedRegion.transcript
WHERE panel='%s' ORDER BY start""" % panel)
db_target_regions = db_cur.fetchall()

cosmicDB_con = sqlite3.connect(global_param['cosmicDB'])
cosmicDB_con.row_factory = dict_factory
cosmicDB_cur = cosmicDB_con.cursor()

############################################################################################

finalReport = openpyxl.Workbook()
annotationSheet = finalReport.create_sheet(title='Annotation')
cnvSheet = finalReport.create_sheet(title='CNV')
plotSheet = finalReport.create_sheet(title='Plot Coverage')
coverageSheet = finalReport.create_sheet(title='Target Coverage')
baseCoverageSheet = finalReport.create_sheet(title='Base Coverage')

try:
	del finalReport['Sheet']
except:
	pass

#                 __  ___      ___    __           __        ___  ___ ___ 
#  /\  |\ | |\ | /  \  |   /\   |  | /  \ |\ |    /__` |__| |__  |__   |  
# /~~\ | \| | \| \__/  |  /~~\  |  | \__/ | \|    .__/ |  | |___ |___  |  


aheader = ['Commentaire','Gene','Exon','Transcript','Chr','Position','Ref','Alt','c.','p.','Region','Consequence','Freq','Var.Cov.','Depth','InterVar','ClinVar','COSMIC','dbSNP','gnomAD','1000G_ALL','1000G_EUR','NCI60','ESP','ExAC','SIFT','POLYPHEN2','PROVEAN','PubMed','VEP_Consequence','VEP_Impact','VEP_Diff','Class.','VC_CALL','c.(annovar)','p.(annovar)','annoWarning','hgvsInfo']

panel_transcripts = {}
db_cur.execute("""SELECT DISTINCT gene,TranscriptID FROM Transcript 
INNER JOIN TargetedRegion ON TargetedRegion.transcript = Transcript.transcriptID
WHERE panel = '%s'""" % panel)
db_transcripts = db_cur.fetchall()
for db_transcript in db_transcripts:
	panel_transcripts[str(db_transcript['transcriptID'])] = str(db_transcript['gene'])

db_cur.execute("SELECT * FROM VariantMetrics WHERE analysis='%s'"% (analysis_id))
db_vms = db_cur.fetchall()

variants = []
for db_vm in db_vms :
	if len(panel_transcripts.keys()) > 1:
		db_cur.execute("""SELECT * FROM VariantAnnotation 
		INNER JOIN Variant ON Variant.variantID = VariantAnnotation.variant
		WHERE variant='%s' AND transcript IN %s""" % (db_vm['variant'],tuple(panel_transcripts.keys())))
	else: # ex : panel TP53 only one gene
		db_cur.execute("""SELECT * FROM VariantAnnotation 
		INNER JOIN Variant ON Variant.variantID = VariantAnnotation.variant
		WHERE variant='%s' AND transcript='%s'""" % (db_vm['variant'],panel_transcripts.keys()[0]))
	db_variant = db_cur.fetchone()

	comm = db_variant['commentaire']
	db_cur.execute("SELECT userComment FROM UserComment WHERE panel='%s' AND variantAnnotation='%s'" % (panel,db_variant['variantAnnotationID']))
	db_userComments = db_cur.fetchall()
	for db_userComment in db_userComments:
		userComment = db_userComment['userComment']
		if comm is None:
			comm = userComment
		else:
			comm = '%s. %s' % (userComment,comm)

	variants.append({
		'Commentaire':comm,
		'Gene':panel_transcripts[db_variant['transcript']],
		'Exon':db_variant['exon'],
		'Transcript':db_variant['transcript'],
		'Chr':int(db_variant['chromosome'].replace('chr','').replace('X','23').replace('Y','24')),
		'Position':int(db_variant['genomicStart']),
		'Ref':db_variant['referenceAllele'],
		'Alt':db_variant['alternativeAllele'],
		'c.':db_variant['transcriptDescription'],
		'p.':db_variant['proteinDescription'],
		'Region':db_variant['region'],
		'Consequence':db_variant['consequence'],
		'Sensitivity':db_variant['actionability'],
		'Freq':(float(db_vm['variantReadDepth'])/float(db_vm['positionReadDepth'])) * 100.0,
		'Var.Cov.':db_vm['variantReadDepth'],
		'Depth':db_vm['positionReadDepth'],
		'InterVar':db_variant['intervar'],
		'ClinVar':db_variant['clinvar'],
		'COSMIC':db_variant['cosmic'],
		'dbSNP':db_variant['dbsnp'],
		'gnomAD':db_variant['gnomad'],
		'1000G_ALL':db_variant['milleGall'],
		'1000G_EUR':db_variant['milleGeur'],
		'NCI60':db_variant['nci60'],
		'ESP':db_variant['esp'],
		'ExAC':db_variant['exac'],
		'SIFT':db_variant['sift'],
		'POLYPHEN2':db_variant['polyphen2'],
		'PROVEAN':db_variant['provean'],
		'PubMed':db_variant['pubmed'],
		'VEP_Consequence':db_variant['vep_consequence'],
		'VEP_Impact':db_variant['vep_impact'],
		'VEP_Diff':db_variant['vep_diff'],
		'Class.':db_variant['variantType'],
		'VC_CALL':db_vm['call'],
		'c.(annovar)':db_variant['annovarTranscriptDescription'],
		'p.(annovar)':db_variant['annovarProteinDescription'],
		'hgvsInfo':db_variant['hgvsInfo'],
		'annoWarning':db_variant['annoWarning'],
		'highlight':db_variant['highlight'],
		'Patho UMD':db_variant['pathoUMD'],
		'Comment UMD':db_variant['commentUMD'],
		'c.p.f.':'%s ; %s ; %s' % (db_variant['transcriptDescription'],db_variant['proteinDescription'],(float(db_vm['variantReadDepth'])/float(db_vm['positionReadDepth']))*100.0),
		'c.p.':'%s ; %s' % (db_variant['transcriptDescription'],db_variant['proteinDescription']),
		'cosm.rs':'%s ; %s' % (str(db_variant['cosmic']).split(',')[0],str(db_variant['dbsnp']).split(',')[0])
	})

print " - [%s] Annotation sheet ..." % time.strftime("%H:%M:%S")

if project in ['SBT','Lymphome_B','Lymphome_T']:
	aheader.insert(aheader.index('p.')+1,'c.p.f.')
	if 'SBT' in project:
		aheader.insert(aheader.index('Consequence')+1,'Sensitivity')
elif 'TP53' in project:
	aheader.insert(aheader.index('COSMIC'),'Patho UMD')
	aheader.insert(aheader.index('COSMIC'),'Comment UMD')
	aheader.insert(len(aheader)+1,'c.p.f.')
else:
	aheader.insert(len(aheader)+1,'c.p.f.')
	aheader.insert(len(aheader)+1,'c.p.')
	aheader.insert(len(aheader)+1,'cosm.rs')

# writing annotation new aheader in finalreport
for i in range(len(aheader)):
	annotationSheet.cell(row=1,column=i+1).value = aheader[i]
	cell_format(annotationSheet.cell(row=1,column=i+1),font='bold',border='thin',color='DarkGrey')

#TRI
variants = sorted(sorted(variants,key=itemgetter('Position')),key=itemgetter('Chr'))

#ECRITURE
vb_variant_list = [] # for alamut vbscript
l=2
for variant in variants:
	variant['Chr'] = 'chr%s' % variant['Chr']
	variant['Chr'] = variant['Chr'].replace('23','X').replace('24','Y')

	#variant['Freq'] = variant['Freq'] * 100.0
	if variant['Freq'] > 1.0 :
		variant['Freq'] = int(round(variant['Freq']))
	else:
		variant['Freq'] = '%0.1f' % variant['Freq']

	variant['c.p.f.'] = '%s ; %s ; %s' % (variant['c.'],variant['p.'],variant['Freq'])

	# LISA RAJOUT TAILLE INS / DEL
	if variant['Class.'] == 'INS' or variant['Class.'] == 'DUP':
		variant['Class.'] = '%s (%s)' % (variant['Class.'],len(variant['Alt']))
	elif variant['Class.'] == 'DEL':
		variant['Class.'] = 'DEL (%s)' % len(variant['Ref'])
	elif variant['Class.'] == 'DELINS':
		variant['Class.'] = 'DELINS (%s/%s)' % (len(variant['Ref']),len(variant['Alt']))

	# LISA ASXL1 ex 12 = ex 13 
	## POUR NM_015338.5, Annovar trouve ex12 (faux), VEP trouve ex13. tant qu'annovar par defaut, garder cette correction.
	if variant['Gene'] == 'ASXL1' and variant['Exon'] == 12:
		variant['Exon'] = 13
		if variant['Commentaire'] != None:
			variant['Commentaire'] = '%s. Exon corrected (12->13)' % variant['Commentaire']
		else:
			variant['Commentaire'] = 'Exon corrected (12->13)'

	# COSMIC occurence for Lymphome B & T (haematopoietic_and_lymphoid_tissue)
	if project in ['Lymphome_B','Lymphome_T']:
		if variant['COSMIC']:
			cosmics = variant['COSMIC'].split(',')
			occ = []
			for c in cosmics:
				cosmicDB_cur.execute("SELECT haematopoietic_and_lymphoid_tissue FROM Cosmic WHERE cosmicID='%s' AND haematopoietic_and_lymphoid_tissue not NULL" % c)
				cosmicDB_data = cosmicDB_cur.fetchone()
				if cosmicDB_data:
					occ.append(str(cosmicDB_data['haematopoietic_and_lymphoid_tissue']))
				else:
					occ.append('0')
			occs = ','.join(occ)
			variant['COSMIC'] = '%s;occurence(haematopoietic_and_lymphoid_tissue)=%s' % (variant['COSMIC'],occs)

	# FORMATING RESULTS
	condition1_green_line = (variant['Region'] in ['exonic','splicing','ncRNA_exonic']) and (variant['Consequence'] != 'synonymous')
	condition2_green_line = (116411873 <= variant['Position'] <= 116411902) or (116412044 <= variant['Position'] <= 116412087)
	condition3_green_line = (variant['Region'] == '' or variant['Region'] == None) # Si pas d'annotation, mettre en vert au cas ou pour ne pas louper de variant important

	if condition1_green_line or condition2_green_line or condition3_green_line:
		vb_variant_list.append('%s:%s' % (variant['Gene'],variant['c.']))
		if variant['hgvsInfo'] != None:
			hgvsinfo = 'HGVS unverified (%s)' % variant['hgvsInfo']
			if variant['Commentaire'] == None:
				variant['Commentaire'] = hgvsinfo
			else:
				variant['Commentaire'] = '%s. %s' % (variant['Commentaire'],hgvsinfo)
		if condition3_green_line :
			variant['Commentaire'] = '%s. %s' % ('Region/Consequence inconnue',variant['Commentaire'])
			variant['Region'] = '?'
			variant['Consequence'] = '?'
		for columnName in aheader:
			cell_val = variant[columnName]
			annotationSheet.cell(row=l,column=aheader.index(columnName)+1).value = cell_val
			cell_format(annotationSheet.cell(row=l,column=aheader.index(columnName)+1),color='LightGreen',border='hair')
		cell_format(annotationSheet.cell(row=l,column=aheader.index('Commentaire')+1),color='LightGreen',border='hair',font='bold')
		cell_format(annotationSheet.cell(row=l,column=aheader.index('Freq')+1),color='LightGrey',border='hair',font='bold',alignment='right')
		cell_format(annotationSheet.cell(row=l,column=aheader.index('Var.Cov.')+1),color='LightGrey',border='hair')
		cell_format(annotationSheet.cell(row=l,column=aheader.index('Depth')+1),color='LightGrey',border='hair')
		if 'c.p.f.' in aheader:
			cell_format(annotationSheet.cell(row=l,column=aheader.index('c.p.f.')+1),color='LightGrey',border='hair')
		if 'c.p.' in aheader:
			cell_format(annotationSheet.cell(row=l,column=aheader.index('c.p.')+1),color='LightGrey',border='hair')
		if 'cosm.rs' in aheader:
			cell_format(annotationSheet.cell(row=l,column=aheader.index('cosm.rs')+1),color='LightGrey',border='hair')
		# intervar et clinvar (bold)
		if variant['InterVar'] != None:
			if 'likely pathogenic' in variant['InterVar']:
				cell_format(annotationSheet.cell(row=l,column=aheader.index('InterVar')+1),font='boldDarkOrange')
			elif 'pathogenic' in variant['InterVar']:
				cell_format(annotationSheet.cell(row=l,column=aheader.index('InterVar')+1),font='boldDarkRed')
			elif 'likely benign' in variant['InterVar'] or 'benign' in variant['InterVar']:
				cell_format(annotationSheet.cell(row=l,column=aheader.index('InterVar')+1),font='boldDarkGreen')
		if variant['ClinVar'] != None:
			s_clinvar = variant['ClinVar'].split('/')
			if 'pathogenic' in s_clinvar:
				cell_format(annotationSheet.cell(row=l,column=aheader.index('ClinVar')+1),font='boldDarkRed')
			elif 'likely_pathogenic' in s_clinvar:
				cell_format(annotationSheet.cell(row=l,column=aheader.index('ClinVar')+1),font='boldDarkOrange')
			elif 'likely_benign' in s_clinvar or 'benign' in s_clinvar:
				cell_format(annotationSheet.cell(row=l,column=aheader.index('ClinVar')+1),font='boldDarkGreen')
	else:
		for columnName in aheader:
			cell_val = variant[columnName]
			annotationSheet.cell(row=l,column=aheader.index(columnName)+1).value = cell_val
			cell_format(annotationSheet.cell(row=l,column=aheader.index(columnName)+1))
		cell_format(annotationSheet.cell(row=l,column=aheader.index('Freq')+1),alignment='right')
		# intervar et clinvar (no bold)
		if variant['InterVar'] != None:
			if 'Likely pathogenic' in variant['InterVar']:
				cell_format(annotationSheet.cell(row=l,column=aheader.index('InterVar')+1),font='DarkOrange')
			elif 'Pathogenic' in variant['InterVar']:
				cell_format(annotationSheet.cell(row=l,column=aheader.index('InterVar')+1),font='DarkRed')
			elif 'Likely benign' in variant['InterVar'] or 'Benign' in variant['InterVar']:
				cell_format(annotationSheet.cell(row=l,column=aheader.index('InterVar')+1),font='DarkGreen')
		if variant['ClinVar'] != None:
			s_clinvar = variant['ClinVar'].split('/')
			if 'Pathogenic' in s_clinvar:
				cell_format(annotationSheet.cell(row=l,column=aheader.index('ClinVar')+1),font='DarkRed')
			elif 'Likely_pathogenic' in s_clinvar:
				cell_format(annotationSheet.cell(row=l,column=aheader.index('ClinVar')+1),font='DarkOrange')
			elif 'Likely_benign' in s_clinvar or 'Benign' in s_clinvar:
				cell_format(annotationSheet.cell(row=l,column=aheader.index('ClinVar')+1),font='DarkGreen')

	# if condition3_green_line:
		# for columnName in aheader:
			# cell_val = variant[columnName]
			# annotationSheet.cell(row=l,column=aheader.index(columnName)+1).value = cell_val
			# cell_format(annotationSheet.cell(row=l,column=aheader.index(columnName)+1),border='hair')
		# annotationSheet.cell(row=l,column=aheader.index('Region')+1).value = '?'
		# cell_format(annotationSheet.cell(row=l,column=aheader.index('Region')+1),color='DarkGreen',border='hair')
		# annotationSheet.cell(row=l,column=aheader.index('Consequence')+1).value = '?'
		# cell_format(annotationSheet.cell(row=l,column=aheader.index('Consequence')+1),color='DarkGreen',border='hair')

	# HEMATO : GATA2 INTRON 4 ET ANKRD26 5'UTR VIOLET
	if (variant['Gene'] == 'GATA2' and (128200788 <= variant['Position'] <= 128202702)) or (variant['Gene'] == 'ANKRD26' and (27389256 <= variant['Position'] <= 27389427)):
		for columnName in aheader:
			cell_format(annotationSheet.cell(row=l,column=aheader.index(columnName)+1),color='LightPurple')
	# HEMATO : HOTSPOTS ROUGE
	if 'LAM' in project and variant['highlight'] == 1:
		cell_format(annotationSheet.cell(row=l,column=aheader.index('c.')+1),color='LightRed')
		cell_format(annotationSheet.cell(row=l,column=aheader.index('p.')+1),color='LightRed')

	# POSITION INTRONIQUE BLEUE
	if variant['c.'] != None:
		if ('+' in variant['c.'] or '-' in variant['c.'] or '*' in variant['c.']) and (variant['Region'] != 'exonic' and variant['Region'] != '?'):
			cell_format(annotationSheet.cell(row=l,column=aheader.index('c.')+1),color='Blue')
			cell_format(annotationSheet.cell(row=l,column=aheader.index('c.p.f.')+1),color='Blue')
			cell_format(annotationSheet.cell(row=l,column=aheader.index('Region')+1),color='Blue')

	# CONSEQUENCE ORANGE / ROUGE
	if variant['Consequence'] != None:
		if 'missense' in variant['Consequence'] or 'nonframeshift' in variant['Consequence']:
			cell_format(annotationSheet.cell(row=l,column=aheader.index('Consequence')+1),font='DarkOrange')
		elif 'stopgain' in variant['Consequence'] or 'stoploss' in variant['Consequence'] or 'frameshift' in variant['Consequence']:
			cell_format(annotationSheet.cell(row=l,column=aheader.index('Consequence')+1),font='DarkRed')
		elif '?' in variant['Consequence']:
			cell_format(annotationSheet.cell(row=l,column=aheader.index('Consequence')+1),font='DarkRed')
			cell_format(annotationSheet.cell(row=l,column=aheader.index('Region')+1),font='DarkRed')

	# SENSITIVITY VERT / ROUGE
	if 'Sensitivity' in aheader and variant['Sensitivity'] != None:
		s = variant['Sensitivity'].lower()
		if s.startswith('sensible'):
			cell_format(annotationSheet.cell(row=l,column=aheader.index('Sensitivity')+1),font='DarkGreen')
		elif s.startswith('resistante'):
			cell_format(annotationSheet.cell(row=l,column=aheader.index('Sensitivity')+1),font='DarkRed')

	l=l+1 # NEXT LINE

# SBT : ajout 'Amplicons <300X: '
if 'SBT' in project:
	annotationSheet.cell(row=l+2,column=1).value = "Amplicons < 300X: "
	cell_format(annotationSheet.cell(row=l+2,column=1))

#  __                __        ___  ___ ___ 
# /  ` |\ | \  /    /__` |__| |__  |__   |  
# \__, | \|  \/     .__/ |  | |___ |___  |  

if os.path.isfile('%s/_CNA/%s/CNV_finalReport.xlsx' % (run_folder,panel)):
	print " - [%s] CNV sheet ..." % time.strftime("%H:%M:%S")
	inBook = openpyxl.load_workbook('%s/_CNA/%s/CNV_finalReport.xlsx' % (run_folder,panel))
	inSheet = inBook['copy number analysis']

	for row_idx in range(1, inSheet.max_row+1):
		for col_idx in range(1, inSheet.max_column+1):
			read_cell = inSheet.cell(row = row_idx, column = col_idx)
			cnvSheet.cell(row=row_idx,column=col_idx).value = read_cell.value
			if read_cell.has_style:
				cnvSheet.cell(row=row_idx,column=col_idx).font = copy.copy(read_cell.font)
				cnvSheet.cell(row=row_idx,column=col_idx).border = copy.copy(read_cell.border)
				cnvSheet.cell(row=row_idx,column=col_idx).fill = copy.copy(read_cell.fill)
				cnvSheet.cell(row=row_idx,column=col_idx).alignment = copy.copy(read_cell.alignment)

	for row_idx in range(1, inSheet.max_row+1):
		c = cnvSheet.cell(row=row_idx,column=1)
		if c.value == sample:
			cell_format(c,color='Yellow')
			break
else:
	print " /!\ CNV finalReport file not found"
	del finalReport['CNV']
	# cnvSheet.cell(row=1,column=1).value = "CNV finalReport file not found for %s. " % sample

#  __        __  ___     __   __        ___  __        __   ___ 
# |__) |    /  \  |     /  ` /  \ \  / |__  |__)  /\  / _` |__  
# |    |___ \__/  |     \__, \__/  \/  |___ |  \ /~~\ \__> |___ 

sample_read_len_histo = '%s/_plotCoverage/read_len_histograms/%s_rawlib.read_len_histogram.png' % (run_folder,barcode)
amplicon_plots = sorted(glob.glob('%s/_plotCoverage/%s/*_S*_all_samples.png' % (run_folder, panel)))

z = 1
if os.path.isfile(sample_read_len_histo):
	z = 10
	plotSheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=18)
	plotSheet.cell(row=1,column=3).value = '%s READ LENGTH' % sample
	plotSheet.merge_cells(start_row=1, start_column=20, end_row=1, end_column=35)
	plotSheet.cell(row=1,column=20).value = 'RUN READ LENGTH'

	with PILimage.open(sample_read_len_histo) as img:
		img.thumbnail((9999,320))
		img.save(sample_read_len_histo.replace('.png','.thumb.png'))
	img = Image(sample_read_len_histo.replace('.png','.thumb.png'))
	plotSheet.add_image(img,'C2')

	run_read_len_histo = '%s/_plotCoverage/read_len_histograms/run.read_len_histogram.png' % run_folder
	with PILimage.open(run_read_len_histo) as img:
		img.thumbnail((9999,320))
		img.save(run_read_len_histo.replace('.png','.thumb.png'))
	img = Image(run_read_len_histo.replace('.png','.thumb.png'))
	plotSheet.add_image(img,'T2')

try:
	if amplicon_plots:
		print " - [%s] Plot Coverage sheet ..." % time.strftime("%H:%M:%S")
		# z = 10
		for ap in amplicon_plots:
			plotSheet.merge_cells(start_row=z, start_column=1, end_row=z, end_column=32)
			plotSheet.cell(row=z,column=1).value = os.path.basename(ap)
			cell_format(plotSheet.cell(row=z,column=1),font='bigBold',alignment='center',color='DarkGrey')
			z += 1
			with PILimage.open(ap) as img:
				img.thumbnail((9999,560)) # 600 / 40 = 15 lines
				img.save(ap.replace('.png','.thumb.png'))
			img = Image(ap.replace('.png','.thumb.png'))
			plotSheet.add_image(img,'B%s' % z)
			z += 14

		for c in range(1,38): # 40/40 pixel cube
			plotSheet.column_dimensions[plotSheet.cell(row=1,column=c).column].width = 5.71
		for r in range(1,plotSheet.max_row+15):# 40/40 pixel cube
			plotSheet.row_dimensions[plotSheet.cell(row=r,column=1).row].height = 30
			for cell in plotSheet["%s:%s" % (r,r)]:
				cell_format(cell,font='bigBold',alignment='center',color='DarkGrey')
	else:
		print " /!\ plotCoverage histograms not found, removing sheet"
		del finalReport['Plot Coverage']
		# plotSheet.cell(row=1,column=1).value = "plotCoverage not found for %s. " % sample
except:
	pass

#  __   __        ___  __        __   ___     __        ___  ___ ___ 
# /  ` /  \ \  / |__  |__)  /\  / _` |__     /__` |__| |__  |__   |  
# \__, \__/  \/  |___ |  \ /~~\ \__> |___    .__/ |  | |___ |___  |  

cov_file = False
amplicon_run = False
if os.path.isfile('%s/coverage/%s_%s.amplicon.cov.xls' % (intermediate_folder,sample,barcode)):
	cov_file = open('%s/coverage/%s_%s.amplicon.cov.xls' % (intermediate_folder,sample,barcode),'r')
	amplicon_run = True
elif os.path.isfile('%s/coverage/%s_%s.target.cov.xls' % (intermediate_folder,sample,barcode)):
	cov_file = open('%s/coverage/%s_%s.target.cov.xls' % (intermediate_folder,sample,barcode),'r')
elif os.path.isfile('%s.zip' % intermediate_folder):
	archive = zipfile.ZipFile('%s.zip' % intermediate_folder, 'r')
	if 'coverage/%s_%s.amplicon.cov.xls' % (sample,barcode) in archive.namelist():
		cov_file = archive.open('coverage/%s_%s.amplicon.cov.xls' % (sample,barcode))
		amplicon_run = True
	if 'coverage/%s_%s.target.cov.xls' % (sample,barcode) in archive.namelist():
		cov_file = archive.open('coverage/%s_%s.target.cov.xls' % (sample,barcode))

if cov_file:
	print " - [%s] Coverage sheet ..." % time.strftime("%H:%M:%S")
	cov_file_reader = csv.reader(cov_file,delimiter='\t')
	header = cov_file_reader.next()
	header[4] = 'gene_id'

	if amplicon_run:
		coverageSheet.title = 'Amplicon Coverage'
		header.append('300x_fwd_rev')
	for i in range(len(header)):
		coverageSheet.cell(row=1,column=i+1).value = header[i]
		cell_format(coverageSheet.cell(row=1,column=i+1),font='bold',border='thin',color='DarkGrey')

	cov_lines = [] # if from zip cannot seek(), so keep lines in list
	for cov_line in cov_file_reader:
		cov_lines.append(cov_line)

	l=2
	red_regions = []
	## HEMATO LAM : MALE / FEMALE (chrY)
	if project in ['LAM','LAM_2018','FLT3']:
		for cov_line in cov_lines:
			if cov_line[0] == 'chrY':
				# cov_line[4] = cov_line[4].split(';')[0].split('GENE_ID=')[-1]
				cov_line[9] = int(round(float(cov_line[9])))
				if cov_line[9] > 20:
					cov_line.append('MALE')
					for i in range(len(cov_line)):
						cell_val = representsInt(cov_line[i])
						coverageSheet.cell(row=l,column=i+1).value = cell_val
						cell_format(coverageSheet.cell(row=l,column=i+1),color='LightBlue')
				else:
					cov_line.append('FEMALE')
					for i in range(len(cov_line)):
						cell_val = representsInt(cov_line[i])
						coverageSheet.cell(row=l,column=i+1).value = cell_val
						cell_format(coverageSheet.cell(row=l,column=i+1),color='LightPink')
				cov_lines.remove(cov_line)
				l=l+1
				break

	region2geneex = {}
	for db_target_region in db_target_regions:
		region2geneex[db_target_region['targetedRegionName']] = '%s_%s' % (db_target_region['gene'],db_target_region['details'])

	sorted_reader = sorted(cov_lines, key=lambda x: float(x[9]))
	for cov_line in sorted_reader:
		# attributes = cov_line[4]
		# gene_id = attributes.split('GENE=')[-1].split(';')[0]
		# details = attributes.split('DETAILS=')[-1].split(';')[0]
		# cov_line[4] = '%s_%s' % (gene_id,details)
		cov_line[4] = region2geneex[cov_line[3]]
		if amplicon_run:
			# YES / NO FORWARD REVERSE 300X
			if float(cov_line[10]) >= minX and float(cov_line[11]) >= minX:
				cov_line.append('yes')
			else:
				cov_line.append('no')
		# TOTAL READS > minX GREEN, < minX RED
		cov_line[9] = int(round(float(cov_line[9])))
		if float(cov_line[9]) < minX:
			red_regions.append(cov_line[3])
			for i in range(len(cov_line)):
				cell_val = representsInt(cov_line[i])
				coverageSheet.cell(row=l,column=i+1).value = cell_val
				cell_format(coverageSheet.cell(row=l,column=i+1),color='LightRed')
				cell_format(coverageSheet.cell(row=l,column=10),font='bold')
		else:
			for i in range(len(cov_line)):
				cell_val = representsInt(cov_line[i])
				coverageSheet.cell(row=l,column=i+1).value = cell_val
				cell_format(coverageSheet.cell(row=l,column=i+1),color='LightGreen')
			if float(cov_line[10]) < minX:
				cell_format(coverageSheet.cell(row=l,column=11),color='LightRed')
				cell_format(coverageSheet.cell(row=l,column=16),color='LightRed')
			if float(cov_line[11]) < minX:
				cell_format(coverageSheet.cell(row=l,column=12),color='LightRed')
				cell_format(coverageSheet.cell(row=l,column=16),color='LightRed')
		if cov_line[3] in ['AMPL7156804406','AMPL7156804405','AMPL7159095228','AMPL7159376611','AMPL7155781411']:
			for i in range(len(cov_line)):
				cell_val = representsInt(cov_line[i])
				coverageSheet.cell(row=l,column=i+1).value = cell_val
				cell_format(coverageSheet.cell(row=l,column=i+1),font='BlueBold')
		l=l+1
	# else:
		# # target cov capture version here
		# print "capture"
	cov_file.close()
else:
	coverageSheet.cell(row=1,column=1).value = "Region Coverage file not found for %s. " % sample
	print " /!\ coverage file not found"

#  __        __   ___     __   __        ___  __        __   ___     __        ___  ___ ___ 
# |__)  /\  /__` |__     /  ` /  \ \  / |__  |__)  /\  / _` |__     /__` |__| |__  |__   |  
# |__) /~~\ .__/ |___    \__, \__/  \/  |___ |  \ /~~\ \__> |___    .__/ |  | |___ |___  |  

base_cov_file = False
if os.path.isfile('%s/tvc_de_novo/depth.txt' % intermediate_folder):
	base_cov_file = open('%s/tvc_de_novo/depth.txt' % intermediate_folder,'r')
elif os.path.isfile('%s/coverage/depth.txt' % intermediate_folder):
	base_cov_file = open('%s/coverage/depth.txt' % intermediate_folder,'r')
elif os.path.isfile('%s.zip' % intermediate_folder):
	archive = zipfile.ZipFile('%s.zip' % intermediate_folder, 'r')
	if 'tvc_de_novo/depth.txt' in archive.namelist():
		base_cov_file = archive.open('tvc_de_novo/depth.txt')
	elif 'coverage/depth.txt' in archive.namelist():
		base_cov_file = archive.open('coverage/depth.txt')

l = 1
if base_cov_file:
	print " - [%s] Base coverage sheet ..." % time.strftime("%H:%M:%S")
	base_cov_file_reader = csv.reader(base_cov_file,delimiter='\t')
	base_cov_lines = [] # if from zip cannot seek(), so keep lines in list
	for base_cov_line in base_cov_file_reader:
		base_cov_lines.append(base_cov_line)

	header = ['Chr','Start','End','Number of bases','Region','Gene','Exon','Hotspots']
	if project in ['Lymphome_B','Lymphome_T']:
		header.append('Cosmic occurences (haematopoietic_and_lymphoid_tissue)')

	xlist_full = [int(minX),2000,500,300,250,200,100]
	xlist_full = sorted(xlist_full,reverse=True)
	xlist = [int(minX)]
	for x in xlist_full:
		if x >= int(minX):
			continue
		xlist.append(x)
		# if len(xlist) >= 3:
			# break

	# CREATE DICT WITH ALL TARGETED BASES (0 depth if not in depth.txt file)
	targeted_regions = []
	chrom_list = ['chr%s' % i for i in range(1,23)+['X','Y']]
	targeted_base_depth = {}
	for chrom in chrom_list:
		targeted_base_depth[chrom] = {} # {chr:{position1:depth,position2:depth}}...

	z = 0
	for db_target_region in db_target_regions:
		targeted_regions.append([db_target_region['chromosome'],db_target_region['start'],db_target_region['stop'],db_target_region['targetedRegionName'],db_target_region['gene'],db_target_region['details']])
		for position in range(db_target_region['start']+1,db_target_region['stop']+1):
			if position not in targeted_base_depth[db_target_region['chromosome']]:
				targeted_base_depth[db_target_region['chromosome']][position] = 0
				z+=1

	# PARSE DEPTH.TXT FILE
	for base_cov_line in base_cov_lines:
		chrom = base_cov_line[0]
		if 'ABL1' in chrom:
			chrom = 'chr9'
		position = int(base_cov_line[1])
		depth = int(base_cov_line[2])
		targeted_base_depth[chrom][position] = depth

	for minx in xlist: # 300, 100, 20...
		region_minx =  []
		start = False
		ampl = False
		for chrom in chrom_list:
			for position in sorted(targeted_base_depth[chrom].keys()):
				depth = targeted_base_depth[chrom][position]
				if depth < minx:
					# TROUVER INFOS AMPLICONS
					for tg in targeted_regions:
						if (chrom == tg[0]) and (tg[1]+1<=position and tg[2]+1>=position):
							tgampl = tg[3]
							tggene = tg[4]
							tgdetails = tg[5]
							continue

					if not start:
						start = position
						ampl = tgampl
						gene = tggene
						details = tgdetails
						end = position
						continue

					if ampl and (ampl != tgampl):
						numOfBase = (end-start)+1
						region_minx.append([chrom,start,end,numOfBase,ampl,gene,details])
						start = position
					ampl = tgampl
					gene = tggene
					details = tgdetails
					end = position
					
				elif start :
					numOfBase = (end-start)+1
					region_minx.append([chrom,start,end,numOfBase,ampl,gene,details])
					start = False
		if start : # si start n'est pas False, s'est termine sur un amplicon < minX, il faut le rajouter!
			numOfBase = (end-start)+1
			region_minx.append([chrom,start,end,numOfBase,ampl,gene,details])

		# Recherche variants d'interet (hotspots pour LAM, actionnables pour SBT). Colonne "hotspots"
		for region in region_minx:
			db_variants = []
			relevant_variants = []
			if 'SBT' in project:
				db_cur.execute("SELECT * FROM VariantAnnotation INNER JOIN Variant ON Variant.variantID = VariantAnnotation.variant WHERE actionability is not Null")
				db_variants = db_cur.fetchall()
			elif 'LAM' in project:
				db_cur.execute("SELECT * FROM VariantAnnotation INNER JOIN Variant ON Variant.variantID = VariantAnnotation.variant WHERE highlight = 1")
				db_variants = db_cur.fetchall()
			for db_variant in db_variants:
				if (db_variant['genomicStart'] >= region[1] and db_variant['genomicStart'] <= region[2]) or (db_variant['genomicStop'] >= region[1] and db_variant['genomicStop'] <= region[2]):
					relevant_variants.append(db_variant['proteinDescription'])
			hotstring = ','.join(relevant_variants)
			region.append(hotstring)

		# panel lymphome : recherche cosmic haematopoietic_and_lymphoid_tissue. Colonne "Cosmic occurences (haematopoietic_and_lymphoid_tissue)"
		if project in ['Lymphome_B','Lymphome_T']:
			if len(region_minx) > 500: # si trop on ne cherche pas. plombe le temps d'execution
				for region in region_minx:
					region.append('n/a')
			else:
				for region in region_minx:
					region.append([])
					cosmicDB_cur.execute("SELECT cosmicID,haematopoietic_and_lymphoid_tissue FROM Cosmic WHERE chr='%s' AND start>=%s AND stop<=%s AND haematopoietic_and_lymphoid_tissue not NULL" % (region[0],region[1],region[2]))
					cosmicDB_data = cosmicDB_cur.fetchall()
					for c in cosmicDB_data:
						region[8].append((c['cosmicID'],c['haematopoietic_and_lymphoid_tissue']))

				for region in region_minx:
					cosm_list = sorted(region[8], key=lambda x: x[1], reverse=True)
					string_list = []
					for item in cosm_list:
						s = '%s(%s)' % (item[0],item[1])
						string_list.append(s)
					cosmstring = ','.join(string_list)
					region[8] = cosmstring

		# red / white
		red_line = []
		white_line = []
		for region in region_minx:
			if region[4] in red_regions:
				red_line.append(region)
			else:
				white_line.append(region)

		###### ECRITURE ################################
		baseCoverageSheet.cell(row=l,column=1).value = 'Bases with coverage < %sX' % minx
		baseCoverageSheet.merge_cells(start_row=l, start_column=1, end_row=l, end_column=8)
		cell_format(baseCoverageSheet.cell(row=l,column=1),font='bigBold',alignment='center',color='DarkGrey')
		l+=1
		for i in range(len(header)):
			baseCoverageSheet.cell(row=l,column=i+1).value = header[i]
			cell_format(baseCoverageSheet.cell(row=l,column=i+1),font='bold',color='LightGrey')
		l+=1
		for region in region_minx:
			if region[4] in red_regions:
				for i in range(len(region)):
					baseCoverageSheet.cell(row=l,column=i+1).value = region[i]
					cell_format(baseCoverageSheet.cell(row=l,column=i+1),color='LightRed')
			else:
				for i in range(len(region)):
					baseCoverageSheet.cell(row=l,column=i+1).value = region[i]
					cell_format(baseCoverageSheet.cell(row=l,column=i+1))
			if baseCoverageSheet.cell(row=l,column=8).value:
				cell_format(baseCoverageSheet.cell(row=l,column=8),color='DarkRed',alignment='wrap')
			l+=1
		if region_minx == []:
			for i in range(len(header)):
				baseCoverageSheet.cell(row=l,column=i+1).value = '-'
			l+=1
		l=l+1

	base_cov_file.close()

else:
	baseCoverageSheet.cell(row=1,column=1).value = "WARNING : Base Coverage file not found for %s (or error parsing the base coverage file). " % sample
	print " /!\ base coverage file not found."

#       __   ___ 
# \  / /  ` |__  
#  \/  \__, |    

print " - [%s] Copying VCFs ..." % time.strftime("%H:%M:%S")
vc_data = {'tvc_de_novo/TSVC_variants':'tvc_de_novo','tvc_only_hotspot/TSVC_variants':'tvc_only_hotspot','mutect2/%s.mutect2.filtered' % sample:'mutect2','varscan2/%s.varscan2.filtered' % sample:'varscan2','lofreq/%s.lofreq.filtered' % sample:'lofreq','vardict/%s.vardict' % sample:'vardict'} # 'deepvariant/%s.deepvariant' % sample:'deepvariant'
for vc in vc_data.keys():
	vcf_file = False
	if os.path.isfile('%s/%s.vcf' % (intermediate_folder,vc)):
		vcf_file = open('%s/%s.vcf' % (intermediate_folder,vc),'r')
	elif os.path.isfile('%s.zip' % intermediate_folder):
		archive = zipfile.ZipFile('%s.zip' % intermediate_folder, 'r')
		if '%s.vcf' % vc in archive.namelist():
			vcf_file = archive.open('%s.vcf' % vc)

	if vcf_file:
		vcfSheet = finalReport.create_sheet(title='VCF (%s)' % vc_data[vc])
		vcf_file_reader = csv.reader(vcf_file,delimiter='\t')

		l=1
		for vcf_line in vcf_file_reader:
			for i in range(len(vcf_line)):
				cell_val = representsInt(vcf_line[i])
				vcfSheet.cell(row=l,column=i+1).value = cell_val
				cell_format(vcfSheet.cell(row=l,column=i+1))
			l=l+1
		vcf_file.close()

## FEUILLE VCF ONLY HOTSPOT ##

# vcf_file = False
# if os.path.isfile('%s/tvc_only_hotspot/TSVC_variants.vcf' % intermediate_folder):
	# vcf_file = open('%s/tvc_only_hotspot/TSVC_variants.vcf' % intermediate_folder,'r')
# elif os.path.isfile('%s.zip' % intermediate_folder):
	# archive = zipfile.ZipFile('%s.zip' % intermediate_folder, 'r')
	# if 'tvc_only_hotspot/TSVC_variants.vcf' in archive.namelist():
		# vcf_file = archive.open('tvc_only_hotspot/TSVC_variants.vcf')	

# if vcf_file:
	# vcfSheet2 = finalReport.create_sheet(title='VCF (only hotspot)')
	# vcf_file_reader = csv.reader(vcf_file,delimiter='\t')
	# l=1
	# for vcf_line in vcf_file_reader:
		# for i in range(len(vcf_line)):
			# cell_val = representsInt(vcf_line[i])
			# vcfSheet2.cell(row=l,column=i+1).value = cell_val
			# cell_format(vcfSheet2.cell(row=l,column=i+1))
		# l=l+1
	# vcf_file.close()
#else:
	#print "WARNING : TSVC_variants.vcf file not found for %s. VCF sheet is skiped." % sample

#                __       ___ 
# |     /\  \ / /  \ |  |  |  
# |___ /~~\  |  \__/ \__/  |  
#                             

#print " - [%s] Layout ..." % time.strftime("%H:%M:%S")
#### AUTO-SIZE COLUMNS ####
sns = ['Annotation','Target Coverage','Amplicon Coverage','Base Coverage']
# if os.path.isfile('%s/tvc_only_hotspot/TSVC_variants.vcf' % intermediate_folder):
	# sns.append('VCF (only hotspot)')

for sn in sns:
	try:
		ws = finalReport[sn]
	except:
		continue
	maxsize = 20
	if sn == 'Base Coverage':
		maxsize = 120
	# elif sn == 'VCF (de novo)' or sn == 'VCF (only hotspot)':
		# maxsize = 12
	dims = {}
	for row in ws.rows:
		for cell in row:
			if cell.value:
				try:
					text_size = len(str(cell.value)) + 2
					dims[cell.column] = min(max(dims.get(cell.column, 0), text_size),maxsize)
				except:
					pass
	for col, value in dims.items():
		ws.column_dimensions[col].width = value

if 'Sensitivity' in aheader:
	annotationSheet.column_dimensions[openpyxl.utils.cell.get_column_letter(aheader.index('Sensitivity')+1)].width = 11
if 'Commentaire' in aheader:
	annotationSheet.column_dimensions[openpyxl.utils.cell.get_column_letter(aheader.index('Commentaire')+1)].width = 15
if 'Chr' in aheader:
	annotationSheet.column_dimensions[openpyxl.utils.cell.get_column_letter(aheader.index('Chr')+1)].width = 6
if 'Freq' in aheader:
	annotationSheet.column_dimensions[openpyxl.utils.cell.get_column_letter(aheader.index('Freq')+1)].width = 9
if 'Var.Cov.' in aheader:
	annotationSheet.column_dimensions[openpyxl.utils.cell.get_column_letter(aheader.index('Var.Cov.')+1)].width = 8
if 'Depth' in aheader:
	annotationSheet.column_dimensions[openpyxl.utils.cell.get_column_letter(aheader.index('Depth')+1)].width = 8
if 'c.p.f.' in aheader:
	annotationSheet.column_dimensions[openpyxl.utils.cell.get_column_letter(aheader.index('c.p.f.')+1)].width = 30
if 'c.p.' in aheader:
	annotationSheet.column_dimensions[openpyxl.utils.cell.get_column_letter(aheader.index('c.p.')+1)].width = 30
if 'cosm.rs' in aheader:
	annotationSheet.column_dimensions[openpyxl.utils.cell.get_column_letter(aheader.index('cosm.rs')+1)].width = 30
if 'COSMIC' in aheader:
	annotationSheet.column_dimensions[openpyxl.utils.cell.get_column_letter(aheader.index('COSMIC')+1)].width = 13
if 'Sensitivity' in aheader:
	annotationSheet.column_dimensions[openpyxl.utils.cell.get_column_letter(aheader.index('Sensitivity')+1)].width = 12
if 'Position' in aheader:
	annotationSheet.column_dimensions[openpyxl.utils.cell.get_column_letter(aheader.index('Position')+1)].width = 10
if 'Ref' in aheader:
	annotationSheet.column_dimensions[openpyxl.utils.cell.get_column_letter(aheader.index('Ref')+1)].width = 8
if 'Alt' in aheader:
	annotationSheet.column_dimensions[openpyxl.utils.cell.get_column_letter(aheader.index('Alt')+1)].width = 8
if 'dbSNP' in aheader:
	annotationSheet.column_dimensions[openpyxl.utils.cell.get_column_letter(aheader.index('dbSNP')+1)].width = 11
if 'InterVar' in aheader:
	annotationSheet.column_dimensions[openpyxl.utils.cell.get_column_letter(aheader.index('InterVar')+1)].width = 13
if 'ClinVar' in aheader:
	annotationSheet.column_dimensions[openpyxl.utils.cell.get_column_letter(aheader.index('ClinVar')+1)].width = 13
if 'SIFT' in aheader:
	annotationSheet.column_dimensions[openpyxl.utils.cell.get_column_letter(aheader.index('SIFT')+1)].width = 12
if 'POLYPHEN2' in aheader:
	annotationSheet.column_dimensions[openpyxl.utils.cell.get_column_letter(aheader.index('POLYPHEN2')+1)].width = 12
if 'PROVEAN' in aheader:
	annotationSheet.column_dimensions[openpyxl.utils.cell.get_column_letter(aheader.index('PROVEAN')+1)].width = 12		

if 'CNV' in finalReport.sheetnames:
	ws = finalReport['CNV']
	ws.column_dimensions['A'].width = 20
	for col_name in ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG']:
		ws.column_dimensions[col_name].width = 12

#         __  ___                    __        __     __      __   __   __     __  ___  __  
# \  / | |__)  |  |  |  /\  |       |__)  /\  /__` | /  `    /__` /  ` |__) | |__)  |  /__` 
#  \/  | |  \  |  \__/ /~~\ |___    |__) /~~\ .__/ | \__,    .__/ \__, |  \ | |     |  .__/ 
#                                                                                           

print " - [%s] VBS scripts ..." % time.strftime("%H:%M:%S")
try:
	alamut_variants_vbscript(sample_folder,vb_variant_list)
	if os.path.exists(bam_path.replace('.bam','.processed.bam')):
		alamut_bam_vbscript(sample_folder,sample,barcode,processed=True)
	else:
		alamut_bam_vbscript(sample_folder,sample,barcode,processed=False)
	print_vbscript(sample_folder,sample,barcode,project)
except:
	print "\t - warning : alamut vbscript creation FAILED "


###### SAVE WORKBOOK ######
output = '%s/%s_%s.finalReport.xlsx' % (sample_folder,sample.split('/')[-1],barcode)
finalReport.save(output)

