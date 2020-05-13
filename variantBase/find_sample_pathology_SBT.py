#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import sys
import sqlite3
import openpyxl

def dict_factory(cursor, row):
    d = {}
    for idx, col in enumerate(cursor.description):
        d[col[0]] = row[idx]
    return d

#####################################################################################
# SBT : PATHOLOGIE (poumon/colon/mela) info

bilan_mut_paths = {
	'bilan_mut_path_2015' : '/media/n06lbth/sauvegardes_pgm/SBT/Bilans_archive/Bilan_mutation_2015.xlsx',
	'bilan_mut_path_2016' : '/media/n06lbth/sauvegardes_pgm/SBT/Bilans_archive/Bilan_mutation_2016.xlsx',
	'bilan_mut_path_2017' : '/media/n06lbth/sauvegardes_pgm/SBT/Bilans_archive/Bilan_mutation_2017.xlsx',
	'bilan_mut_path_2018' : '/media/n06lbth/sauvegardes_pgm/SBT/Bilans_archive/Bilan_mutation_2018.xlsx',
	'bilan_mut_path_2019' : '/media/n06lbth/sauvegardes_pgm/SBT/Bilan_mutation_2019.xlsx',
	'bilan_mut_path_2020' : '/media/n06lbth/sauvegardes_pgm/SBT/Bilan_mutation_2020.xlsx'
	}

pipeline_folder = os.environ['NGS_PIPELINE_BX_DIR']
db_con = sqlite3.connect('%s/variantBase/VariantBase.db' % pipeline_folder)
db_con.row_factory = dict_factory
db_cur = db_con.cursor()
#####################################################################################

# PARSE ALL BILAN MUTATION
print "- parsing bilan mutation..."
sample2patho = {}
for bm in bilan_mut_paths.values():
	bilan_mut = openpyxl.load_workbook(bm)
	sheetnames = bilan_mut.sheetnames
	sheet2path = {sheetnames[0]:'poumon',sheetnames[1]:'colon',sheetnames[2]:'melanome'}  # poumon > colon > mela
	for i in range(3):
		sheet = bilan_mut[sheetnames[i]]
		for nrow in range(sheet.max_row):
			sample2patho[sheet.cell(row=nrow+1,column=4).value] = sheet2path[sheetnames[i]]

# FIND AND INSERT PATHOLOGY INTO DB
print "- updating db..."
SBT_samples = []
db_cur.execute("SELECT sampleID FROM SAMPLE INNER JOIN Analysis ON Analysis.sample = Sample.sampleID INNER JOIN Panel ON Panel.panelID = Analysis.panel WHERE panelProject = '%s'" % 'SBT')
db_samples = db_cur.fetchall()
for db_sample in db_samples:
	SBT_samples.append(db_sample['sampleID'])
	
print "\t - %s SBT samples" % len(SBT_samples)
for sampleID in sample2patho.keys():
	if sampleID in SBT_samples:
		db_cur.execute("UPDATE Sample SET pathology = '%s' WHERE sampleID='%s'" % (sample2patho[sampleID],sampleID))
		print "\t\t - %s -> %s" % (sampleID,sample2patho[sampleID])
	else:
		try:
			print "\t\t - %s not found in bilan mutation" % sampleID
		except Exception as e:
			print str(e)
	
db_con.commit()
db_con.close()
