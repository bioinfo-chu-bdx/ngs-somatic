#!/usr/bin/python
import sys
import re
import json
import xlrd
import glob
import os
import openpyxl

# THIS SCRIPT UPDATE THE COMPLETE VariantList_ALL.json FROM RUN LIST OR SINGLE RUN FOLDER IN ARGV.
# VARIANTLIST JSON CONTAINS RunName, Sample, Variant Data (chrm, start, stop, ref, alt, varcov, poscov). EXEMPLE :


#"Auto_user_PGM-165-Run98_35pM_Chef_SBT_colon_lung_v4_318v2_234": {
	#"SAMPLE-AF454G": [
		#["chr4", 55599436, 55599436, "T", "C", 405, 1245],
		#["chr7", 55599438, 55599438, "G", "C", 120, 1040],
#	]
                
# USAGE : python collect_run_variants_to_variantlist.py
#    OR : python collect_run_variants_to_variantlist.py /path/to/run_folder

pipeline_folder = os.environ['NGS_PIPELINE_BX_DIR']
variant_list_path = '%s/variantBase/variantList/variantList_ALL.json' % pipeline_folder
run_list_path = '%s/variantBase/runList/runList_ALL.fullpath.txt' % pipeline_folder

variantlist = {}
if os.path.exists(variant_list_path):
	with open(variant_list_path, 'r') as g:
		variantlist = json.load(g)

run2write_ordered = []
if len(sys.argv)>1:
	run_list = [sys.argv[1]]
else:
	run_list = []
	rl = open(run_list_path,'r')
	for run in rl:
		run_folder = run.replace('\n','')
		run_list.append(run_folder)
	
for run_folder in run_list:
	# RUN TABLE
	if run_folder.endswith('/'):
		run = run_folder.split('/')[-2]
	else:
		run = run_folder.split('/')[-1]
		
	run2write_ordered.append(run)
	barcodes_json = False
	with open(run_folder+'/barcodes.json', 'r') as g:
		barcodes_json = json.load(g)
	if run not in variantlist:
		variantlist[run] = {}
	else:
		print "*already in variantlist* %s" % run
		continue
	print "- collecting %s ..." % run
	for barcode in barcodes_json:
		sample = barcodes_json[barcode]['sample']
		dna_number = barcodes_json[barcode]['sample_id']
		#if dna_number == 'CONTROL':
			#continue
		if sample not in variantlist[run]:
			variantlist[run][sample] = []
		
		finalreport_paths = glob.glob('%s/%s/*%s_%s*finalReport*'%(run_folder,sample,sample,barcode))
		if finalreport_paths:
			for fp_path in finalreport_paths:
				if '~$' in fp_path: # fichier temporaire
					continue
				if fp_path.endswith('.xls'):
					#xlrd
					fp = xlrd.open_workbook(fp_path)
					anno_sheet = fp.sheet_by_index(0)
					for j in range(anno_sheet.ncols):
						if anno_sheet.cell_value(0,j) in ['Chr','chr','Chromosome','chromosome','chrom','Chrom']:
							chromosome_index = j
						elif anno_sheet.cell_value(0,j) in ['Start_Position','Position','Start.Pos','Start.Pos.','Start','start','Position','Pos.']:
							start_index = j
						elif anno_sheet.cell_value(0,j) in ['Ref.seq','Ref.Seq','Ref.seq.','Ref.Seq.','Ref','ref']:
							ref_index = j
						elif anno_sheet.cell_value(0,j) in ['Var.seq','Var.Seq','Alt','Var.seq.','Var.Seq.','alt']:
							alt_index = j
						elif anno_sheet.cell_value(0,j) in ['Var.Cov.','var.cov.']:
							varcov_index = j
						elif anno_sheet.cell_value(0,j) in ['Pos.Cov.','Depth']:
							poscov_index = j
					### PARSE XLS
					for i in range(1,anno_sheet.nrows):
						chrm = anno_sheet.cell_value(i,chromosome_index)
						ref = anno_sheet.cell_value(i,ref_index)
						alt = anno_sheet.cell_value(i,alt_index)
						if chrm and ref and alt :
							start = int(anno_sheet.cell_value(i,start_index))
							varcov = int(anno_sheet.cell_value(i,varcov_index))
							poscov = int(anno_sheet.cell_value(i,poscov_index))
							if ref == '-':
								stop = start + 1
							elif alt == '-':
								if len(ref) > 1:
									stop = start+(len(ref)-1)
								else:
									stop = start
							elif len(ref) > 1 or len(alt) > 1:
								if len(ref) > 1:
									stop = start+(len(ref)-1)
								else:
									stop = start
							else:
								stop = start
							variant = [str(chrm),start,stop,str(ref),str(alt),varcov,poscov]
							if variant not in variantlist[run][sample]: 
								variantlist[run][sample].append(variant)
				elif fp_path.endswith('.xlsx'):
					#openpyxl
					fp = openpyxl.load_workbook(fp_path)
					anno_sheetname = fp.sheetnames[0]
					anno_sheet = fp[anno_sheetname]
					for ncol in range(anno_sheet.max_column):
						if anno_sheet.cell(row=1,column=ncol+1).value in ['Chr','chr','Chromosome','chromosome','chrom','Chrom']:
							chromosome_index = ncol+1
						elif anno_sheet.cell(row=1,column=ncol+1).value in ['Start_Position','Position','Start.Pos','Start.Pos.','Start','start','Position','Pos.']:
							start_index = ncol+1
						elif anno_sheet.cell(row=1,column=ncol+1).value in ['Ref.seq','Ref.Seq','Ref.seq.','Ref.Seq.','Ref','ref']:
							ref_index = ncol+1
						elif anno_sheet.cell(row=1,column=ncol+1).value in ['Var.seq','Var.Seq','Alt','Var.seq.','Var.Seq.','alt']:
							alt_index = ncol+1
						if anno_sheet.cell(row=1,column=ncol+1).value in ['Var.Cov.','var.cov.']:
							varcov_index = ncol+1
						if anno_sheet.cell(row=1,column=ncol+1).value in ['Pos.Cov.','Depth']:
							poscov_index = ncol+1
					### PARSE XLSX
					for nrow in range(2,anno_sheet.max_row+1):
						chrm = anno_sheet.cell(row=nrow,column=chromosome_index).value
						ref = anno_sheet.cell(row=nrow,column=ref_index).value
						alt = anno_sheet.cell(row=nrow,column=alt_index).value
						if chrm and ref and alt :
							start = int(anno_sheet.cell(row=nrow,column=start_index).value)
							varcov = int(anno_sheet.cell(row=nrow,column=varcov_index).value)
							poscov = int(anno_sheet.cell(row=nrow,column=poscov_index).value)
							if ref == '-': 
								stop = start + 1
							elif alt == '-':
								if len(ref) > 1:
									stop = start+(len(ref)-1)
								else:
									stop = start
							elif len(ref) > 1 or len(alt) > 1:
								if len(ref) > 1:
									stop = start+(len(ref)-1)
								else:
									stop = start
							else:
								stop = start
							variant = [str(chrm),start,stop,str(ref),str(alt),varcov,poscov]
							if variant not in variantlist[run][sample]:
								variantlist[run][sample].append(variant)				
				else:
					print "**WARNING (FINALREPORT FILE EXTENSION weird )** %s" % fp_path
			#print "\t- %s : %s variants" % (sample,len(variantlist[run][sample]))
		else:
			print "**WARNING (NO FINALREPORT found for SAMPLE )** %s" % sample
                                                                                                                                                                                                                        
print "- WRITING VARIANTLIST JSON..."
# routine d'ecriture differente de json dumps indent qui prend trop de lignes
with open(variant_list_path,'w') as vljson:
	vljson.write('{\n')
	for i in range(len(run2write_ordered)) :
		run = run2write_ordered[i]
		vljson.write('\t"%s": {\n' % run)
		for j in range(len(variantlist[run].keys())):
			sample = variantlist[run].keys()[j]
			vljson.write('\t\t"%s": [\n' % (sample))
			for k in range(len(variantlist[run][sample])):
				variant = str(variantlist[run][sample][k]).replace('\'','"').replace('u"','"')
				if k == (len(variantlist[run][sample])-1):
					vljson.write('\t\t\t%s\n' % variant)
				else:
					vljson.write('\t\t\t%s,\n' % variant)
			if j == (len(variantlist[run].keys())-1):
				vljson.write('\t\t]\n')
			else:
				vljson.write('\t\t],\n')
		if i == (len(run2write_ordered)-1):
			vljson.write('\t}\n')
		else:
			vljson.write('\t},\n')
	vljson.write('}\n')
