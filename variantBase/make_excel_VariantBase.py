#!/usr/bin/env python
# coding: utf-8
import os
import sys
import json
import sqlite3
import openpyxl
from optparse import OptionParser

def dict_factory(cursor, row):
	d = {}
	for idx, col in enumerate(cursor.description):
		d[col[0]] = row[idx]
	return d

def cell_format(cell, alignment='left', style=None, font=None):
	if alignment == 'center':
		cell.alignment = openpyxl.styles.Alignment(horizontal='center')
	elif alignment == 'left':
		cell.alignment = openpyxl.styles.Alignment(horizontal='left')
	if style == 'Grey':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='D9D9D9')
		cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='hair'))
	elif style == 'LightGrey':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='F2F2F2')
		cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='hair'))
	elif style == 'Purple':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='CCC0D0')
		cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='hair'))
	elif style == 'LightPurple':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='E4DFEC')
		cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='hair'))
	elif style == 'Blue':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='B7DEE8')
		cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='hair'))
	elif style == 'LightBlue':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='DAEEF3')
		cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='hair'))
	elif style == 'Green':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='D8E4BC')
		cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='hair'))
	elif style == 'LightGreen':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='EBF1DE')
		cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='hair'))
	elif style == 'DarkGreen':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='C4D79B')
	elif style == 'LightRed':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='F2DCDB')
	elif style == 'TextBlue':
		cell.font = openpyxl.styles.Font(color='004c99')
	else:
		cell.fill = openpyxl.styles.PatternFill(fill_type=None,start_color='FFFFFF')
	if font == 'Bold':
		cell.font = openpyxl.styles.Font(bold=True)

#####################################################################################################	
parser = OptionParser()
parser.add_option('-p', '--project', help='one or more project comma separated. ex:(LAM,FLT3)', dest='project',default=False)
parser.add_option('-x', '--panel', help='one or more panel comma separated. ex:(LAM-illumina-v1,LAM-illumina-v2)', dest='panel',default=False)
parser.add_option('-o', '--output', help='output folder', dest='output')
(options, args) = parser.parse_args()

pipeline_folder = os.environ['NGS_PIPELINE_BX_DIR']
db_con = sqlite3.connect('%s/variantBase/VariantBase.db' % pipeline_folder)
db_con.row_factory = dict_factory
db_cur = db_con.cursor()

if options.project and options.pane:
	sys.stderr.write("Error: choose either <--project> or <--panel>, not both\n")
	sys.exit()

vbformat = 'Classic'
if options.project:
	projects = options.project.split(',')
	in_projects = str(projects).replace('[','(').replace(']',')')
	if options.project == 'SBT': # Si project SBT uniquement; format special avec colon lung mela colums.
		vbformat = 'ColonLungMela'
elif options.panel:
	panels = options.panel.split(',')
	in_panels = str(panels).replace('[','(').replace(']',')')
	db_cur.execute("SELECT DISTINCT panelProject FROM Panel WHERE panelID in %s" % in_panels)
	db_panels = db_cur.fetchall()
	for db_panel in db_panels:
		if db_panel['panelProject'] == 'SBT' and len(db_panels) == 1:
			vbformat = 'ColonLungMela'
			break

if vbformat == 'ColonLungMela':
	excelVB = openpyxl.load_workbook('%s/variantBase/excel_files/VariantBase_EMPTY_SBT.xlsx' % pipeline_folder)
else:
	excelVB = openpyxl.load_workbook('%s/variantBase/excel_files/VariantBase_EMPTY.xlsx' % pipeline_folder)
normal = excelVB._named_styles['Normal']
normal.font.name = 'Calibri'
normal.font.size = 11

output = options.output

#####################################################################################################

## DATA SHEET
print "- filling data sheet..."
dataSheet = excelVB['Data']
if options.project:
	db_cur.execute("""SELECT DISTINCT runID FROM Run 
	INNER JOIN Analysis ON Analysis.run = Run.runID 
	INNER JOIN Panel ON Panel.panelID = Analysis.panel 
	WHERE panelProject in %s ORDER BY runDate""" % in_projects)
elif options.panel:
	db_cur.execute("""SELECT DISTINCT runID FROM Run 
	INNER JOIN Analysis ON Analysis.run = Run.runID 
	INNER JOIN Panel ON Panel.panelID = Analysis.panel 
	WHERE panelID in %s ORDER BY runDate""" % in_panels)
db_runs = db_cur.fetchall()
runs = []

for db_run in db_runs:
	dataSheet_newrow = dataSheet.max_row+1
	dataSheet.cell(row=dataSheet_newrow,column=1).value = db_run['runID']
	runs.append(db_run['runID'])

	if options.project :
		db_cur.execute("""SELECT sampleName, sampleID, pathology FROM Sample 
		INNER JOIN Analysis ON Analysis.sample = Sample.sampleID INNER JOIN Panel ON Panel.panelID = Analysis.panel 
		INNER JOIN Run ON Run.runID = Analysis.run 
		WHERE runID = '%s' AND panelProject in %s AND isControl = 0""" % (db_run['runID'],in_projects))
	elif options.panel:
		db_cur.execute("""SELECT sampleName, sampleID, pathology FROM Sample 
		INNER JOIN Analysis ON Analysis.sample = Sample.sampleID 
		INNER JOIN Panel ON Panel.panelID = Analysis.panel 
		INNER JOIN Run ON Run.runID = Analysis.run 
		WHERE runID = '%s' AND panelID in %s AND isControl = 0""" % (db_run['runID'],in_panels))
	db_samples = db_cur.fetchall()
	if vbformat == 'ColonLungMela':
		lsamples = []
		csamples = []
		msamples = []
		osamples = []
		for db_sample in db_samples:
			if db_sample['pathology'] == 'poumon':
				lsamples.append('%s-%s' % (db_sample['sampleName'],db_sample['sampleID']))
			elif db_sample['pathology'] == 'colon':
				csamples.append('%s-%s' % (db_sample['sampleName'],db_sample['sampleID']))
			elif db_sample['pathology'] == 'melanome':
				msamples.append('%s-%s' % (db_sample['sampleName'],db_sample['sampleID']))
			else:
				osamples.append('%s-%s' % (db_sample['sampleName'],db_sample['sampleID']))
		dataSheet.cell(row=dataSheet_newrow,column=7).value = len(lsamples) + len(csamples) + len(msamples) + len(osamples)
		dataSheet.cell(row=dataSheet_newrow,column=9).value = ','.join(lsamples)
		dataSheet.cell(row=dataSheet_newrow,column=12).value = ','.join(csamples)
		dataSheet.cell(row=dataSheet_newrow,column=15).value = ','.join(msamples)
		dataSheet.cell(row=dataSheet_newrow,column=18).value = ','.join(osamples)
		dataSheet['C2'].value += (len(lsamples) + len(csamples) + len(msamples) + len(osamples))
		dataSheet['D2'].value += len(lsamples)
		dataSheet['E2'].value += len(csamples)
		dataSheet['F2'].value += len(msamples)

		dataSheet.merge_cells(start_row=dataSheet_newrow,start_column=1,end_row=dataSheet_newrow,end_column=6)
		dataSheet.merge_cells(start_row=dataSheet_newrow,start_column=7,end_row=dataSheet_newrow,end_column=8)
		dataSheet.merge_cells(start_row=dataSheet_newrow,start_column=9,end_row=dataSheet_newrow,end_column=11)
		dataSheet.merge_cells(start_row=dataSheet_newrow,start_column=12,end_row=dataSheet_newrow,end_column=14)
		dataSheet.merge_cells(start_row=dataSheet_newrow,start_column=15,end_row=dataSheet_newrow,end_column=17)
		dataSheet.merge_cells(start_row=dataSheet_newrow,start_column=18,end_row=dataSheet_newrow,end_column=20)
		cell_format(dataSheet.cell(row=dataSheet_newrow,column=1),alignment='left')
		cell_format(dataSheet.cell(row=dataSheet_newrow,column=7),alignment='center')
		cell_format(dataSheet.cell(row=dataSheet_newrow,column=9),alignment='left')
		cell_format(dataSheet.cell(row=dataSheet_newrow,column=12),alignment='left')
		cell_format(dataSheet.cell(row=dataSheet_newrow,column=15),alignment='left')
		cell_format(dataSheet.cell(row=dataSheet_newrow,column=18),alignment='left')
	else:
		samples = []
		for db_sample in db_samples:
			samples.append('%s-%s' % (db_sample['sampleName'],db_sample['sampleID']))
		dataSheet.cell(row=dataSheet_newrow,column=7).value = len(samples)
		dataSheet.cell(row=dataSheet_newrow,column=9).value = ','.join(samples)
		dataSheet['C2'].value += len(samples)

		dataSheet.merge_cells(start_row=dataSheet_newrow,start_column=1,end_row=dataSheet_newrow,end_column=6)
		dataSheet.merge_cells(start_row=dataSheet_newrow,start_column=7,end_row=dataSheet_newrow,end_column=8)
		dataSheet.merge_cells(start_row=dataSheet_newrow,start_column=9,end_row=dataSheet_newrow,end_column=17)
		dataSheet.merge_cells(start_row=dataSheet_newrow,start_column=18,end_row=dataSheet_newrow,end_column=20)
		cell_format(dataSheet.cell(row=dataSheet_newrow,column=1),alignment='left')
		cell_format(dataSheet.cell(row=dataSheet_newrow,column=7),alignment='center')
		cell_format(dataSheet.cell(row=dataSheet_newrow,column=9),alignment='left')
		cell_format(dataSheet.cell(row=dataSheet_newrow,column=18),alignment='left')

dataSheet['A2'].value = len(runs)

## MAKE ALL GENE SHEET
print "- making gene sheets..."
if options.project:
	db_cur.execute("""SELECT DISTINCT gene,transcriptID FROM Transcript 
	INNER JOIN TargetedRegion ON TargetedRegion.transcript = Transcript.transcriptID 
	INNER JOIN Panel ON Panel.panelID = TargetedRegion.panel 
	WHERE panelProject in %s ORDER BY gene""" % in_projects)
elif options.panel:
	db_cur.execute("""SELECT DISTINCT gene,transcriptID FROM Transcript 
	INNER JOIN TargetedRegion ON TargetedRegion.transcript = Transcript.transcriptID 
	INNER JOIN Panel ON Panel.panelID = TargetedRegion.panel 
	WHERE panelID in %s ORDER BY gene""" % in_panels)
db_transcripts = db_cur.fetchall()
gene2transcript = {}
geneTemplateSheet = excelVB['Gene']
for db_transcript in db_transcripts:
	gene = db_transcript['gene']
	transcript = db_transcript['transcriptID']
	if gene not in gene2transcript.keys():
		gene2transcript[gene] = transcript
		newGeneSheet = excelVB.copy_worksheet(geneTemplateSheet)
		newGeneSheet.title = db_transcript['gene']
	else:
		print "WARNING : duplicate gene-transcript %s:%s : using transcript %s" % (gene,transcript,gene2transcript[gene])
del excelVB['Gene']

genes = gene2transcript.keys()
transcripts = gene2transcript.values()

# PARSE ALL VARIANTS
print "- parsing and formatting all db variants..."
header = [
		'Commentaire',
		'Gene',
		'Transcript',
		'Chr',
		'Exon',
		'Intron',
		'Position',
		'Ref',
		'Alt',
		'c.',
		'p.',
		'Region',
		'Consequence',
		'Count',
		'Patients',
		'Freq.Mean',
		'Freq.Range',
		'Freqs',
		'Cov.Mean',
		'Cov.Range',
		'Depth.Mean',
		'Depth.Range',
		'InterVar',
		'ClinVar',
		'COSMIC',
		'dbSNP',
		'gnomAD',
		'1000G_ALL',
		'1000G_EUR',
		'NCI60',
		'ESP',
		'ExAC',
		'SIFT',
		'POLYPHEN2',
		'PROVEAN',
		'PubMed',
		'VEP_Consequence',
		'VEP_Impact',
		'VEP_Diff',
		'Class.',
		'c.(annovar)',
		'p.(annovar)',
		'annoWarning'
		]

if vbformat == 'ColonLungMela':
	header[header.index('Count')] = 'Lung Count'
	header.insert(header.index('Lung Count')+1,'Colon Count')
	header.insert(header.index('Colon Count')+1,'Mela Count')
	header.insert(header.index('Mela Count')+1,'Other Count')

if options.project:
	db_cur.execute("""SELECT DISTINCT VariantAnnotation.*,Variant.*,gene,transcript, variantReadDepth, positionReadDepth, sampleName, sampleID, pathology, panelID 
	FROM VariantAnnotation 
	INNER JOIN Variant ON Variant.variantID = VariantAnnotation.variant 
	INNER JOIN VariantMetrics ON VariantMetrics.variant = Variant.variantID 
	INNER JOIN Analysis ON Analysis.analysisID = VariantMetrics.analysis 
	INNER JOIN Sample ON Sample.sampleID = Analysis.sample 
	INNER JOIN Run ON Run.runID = Analysis.run 
	INNER JOIN Panel ON Panel.panelID = Analysis.panel 
	INNER JOIN Transcript ON Transcript.transcriptID = VariantAnnotation.transcript 
	WHERE panelProject in %s AND isControl = 0 ORDER BY genomicStart, variantID""" % in_projects)
elif options.panel:
	db_cur.execute("""SELECT DISTINCT VariantAnnotation.*,Variant.*,gene,transcript, variantReadDepth, positionReadDepth, sampleName, sampleID, pathology, panelID 
	FROM VariantAnnotation 
	INNER JOIN Variant ON Variant.variantID = VariantAnnotation.variant 
	INNER JOIN VariantMetrics ON VariantMetrics.variant = Variant.variantID 
	INNER JOIN Analysis ON Analysis.analysisID = VariantMetrics.analysis 
	INNER JOIN Sample ON Sample.sampleID = Analysis.sample 
	INNER JOIN Run ON Run.runID = Analysis.run 
	INNER JOIN Panel ON Panel.panelID = Analysis.panel 
	INNER JOIN Transcript ON Transcript.transcriptID = VariantAnnotation.transcript 
	WHERE panelID in %s AND isControl = 0 ORDER BY genomicStart, variantID""" % in_panels)
db_variants = db_cur.fetchall()
## note : RUN dans la requete permet de laisser contrainte que run existe bien. il y a des TESTILLUMINA truc qui se baladent... 

actualVariant = None
for db_variant in db_variants:
	if db_variant['transcript'] not in transcripts:
		continue
	if db_variant['variantID'] != actualVariant:
		if actualVariant != None:
			if vbformat == 'ColonLungMela':
				sheet.cell(row=l,column=header.index('Lung Count')+1).value = lcount
				cell_format(sheet.cell(row=l,column=header.index('Lung Count')+1),alignment='center',style='Purple',font='Bold')
				sheet.cell(row=l,column=header.index('Colon Count')+1).value = ccount
				cell_format(sheet.cell(row=l,column=header.index('Colon Count')+1),alignment='center',style='Blue',font='Bold')
				sheet.cell(row=l,column=header.index('Mela Count')+1).value = mcount
				cell_format(sheet.cell(row=l,column=header.index('Mela Count')+1),alignment='center',style='Green',font='Bold')
				sheet.cell(row=l,column=header.index('Other Count')+1).value = ocount
				cell_format(sheet.cell(row=l,column=header.index('Other Count')+1),alignment='center',style='Grey')
			else:
				sheet.cell(row=l,column=header.index('Count')+1).value = len(variantSamples)
				cell_format(sheet.cell(row=l,column=header.index('Count')+1),alignment='center',style='Purple',font='Bold')
			sheet.cell(row=l,column=header.index('Patients')+1).value = ','.join(variantSamples)
			cell_format(sheet.cell(row=l,column=header.index('Patients')+1),alignment='left',style='LightPurple')
			sheet.cell(row=l,column=header.index('Freq.Mean')+1).value = sum(variantFreqs)/len(variantFreqs)
			cell_format(sheet.cell(row=l,column=header.index('Freq.Mean')+1),alignment='center',style='Grey') #,format='Percent'
			sheet.cell(row=l,column=header.index('Freq.Range')+1).value = '(%s-%s)' % (min(variantFreqs),max(variantFreqs))
			cell_format(sheet.cell(row=l,column=header.index('Freq.Range')+1),alignment='center',style='LightGrey')
			sheet.cell(row=l,column=header.index('Freqs')+1).value = str(variantFreqs)
			cell_format(sheet.cell(row=l,column=header.index('Freqs')+1),alignment='left')
			sheet.cell(row=l,column=header.index('Cov.Mean')+1).value = sum(variantCovs)/len(variantCovs)
			cell_format(sheet.cell(row=l,column=header.index('Cov.Mean')+1),alignment='center',style='Grey')
			sheet.cell(row=l,column=header.index('Cov.Range')+1).value = '(%s-%s)' % (min(variantCovs),max(variantCovs))
			cell_format(sheet.cell(row=l,column=header.index('Cov.Range')+1),alignment='center',style='LightGrey')
			sheet.cell(row=l,column=header.index('Depth.Mean')+1).value = sum(variantDepths)/len(variantDepths)
			cell_format(sheet.cell(row=l,column=header.index('Depth.Mean')+1),alignment='center',style='Grey')
			sheet.cell(row=l,column=header.index('Depth.Range')+1).value = '(%s-%s)' % (min(variantDepths),max(variantDepths))
			cell_format(sheet.cell(row=l,column=header.index('Depth.Range')+1),alignment='center',style='LightGrey')

		actualVariant = db_variant['variantID']
		variantCovs = []
		variantDepths = []
		variantFreqs = []
		variantSamples = []
		dataSheet['B2'].value += 1
		if vbformat == 'ColonLungMela':
			lcount = 0
			ccount = 0
			mcount = 0
			ocount = 0

		gene = db_variant['gene']
		sheet = excelVB[gene]
		l = sheet.max_row+1

		comm = db_variant['commentaire']
		db_cur.execute("SELECT userComment FROM UserComment WHERE variantAnnotation='%s' AND panel='%s'" % (db_variant['variantAnnotationID'],db_variant['panelID']))
		db_userComments = db_cur.fetchall()
		for db_userComment in db_userComments:
			userComment = db_userComment['userComment']
			if comm is None:
				comm = userComment
			else:
				comm = '%s. %s' % (userComment,comm)
		
		sheet.cell(row=l,column=header.index('Commentaire')+1).value = comm
		sheet.cell(row=l,column=header.index('Gene')+1).value = db_variant['gene']
		sheet.cell(row=l,column=header.index('Transcript')+1).value = db_variant['transcript']
		sheet.cell(row=l,column=header.index('Chr')+1).value = db_variant['chromosome']
		sheet.cell(row=l,column=header.index('Exon')+1).value = db_variant['exon']
		sheet.cell(row=l,column=header.index('Intron')+1).value = db_variant['intron']
		sheet.cell(row=l,column=header.index('Position')+1).value = db_variant['genomicStart']
		sheet.cell(row=l,column=header.index('Ref')+1).value = db_variant['referenceAllele']
		sheet.cell(row=l,column=header.index('Alt')+1).value = db_variant['alternativeAllele']
		sheet.cell(row=l,column=header.index('c.')+1).value = db_variant['transcriptDescription']
		sheet.cell(row=l,column=header.index('p.')+1).value = db_variant['proteinDescription']
		sheet.cell(row=l,column=header.index('Region')+1).value = db_variant['region']
		sheet.cell(row=l,column=header.index('Consequence')+1).value = db_variant['consequence']
		sheet.cell(row=l,column=header.index('InterVar')+1).value = db_variant['intervar']
		sheet.cell(row=l,column=header.index('ClinVar')+1).value = db_variant['clinvar']
		sheet.cell(row=l,column=header.index('COSMIC')+1).value = db_variant['cosmic']
		sheet.cell(row=l,column=header.index('dbSNP')+1).value = db_variant['dbsnp']
		sheet.cell(row=l,column=header.index('gnomAD')+1).value = db_variant['gnomad']
		sheet.cell(row=l,column=header.index('1000G_ALL')+1).value = db_variant['milleGall']
		sheet.cell(row=l,column=header.index('1000G_EUR')+1).value = db_variant['milleGeur']
		sheet.cell(row=l,column=header.index('NCI60')+1).value = db_variant['nci60']
		sheet.cell(row=l,column=header.index('ESP')+1).value = db_variant['esp']
		sheet.cell(row=l,column=header.index('ExAC')+1).value = db_variant['exac']
		sheet.cell(row=l,column=header.index('SIFT')+1).value = db_variant['sift']
		sheet.cell(row=l,column=header.index('POLYPHEN2')+1).value = db_variant['polyphen2']
		sheet.cell(row=l,column=header.index('PubMed')+1).value = db_variant['pubmed']
		sheet.cell(row=l,column=header.index('VEP_Consequence')+1).value = db_variant['vep_consequence']
		sheet.cell(row=l,column=header.index('VEP_Impact')+1).value = db_variant['vep_impact']
		sheet.cell(row=l,column=header.index('VEP_Diff')+1).value = db_variant['vep_diff']
		sheet.cell(row=l,column=header.index('Class.')+1).value = db_variant['variantType']
		sheet.cell(row=l,column=header.index('c.(annovar)')+1).value = db_variant['annovarTranscriptDescription']
		sheet.cell(row=l,column=header.index('p.(annovar)')+1).value = db_variant['annovarProteinDescription']
		sheet.cell(row=l,column=header.index('annoWarning')+1).value = db_variant['annoWarning']

		if db_variant['region'] in ['exonic','splicing'] and db_variant['consequence'] != 'synonymous':
			cell_format(sheet.cell(row=l,column=header.index('c.')+1),alignment='left',style='LightGreen')
			cell_format(sheet.cell(row=l,column=header.index('p.')+1),alignment='left',style='LightGreen')
			cell_format(sheet.cell(row=l,column=header.index('Region')+1),alignment='left',style='LightGreen')
			cell_format(sheet.cell(row=l,column=header.index('Consequence')+1),alignment='left',style='LightGreen')
		elif db_variant['region'] == 'exonic;splicing':
			cell_format(sheet.cell(row=l,column=header.index('c.')+1),alignment='left',style='LightGreen')
			cell_format(sheet.cell(row=l,column=header.index('p.')+1),alignment='left',style='LightGreen')
			cell_format(sheet.cell(row=l,column=header.index('Region')+1),alignment='left',style='LightGreen')
			cell_format(sheet.cell(row=l,column=header.index('Consequence')+1),alignment='left',style='LightGreen')
		if db_variant['transcriptDescription'] != None:
			if ('+' in db_variant['transcriptDescription'] or '-' in db_variant['transcriptDescription'] or '*' in db_variant['transcriptDescription']) and (db_variant['region'] != 'exonic' and db_variant['region'] != '?'):
				cell_format(sheet.cell(row=l,column=header.index('c.')+1),alignment='left',style='TextBlue')
				cell_format(sheet.cell(row=l,column=header.index('Region')+1),alignment='left',style='TextBlue')

		variantCovs.append(db_variant['variantReadDepth'])
		variantDepths.append(db_variant['positionReadDepth'])
		variantFreqs.append(int((float(db_variant['variantReadDepth'])/float(db_variant['positionReadDepth']))*100))
		variantSamples.append('%s-%s' % (db_variant['sampleName'],db_variant['sampleID']))
		if vbformat == 'ColonLungMela':
			if db_variant['pathology'] == 'poumon':
				lcount += 1
			elif db_variant['pathology'] == 'colon':
				ccount += 1
			elif db_variant['pathology'] == 'melanome':
				mcount += 1
			else:
				ocount += 1
	else:
		variantCovs.append(db_variant['variantReadDepth'])
		variantDepths.append(db_variant['positionReadDepth'])
		variantFreqs.append(int((float(db_variant['variantReadDepth'])/float(db_variant['positionReadDepth']))*100))
		variantSamples.append('%s-%s' % (db_variant['sampleName'],db_variant['sampleID']))
		if vbformat == 'ColonLungMela':
			if db_variant['pathology'] == 'poumon':
				lcount += 1
			elif db_variant['pathology'] == 'colon':
				ccount += 1
			elif db_variant['pathology'] == 'melanome':
				mcount += 1
			else:
				ocount += 1

excelVB.save(output)
