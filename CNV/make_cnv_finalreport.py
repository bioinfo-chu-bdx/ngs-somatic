#!/usr/bin/env python
import os
import sys
import csv
import glob
import openpyxl

def cell_format(cell, copy, amp_colors, del_colors, alignment='center', border=None):
	color = 'Grey'
	clearText = False

	if copy == 'insufficient':
		color = 'Grey'
	elif copy == 'unchanged':
		color = 'LightGrey'
	else:		
		if copy >= 3.0:
			copy_rounded = int(round(copy))
			if copy_rounded > 15:
				color = amp_colors[15]
			else:
				color = amp_colors[copy_rounded]
		elif copy <= 1.7:
			copy_rounded = float("%.1f"%(copy))
			if copy_rounded < 0.5:
				color = del_colors[0.5]
			else:
				color = del_colors[copy_rounded]
		if copy <= 0.9 or copy >= 11.0:
			clearText=True

	cell.alignment = openpyxl.styles.Alignment(horizontal=alignment)
	if border:
		cell.border = openpyxl.styles.Border(right=openpyxl.styles.Side(border_style=border))

	if color == 'Grey':
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, color='77787c')
	elif color == 'LightGrey':
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, color='BFBFBF')#green 77a513 brow 949454 purple B1A0C7
	else:
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color=color)
		if clearText:
			cell.font = openpyxl.styles.Font(name='Calibri', size=11, color="FFFFFF")
		else:
			cell.font = openpyxl.styles.Font(name='Calibri', size=11)


cna_folder = sys.argv[1]
pipeline_folder = os.environ['NGS_PIPELINE_BX_DIR']
CNV_report_path = "%s/CNV/CNV_Report_Empty_ioncopy.xlsx" % pipeline_folder

amp_colors = {3:'dfffda', 4:'cffcc7', 5:'bff9b5', 6:'aff6a3', 7:'9ff391', 8:'8ff07f', 9:'80ee6d', 10:'6dd95b', 11:'5bc449', 12:'49b037', 13:'379b25', 14:'258613', 15:'137202'}
del_colors = {1.7:'f2dcdb', 1.6:'efd3d2', 1.5:'eccac9', 1.4:'e9c1c0', 1.3:'e6b8b7', 1.2:'e3afae', 1.1:'e0a7a5', 1.0:'dd9e9c', 0.9:'da9694', 0.8:'c97e7c', 0.7:'b86664', 0.6:'a74e4c', 0.5:'963634'}

CNV_finalreport = openpyxl.load_workbook(CNV_report_path)
results_sheet = CNV_finalreport.get_sheet_by_name("copy number analysis")

current_row = 3

### IONCOPY RESULTS ###
ioncopy_gain_file = open(cna_folder+'/ioncopy_gain.tsv','r')
gain_reader = csv.reader(ioncopy_gain_file,delimiter='\t')
ioncopy_loss_file = open(cna_folder+'/ioncopy_loss.tsv','r')
loss_reader = csv.reader(ioncopy_loss_file,delimiter='\t')

### GAIN RESULTS
column2sample = {}
header = gain_reader.next()
for i in range(len(header)):
	column2sample[i+1] = header[i] # i+1 car decalage entre header et le reste d'une case

sample_amplicon_cn = {}
for sample in column2sample.values():
	sample_amplicon_cn[sample] = {}

for line in gain_reader:
	for i in range(1,len(line)):
		sample = column2sample[i]
		if line[0] not in sample_amplicon_cn[sample].keys():
			sample_amplicon_cn[sample][line[0]] = {}
		if line[i] == '':
			cn = 2.0
		else:
			cn = line[i].split(',')[0].replace('CN=','')
		sample_amplicon_cn[sample][line[0]]["gain"] = float(cn)
		
### LOSS RESULTS
column2sample = {}
header = loss_reader.next()
for i in range(len(header)):
	column2sample[i+1] = header[i] # i+1 car decalage entre header et le reste d'une case

for line in loss_reader:
	for i in range(1,len(line)):
		sample = column2sample[i]
		if line[i] == '':
			cn = 2.0
		else:
			cn = line[i].split(',')[0].replace('CN=','')
		sample_amplicon_cn[sample][line[0]]["loss"] = float(cn)
		
# COMBINE RESULTS
sample_gene_cn = {}
for sample in column2sample.values():
	sample_gene_cn[sample] = {}
	
for sample in sample_amplicon_cn.keys():
	for amplicon in sample_amplicon_cn[sample].keys():
		gene = amplicon.split('_')[0]
		if gene not in sample_gene_cn[sample].keys():
			sample_gene_cn[sample][gene] = []
		gain_value = sample_amplicon_cn[sample][amplicon]["gain"]
		loss_value = sample_amplicon_cn[sample][amplicon]["loss"]
		
		if gain_value == 2.0 and loss_value == 2.0:
			sample_gene_cn[sample][gene].append(2.0)
		elif gain_value == 2.0:
			sample_gene_cn[sample][gene].append(loss_value)
		elif loss_value == 2.0:
			sample_gene_cn[sample][gene].append(gain_value)
		else:
			print "Problem : both gain and loss for %s - %s" % (sample,amplicon)
	
# Ecriture resultats
gene_index = 2
gene2column = {}
for gene in sample_gene_cn[sample].keys(): # a partir du dernier "sample" en variable
	results_sheet.cell(row=1,column=gene_index).value = gene
	results_sheet.cell(row=2,column=gene_index).value = 'IONCOPY'
	gene2column[gene] = gene_index
	gene_index = gene_index + 1

sample2row = {}
for sample in sample_gene_cn.keys():
	results_sheet.cell(row=current_row,column=1).value = sample.replace('.','-')
	results_sheet.cell(row=current_row,column=1).border = openpyxl.styles.Border(right=openpyxl.styles.Side(border_style="medium"))
	sample2row[sample] = current_row
	for gene in sample_gene_cn[sample].keys():
		mean_copy = sum(sample_gene_cn[sample][gene])/len(sample_gene_cn[sample][gene])
		if mean_copy == 2.0:
			results_sheet.cell(row=current_row,column=gene2column[gene]).value = "unchanged"
			cell_format(results_sheet.cell(row=current_row,column=gene2column[gene]), "unchanged", amp_colors, del_colors, border="medium")
		else:
			results_sheet.cell(row=current_row,column=gene2column[gene]).value = round(mean_copy,1)
			cell_format(results_sheet.cell(row=current_row,column=gene2column[gene]), mean_copy, amp_colors, del_colors, border="medium")
	current_row = current_row + 1


#ioncopy_file.seek(1)
### GRISEE LIGNE mauvais sequencage en utilisant donnees du ioncopy_input#####################
#ioncopy_input_path = cna_folder.replace('/_CNA','/ioncopy_input.tsv')
ioncopy_input_path = cna_folder + '/ioncopy_input.tsv'
ioncopy_input = open(ioncopy_input_path,'r')
ioncopy_input_reader = csv.reader(ioncopy_input,delimiter='\t')
data = list(ioncopy_input_reader)
row_number = len(data)
samples = data[0]

for i in range(1,len(samples)):
	inf300X = 0
	for j in range(1,row_number):
		if int(data[j][i]) < 300:
			inf300X = inf300X + 1
	if inf300X >= int(row_number/4): # un quart des cibles < 300X... a voir. Aussi si grand panel, 300X est peut-etre trop
		try:
			f_name = samples[i].replace('-','.').replace(' ','.')
			current_row = sample2row[f_name]
		except:
			continue
		cell = results_sheet.cell(row=current_row,column=1)
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, color='963634')
		for k in range (1,results_sheet.max_column+1):
			cell = results_sheet.cell(row=current_row,column=k)
			cell.fill = openpyxl.styles.PatternFill(fill_type='lightUp')
##############################################################################################

#run_name = cna_folder.split("/")[-4].replace("Auto_user_","")
#run_name = run_name.split("_")[0]
#CNV_finalreport.save(cna_folder+"/CNV_finalReport_%s.xlsx" % run_name)
CNV_finalreport.save(cna_folder+"/CNV_finalReport.xlsx")

##amp_colors = {3:'d3d6f2', 4:'ccbeee', 5:'c5a6eb', 6:'be8ee8', 7:'b776e5', 8:'a159d4', 9:'8c3cc4', 10:'761fb3', 11:'6103a3', 12:'53038a', 13:'450472', 14:'37055a', 15:'290642'}
##del_colors = {1.7:'e3d54f', 1.6:'e5cc3b', 1.5:'e7c327', 1.4:'e9ba13', 1.3:'ecb200', 1.2:'e6a106', 1.1:'e1910d', 1.0:'db8013', 0.9:'d6701a', 0.8:'d35c15', 0.7:'d14910', 0.6:'ce360b', 0.5:'cc2306'}
##del_colors = {1.7:'eea3a3', 1.6:'ed9b9b', 1.5:'ed9494', 1.4:'ec8d8d', 1.3:'ec8686', 1.2:'ee7a7a', 1.1:'f16f6f', 1.0:'f36363', 0.9:'f65858', 0.8:'d94747', 0.7:'bc3636', 0.6:'9f2525', 0.5:'831515'}

