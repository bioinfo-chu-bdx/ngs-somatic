#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import sys
import openpyxl

workbook = openpyxl.load_workbook('/media/stuff/Archer_Acrometrix_compare.xlsx')
# acro_sheet = workbook['variants_acrometrix']
# archer_sheet = workbook['hotspots_archer']
# compare_sheet = workbook['compare']
acro_sheet = workbook['variants_acrometrix']
archer_sheet = workbook['test "dna analysis"']
compare_sheet = workbook['compare2']

acro_variants = []
var2genomic = {}

for nrow in range(14,acro_sheet.max_row+1):
	gene = acro_sheet.cell(row=nrow,column=7).value
	cpos = acro_sheet.cell(row=nrow,column=8).value
	chromosome = acro_sheet.cell(row=nrow,column=1).value
	position = acro_sheet.cell(row=nrow,column=2).value
	reference = acro_sheet.cell(row=nrow,column=3).value
	alternate = acro_sheet.cell(row=nrow,column=4).value
	genomic_description = 'chr%s:%s:%s>%s' % (chromosome,position,reference,alternate)
	acro_variants.append((gene,cpos))
	var2genomic[(gene,cpos)] = genomic_description

compare_row = 2
for nrow in range(4,archer_sheet.max_row+1):
	gene = archer_sheet.cell(row=nrow,column=4).value
	transcript_cpos = archer_sheet.cell(row=nrow,column=6).value
	cpos = transcript_cpos.split(':')[-1]
	genodesc = archer_sheet.cell(row=nrow,column=32).value
	refalt = archer_sheet.cell(row=nrow,column=33).value.replace(' ','').replace('/','>')
	genomic_description = '%s:%s' % (genodesc,refalt)

	if ((gene,cpos) in acro_variants) or (genomic_description in var2genomic.values()):
		depth = archer_sheet.cell(row=nrow,column=7).value
		ao = archer_sheet.cell(row=nrow,column=8).value
		af = archer_sheet.cell(row=nrow,column=9).value
		compare_sheet.cell(row=compare_row,column=1).value = '%s:%s' % (gene,cpos)
		compare_sheet.cell(row=compare_row,column=2).value = af
		compare_sheet.cell(row=compare_row,column=3).value = depth
		compare_sheet.cell(row=compare_row,column=4).value = ao
		if genomic_description in var2genomic.values():
			compare_sheet.cell(row=compare_row,column=5).value = genomic_description
			compare_sheet.cell(row=compare_row,column=7).value = 'differente cpos'
		else:
			compare_sheet.cell(row=compare_row,column=5).value = var2genomic[(gene,cpos)]
		compare_sheet.cell(row=compare_row,column=6).value = genomic_description
		compare_row += 1

workbook.save('/media/stuff/Archer_Acrometrix_compare.xlsx')
