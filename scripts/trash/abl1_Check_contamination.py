#!/usr/bin/env python
import subprocess
import os
import pysam
import json
import csv
import openpyxl
import copy
from optparse import OptionParser

#### MODIF ABl1 : annoSheet_rows[i][5].value ---> annoSheet_rows[i][6].value    pareil avec contaSheet
#### MODIF ABl1 : run_analysis.py ---> run_cdna_abl1_analysis.py


"""Plugin object to check control sample contamination"""

def representsInt(s):
	try: 
		s = int(s)
		return s
	except ValueError:
		return s
		
############################################################################################
parser = OptionParser()
parser.add_option('-i', '--bam', 			help="Input bam file ", 			dest='bam_file')
parser.add_option('-l', '--read-len', 		help="Min read len for filtering ", dest='read_len')
(options, args) = parser.parse_args()

sample = options.bam_file.split('/')[-1].split('_IonXpress')[0]
barcode = 'IonXpress_' + options.bam_file.split('IonXpress_')[-1].split('.bam')[0]
sample_folder = os.path.dirname(options.bam_file)
run_folder = os.path.dirname(sample_folder)

with open('/DATA/work/global_parameters.json', 'r') as g:
	global_param = json.load(g)
	
if os.path.isfile(run_folder+'/barcodes.json'):
	with open(run_folder+'/barcodes.json', 'r') as g:
		barcodes_json = json.load(g)
		sample_bed = barcodes_json[barcode]['target_region_filepath'].split('/unmerged/detail/')[-1]
		sample_lib = barcodes_json[barcode]['barcode_description']
		for _run_type in global_param['run_type']:
			if global_param['run_type'][_run_type]['target_bed'].split('/')[-1] == sample_bed:
				run_type = _run_type
				break
else:
	print "error : barcodes.json not found in run folder"
target_bed_unmerged = global_param['run_type'][run_type]['target_bed']
	
############################################################################################

# dossier results
results_name = options.bam_file.split('/')[-1].replace(' ','@')
results_folder = '/DATA/work/results/' + results_name
if not os.path.isdir(results_folder):
	subprocess.call(['mkdir', results_folder])
conta_folder = results_folder + '/check-contamination'
if not os.path.isdir(conta_folder):
	subprocess.call(['mkdir', conta_folder])

	######################
	### BAM FILTERING  ###
	######################

bamfile = pysam.Samfile(options.bam_file,'rb')
bamfile_filtered = pysam.Samfile('%s/%s_%s.filtered.bam' % (conta_folder,sample,barcode), 'wb', template=bamfile)

### BED DATA ###
amplicons = {}
bedfile = open(target_bed_unmerged,'r')
target_reader = csv.reader(bedfile, delimiter = '\t')
target_reader.next()
for row in target_reader:
	s = row[7].split(';')
	s[0] = s[0].split('=')[-1]
	s[1] = s[1].split('=')[-1]
	amplicons[row[3]] = {}
	amplicons[row[3]]['chr'] = row[0]
	amplicons[row[3]]['startpos'] = row[1]
	amplicons[row[3]]['endpos'] = row[2]
	amplicons[row[3]]['gene_id'] = s[0]
	amplicons[row[3]]['transcrit'] = s[1]
	amplicons[row[3]]['contamination count'] = 0
bedfile.close()

### CREATE NEW BAM WITH READS > 100 nc ###
print "bam filtering..."
for read in bamfile.fetch():
	if (len(read.query) >= int(options.read_len)): 	# aligned portion of the read, exclude soft-clipped bases
		bamfile_filtered.write(read)
bamfile.close()
bamfile_filtered.close()

# SORT NEW BAM AND CREATE BAI
print "samtools sort..."
cmd = subprocess.call(['samtools','sort', '%s/%s_%s.filtered.bam' % (conta_folder,sample,barcode),'-o', '%s/%s_%s.filtered.sorted.bam' % (conta_folder,sample,barcode)])

print "samtools index..."
cmd = subprocess.call(['samtools','index','%s/%s_%s.filtered.sorted.bam' % (conta_folder,sample,barcode)])

	#########################
	### NEW BAM ANALYSIS  ###
	#########################
	
print "filtered bam complete analysis..."
cmd = subprocess.Popen(['python','/DATA/work/run_cdna_abl1_analysis.py','--bam','%s/%s_%s.filtered.sorted.bam' % (conta_folder,sample,barcode),'--run-type',run_type],stdout=subprocess.PIPE)
out, err = cmd.communicate()
print 'OUT: %s\nERR: %s'%(out, err)	

##################################################################
### COMPARAISON AVEC COUVERTURE / VARIANTS DES AUTRES PATIENTS ###
##################################################################

# creation dic avec variants de l'annotation du check conta
variantSearch = {}
conta_finalreport = openpyxl.load_workbook('%s/%s_%s.filtered.sorted.finalReport.xlsx' % (conta_folder,sample,barcode))
amplSheet = conta_finalreport.get_sheet_by_name('Amplicon Coverage')
amplSheet_rows = tuple(amplSheet.rows)
for i in range(1,len(amplSheet_rows)):
	amplicons[amplSheet_rows[i][3].value]['contamination count'] = int(amplSheet_rows[i][9].value)
contaSheet = conta_finalreport.get_sheet_by_name('Annotation')
contaSheet_rows = tuple(contaSheet.rows)
for i in range(1,len(contaSheet_rows)):
	if contaSheet_rows[i][2].value:
		variantSearch[(contaSheet_rows[i][2].value,contaSheet_rows[i][6].value)] = [] #(NM,c)

# recuperation de la liste des patients a comparer
sample2compare = []
for bc in barcodes_json:
	bc_bed = barcodes_json[bc]['target_region_filepath'].split('/unmerged/detail/')[-1]
	bc_lib = barcodes_json[bc]['barcode_description']
	if ((bc_lib == sample_lib) and (bc_bed == sample_bed)) and bc != barcode:
		sample2compare.append((barcodes_json[bc]['sample'],bc))

# pour les patients a comparer, recuperation de leur couverture par amplicon
for s in sample2compare:
	finalReport_path = '%s/%s/%s_%s.finalReport.xlsx' % (run_folder,s[0],s[0],s[1])
	if os.path.isfile(finalReport_path):
		# ouverture du finalReport, onglet couverture par amplicon, recuperation des infos
		finalReport = openpyxl.load_workbook(finalReport_path)
		covSheet = finalReport.get_sheet_by_name('Amplicon Coverage')
		covSheet_rows = tuple(covSheet.rows)
		for i in range(1,len(covSheet_rows)):
			amplicon = covSheet_rows[i][3].value
			totalReads = covSheet_rows[i][9].value
			amplicons[amplicon][s[0]] = int(totalReads)
		# onglet annotation, recherche si mutation present dans le bam filtre
		annoSheet = finalReport.get_sheet_by_name('Annotation')
		annoSheet_rows = tuple(annoSheet.rows)
		for i in range(1,len(annoSheet_rows)):
			if (annoSheet_rows[i][2].value,annoSheet_rows[i][6].value) in variantSearch.keys():
				variantSearch[(annoSheet_rows[i][2].value,annoSheet_rows[i][6].value)].append(s[0]+'_'+s[1])

###############################
#### ecriture des resultats ###
###############################

conta_report = openpyxl.Workbook()
ampliconSheet = conta_report.create_sheet(title='Amplicons Contamination')
annotationSheet = conta_report.create_sheet(title='Annotation')
try:
	del conta_report['Sheet']
except:
	pass

# AMPLICON SHEET
header = ['Amplicon','Chr','Start','End','Gene_id', 'Transcript', '%s reads > %s b' % (sample,options.read_len)]
for s in sample2compare:
	header.append(s[0])
for i in range(len(header)):
	ampliconSheet.cell(row=1,column=i+1).value = header[i]
	ampliconSheet.cell(row=1,column=i+1).font = openpyxl.styles.Font(name='Calibri', size=11, bold=True)
	ampliconSheet.cell(row=1,column=i+1).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'),bottom=openpyxl.styles.Side(style='thin'))

to_write = []
for amplicon in amplicons.keys():
	line = [amplicon,amplicons[amplicon]['chr'],amplicons[amplicon]['startpos'],amplicons[amplicon]['endpos'],amplicons[amplicon]['gene_id'],amplicons[amplicon]['transcrit'],amplicons[amplicon]['contamination count']]
	for s in sample2compare:
		line.append(amplicons[amplicon][s[0]])
	to_write.append(line)
to_write.sort(key=lambda x: x[6]) # classement par ordre decroissant de nombre de reads > read_len
to_write.reverse()

for i in range(len(to_write)):
	for j in range(len(to_write[i])):
		ampliconSheet.cell(row=i+2,column=j+1).font = openpyxl.styles.Font(name='Calibri', size=11)
		if j <= 6:
			ampliconSheet.cell(row=i+2,column=j+1).value = representsInt(to_write[i][j])
			if j == 6 and int(to_write[i][6]) >= 100:
				ampliconSheet.cell(row=i+2,column=j+1).font = openpyxl.styles.Font(name='Calibri', size=11, color='ff0000')
		elif j > 6:
			if int(to_write[i][j]) < 5*(int(to_write[i][6])):
				ampliconSheet.cell(row=i+2,column=j+1).value = representsInt(to_write[i][j])
				ampliconSheet.cell(row=i+2,column=j+1).fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='d28e8e')
			else:
				ampliconSheet.cell(row=i+2,column=j+1).value = representsInt(to_write[i][j])

# ANNOTATION SHEET
cosmic_index = False
annotationSheet.cell(row=1,column=1).value = 'Library Search'
annotationSheet.cell(row=1,column=1).font = openpyxl.styles.Font(name='Calibri', size=11, bold=True)
annotationSheet.cell(row=1,column=1).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'),bottom=openpyxl.styles.Side(style='thin'))
for j in range(len(contaSheet_rows[0])):
	annotationSheet.cell(row=1,column=j+2).value = contaSheet_rows[0][j].value
	annotationSheet.cell(row=1,column=j+2).font = openpyxl.styles.Font(name='Calibri', size=11, bold=True)
	annotationSheet.cell(row=1,column=j+2).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'),bottom=openpyxl.styles.Side(style='thin'))
	if contaSheet_rows[0][j].value == 'COSMIC':
		cosmic_index = j


for i in range(1,len(contaSheet_rows)):
	if contaSheet_rows[i][2].value:
		if not contaSheet_rows[i][0].value:
			contaSheet_rows[i][0].value = '.'
		if variantSearch[contaSheet_rows[i][2].value,contaSheet_rows[i][6].value]: # list not empty = variant found somewhere
			foundstring = ','.join(variantSearch[(contaSheet_rows[i][2].value,contaSheet_rows[i][6].value)])
			annotationSheet.cell(row=i+1,column=1).value = 'Variant found in ' + foundstring
			annotationSheet.cell(row=i+1,column=1).font = openpyxl.styles.Font(name='Calibri', size=11, color='ff0000')
	for j in range(len(contaSheet_rows[i])):
		if contaSheet_rows[i][j].value:
			annotationSheet.cell(row=i+1,column=j+2).value = representsInt(contaSheet_rows[i][j].value)
			annotationSheet.cell(row=i+1,column=j+2).font = openpyxl.styles.Font(name='Calibri', size=11)
			annotationSheet.cell(row=i+1,column=j+2).fill = copy.copy(contaSheet_rows[i][j].fill)

# coloration rouge cosmic
if cosmic_index:
	for i in range(len(contaSheet_rows)):
		if annotationSheet.cell(row=i+2,column=cosmic_index+2).value:
			if annotationSheet.cell(row=i+2,column=cosmic_index+2).value != '.':
				annotationSheet.cell(row=i+2,column=cosmic_index+2).font = openpyxl.styles.Font(name='Calibri', size=11, color='ff0000')

## LAYOUT
ws = conta_report['Amplicons Contamination']
ws.column_dimensions[ws['A1'].column].width = 18
ws.column_dimensions[ws['B1'].column].width = 8
ws.column_dimensions[ws['C1'].column].width = 12
ws.column_dimensions[ws['D1'].column].width = 12
ws.column_dimensions[ws['E1'].column].width = 15
ws.column_dimensions[ws['F1'].column].width = 15
ws = conta_report['Annotation']
maxsize = 15
dims = {}
for row in ws.rows:
	for cell in row:
		if cell.value:
			try:
				text_size = len(str(cell.value)) + 2
				dims[cell.column] = min(max(dims.get(cell.column, 0), text_size),maxsize)
			except:
				pass
for col, value in dims.items():
	ws.column_dimensions[col].width = value
ws.column_dimensions[ws['J1'].column].width = 8
				
conta_report.save(conta_folder+'/Check-contamination_%s.xlsx'%sample)
subprocess.call(['cp',conta_folder+'/Check-contamination_%s.xlsx'%sample,run_folder+'/_Check-contamination/Check-contamination_%s.xlsx'%sample])
