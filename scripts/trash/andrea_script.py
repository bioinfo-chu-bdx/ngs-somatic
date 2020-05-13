#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl
from datetime import date

andrea_file_path = '/media/stuff/1 - Liste des échantillons requêtés.xlsx'
andrea_file = openpyxl.load_workbook(andrea_file_path)
output = openpyxl.Workbook()
del output['Sheet']

inputSheet = andrea_file['Feuil1']
outputSheet = output.create_sheet(title='Samples')

header_row = 9 # header row in input
actual_sample = ''
sample_oldest_date = False
MinDays = 456

# REWRITE HEADER
for j in range(1,inputSheet.max_column+1):
	outputSheet.cell(row=1,column=j).value = inputSheet.cell(row=header_row,column=j).value

# PARSE LINES AND FILTER
r = 2
for i in range(header_row+1,inputSheet.max_row+1):
	sample = inputSheet.cell(row=i,column=2).value
	dp = inputSheet.cell(row=i,column=6).value
	if sample != actual_sample:
		sample_first_line = []
		for j in range(1,inputSheet.max_column+1):
			sample_first_line.append(inputSheet.cell(row=i,column=j).value)
		first_line_written = False
		actual_sample = sample
		sample_oldest_date = dp
		continue
	try:
		delta = dp - sample_oldest_date
	except:
		continue
	print r
	if delta.days >= MinDays:
		if first_line_written == False:
			for z in range(len(sample_first_line)):
				outputSheet.cell(row=r,column=z+1).value = sample_first_line[z]
			r += 1
			first_line_written = True
		for j in range(1,inputSheet.max_column+1):
			outputSheet.cell(row=r,column=j).value = inputSheet.cell(row=i,column=j).value
		r += 1

output.save('/media/stuff/Liste echantillon filtree.xlsx')
