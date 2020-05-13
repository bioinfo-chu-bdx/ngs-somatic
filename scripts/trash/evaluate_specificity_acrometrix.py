#!/usr/bin/env python
import sys
import os
import glob
import openpyxl


#acrometrix_file = '/media/n06lbth/sauvegardes_pgm/SBT/Suivi_temoin_Acrometrix.xlsx'
acrometrix_file = '/media/n06lbth/sauvegardes_pgm/tests_qualification_pipeline/Specificity/HEMATO/Suivi_temoin_Acrometrix.xlsx'

acrometrix_finalreport_list = [
#'/media/n06lbth/sauvegardes_pgm/SBT/Run_300-399/Auto_user_S5-0198-185-Run319_320_35pM_Chef_SBT-colon-lung_v8_530_354/AcroMetrix_lib1_NGS319-/AcroMetrix_lib1_NGS319-_IonXpress_006.finalReport.xlsx',
#'/media/n06lbth/sauvegardes_pgm/SBT/Run_300-399/Auto_user_S5-0198-185-Run319_320_35pM_Chef_SBT-colon-lung_v8_530_354/AcroMetrix_lib2_NGS320-/AcroMetrix_lib2_NGS320-_IonXpress_026.finalReport.xlsx',
#'/media/n06lbth/sauvegardes_pgm/SBT/Run_300-399/Auto_user_S5-0198-186-NGS321_322_35pM_Chef_SBT-colon-lung_v8_530_356/AcroMetrix_lib1_NGS321-/AcroMetrix_lib1_NGS321-_IonXpress_053.finalReport.xlsx',
#'/media/n06lbth/sauvegardes_pgm/SBT/Run_300-399/Auto_user_S5-0198-186-NGS321_322_35pM_Chef_SBT-colon-lung_v8_530_356/AcroMetrix_lib2_NGS322-/AcroMetrix_lib2_NGS322-_IonXpress_080.finalReport.xlsx',
#'/media/n06lbth/sauvegardes_pgm/SBT/Run_300-399/Auto_user_S5-0198-188-run323_324_35pM_Chef_SBT-colon-lung_v8_530_359/AcroMetrix_lib1_NGS323-/AcroMetrix_lib1_NGS323-_IonXpress_011.finalReport.xlsx',
#'/media/n06lbth/sauvegardes_pgm/SBT/Run_300-399/Auto_user_S5-0198-192-Run324bis_35pM_Chef_SBT-colon-lung_v8_530_364/AcroMetrix_lib2_NGS324bis-/AcroMetrix_lib2_NGS324bis-_IonXpress_071.finalReport.xlsx',
#'/media/n06lbth/sauvegardes_pgm/SBT/Run_300-399/Auto_user_S5-0198-193-Run325_326_35pM_Chef_SBT-colon-lung_v8_530_365/AcroMetrix_lib2_NGS325-/AcroMetrix_lib2_NGS325-_IonXpress_003.finalReport.xlsx',
#'/media/n06lbth/sauvegardes_pgm/SBT/Run_300-399/Auto_user_S5-0198-193-Run325_326_35pM_Chef_SBT-colon-lung_v8_530_365/AcroMetrix_lib2_NGS326-/AcroMetrix_lib2_NGS326-_IonXpress_029.finalReport.xlsx',
#'/media/n06lbth/sauvegardes_pgm/SBT/Run_300-399/Auto_user_S5-0198-195-NGS327_328_35pM_Chef_SBT-colon-lung_v8_530_369/AcroMetrix_lib1_NGS327-/AcroMetrix_lib1_NGS327-_IonXpress_056.finalReport.xlsx',
#'/media/n06lbth/sauvegardes_pgm/SBT/Run_300-399/Auto_user_S5-0198-195-NGS327_328_35pM_Chef_SBT-colon-lung_v8_530_369/AcroMetrix_lib2_NGS328-/AcroMetrix_lib2_NGS328-_IonXpress_078.finalReport.xlsx',
#'/media/n06lbth/sauvegardes_pgm/SBT/Run_300-399/Auto_user_S5-0198-199-NGS329_330_35pM_Chef_SBT-colon-lung_v8_530_376/AcroMetrix_lib1_NGS329-/AcroMetrix_lib1_NGS329-_IonXpress_027.finalReport.xlsx',
#'/media/n06lbth/sauvegardes_pgm/SBT/Run_300-399/Auto_user_S5-0198-199-NGS329_330_35pM_Chef_SBT-colon-lung_v8_530_376/AcroMetrix_lib2_NGS330-/AcroMetrix_lib2_NGS330-_IonXpress_042.finalReport.xlsx',
#'/media/n06lbth/sauvegardes_pgm/SBT/Run_300-399/Auto_user_S5-0198-201-Run331_332_35pM_Chef_SBT-colon-lung_v8_530_378/AcroMetrix_lib1_NGS331-/AcroMetrix_lib1_NGS331-_IonXpress_065.finalReport.xlsx',
#'/media/n06lbth/sauvegardes_pgm/SBT/Run_300-399/Auto_user_S5-0198-201-Run331_332_35pM_Chef_SBT-colon-lung_v8_530_378/AcroMetrix_lib2_NGS332-/AcroMetrix_lib2_NGS332-_IonXpress_084.finalReport.xlsx',
#'/media/n06lbth/sauvegardes_pgm/SBT/Run_300-399/Auto_user_S5-0198-203-Run_bis_331_332_35pM_Chef_SBT-colon-lung_v8_530_379/AcroMetrix_lib1_NGS331-/AcroMetrix_lib1_NGS331-_IonXpress_065.finalReport.xlsx',
#'/media/n06lbth/sauvegardes_pgm/SBT/Run_300-399/Auto_user_S5-0198-203-Run_bis_331_332_35pM_Chef_SBT-colon-lung_v8_530_379/AcroMetrix_lib2_NGS332-/AcroMetrix_lib2_NGS332-_IonXpress_084.finalReport.xlsx',
#'/media/n06lbth/sauvegardes_pgm/SBT/Run_300-399/Auto_user_S5-0198-205-Run333_334_35pM_Chef_SBT-colon-lung_v9_530_381/AcroMetrix_lib1_NGS333-/AcroMetrix_lib1_NGS333-_IonXpress_017.finalReport.xlsx',
#'/media/n06lbth/sauvegardes_pgm/SBT/Run_300-399/Auto_user_S5-0198-205-Run333_334_35pM_Chef_SBT-colon-lung_v9_530_381/AcroMetrix_lib2_NGS334-/AcroMetrix_lib2_NGS334-_IonXpress_034.finalReport.xlsx',
#'/media/n06lbth/sauvegardes_pgm/SBT/Run_300-399/Auto_user_S5-0198-207-Run335_336_35pM_Chef_SBT-colon-lung_v9_530_383/AcroMetrix_lib1_NGS335-/AcroMetrix_lib1_NGS335-_IonXpress_061.finalReport.xlsx',
#'/media/n06lbth/sauvegardes_pgm/SBT/Run_300-399/Auto_user_S5-0198-207-Run335_336_35pM_Chef_SBT-colon-lung_v9_530_383/AcroMetrix_lib2_NGS336-/AcroMetrix_lib2_NGS336-_IonXpress_088.finalReport.xlsx',
'/media/n06lbth/sauvegardes_pgm/LAM/Auto_user_S5-0198-288-RUN135-2019-06-05-LAMV8-CF_494/acrometrix-SBT-/acrometrix-SBT-_IonXpress_024.finalReport.xlsx',
'/media/n06lbth/sauvegardes_pgm/LAM/Auto_user_S5-0198-290-Run136-LAMv8-TP53-ABL-FLT3-MM_497/acrometrix-SBT-exp-01-20/acrometrix-SBT-exp-01-20_IonXpress_025.finalReport.xlsx',
'/media/n06lbth/sauvegardes_pgm/LAM/Auto_user_S5-0198-293-Run137-LAMv8-TP53-CF_499/acrometrix-SBT-/acrometrix-SBT-_IonXpress_050.finalReport.xlsx'
]

nomen_error_list = ['c.1616_1624delinsGTCTCTCTA','c.746_753delinsGCCCGCAT','c.1138_1150delinsAGGGTGGGCTTCC','c.3778_3785delinsTGTGCAAG',
'c.2197_2203delinsTCAGAAA','c.2222_2249delinsTCGTCGCTATCAA','c.2429_2441delinsACTCCCAGTACCC','c.2485_2515delinsAAGGACCGTCGCGTGGTGCTCCGCGACCTGA',
'c.2573_2588delinsGGGCCAAACAGCTGGA','c.2582_2588delinsAGCTGGA','c.175_182delinsACCGGCCG','c.1405_1416delinsCTTGGAAAGCTG','c.1526_1535delinsTAGGTAACAG',
'c.175_183delinsACAGGTCAC','c.24_35delinsGGTTGGAGCTGA','c.3336_3370delinsCGGGACTTTGTTGGACAATGATGGCAAGAAAATTG','c.3778_3785delinsTGTGCAAG','c.174_182delinsGGCTGGACG',
'c.35_52delinsATGGTGTTGGGAAAAGCA','c.2517_2525delinsTGCCAGAGT','c.1616_1624delinsGTCTCTCTA','c.1633_1640delinsAAGCAGGG','c.580_595delinsTACCTGGGCGTGGCCT',

'c.578_615delinsCGTTGTTTCACAAGATGTTTGAAACTATTACAATA','c.860_863dup','c.1510_1535delinsGCCTATTTTAACCTTGCATTTATAGGTAACAG','c.1504_1509dup',
'c.860_861insTCTGT']

true_variants_exceptions = ['JAK2_c.1849G>T', 'JAK2_c.1860C>A', 'PTEN_c.477G>T', 'KRAS_c.183A>C', 'KRAS_c.175G>A', 'TP53_c.991C>T', 'TP53_c.981T>G', 'TP53_c.963A>G', 'TP53_c.949C>T']

error_list = []
acro_gene_cpos = []
targeted_gene_exon_list = []
sensitivities = []
sensitivities_without_exception = []
specificities = []

final_results = []

suivi_temoin_workbook = openpyxl.load_workbook(acrometrix_file)
suivi_temoin_sheet = suivi_temoin_workbook['Acrometrix']
for row_idx in range(3, suivi_temoin_sheet.max_row+1):
	gene = suivi_temoin_sheet.cell(row=row_idx,column=3).value
	exon = suivi_temoin_sheet.cell(row=row_idx,column=2).value
	cpos = suivi_temoin_sheet.cell(row=row_idx,column=6).value
	if exon == None:
		exon = ''
	targeted_gene_exon = '%s_%s' % (gene,exon)
	gene_cpos = '%s_%s' % (gene,cpos)
	acro_gene_cpos.append(gene_cpos)
	targeted_gene_exon_list.append((targeted_gene_exon))
targeted_gene_exon_list = list(set(targeted_gene_exon_list))

for acrometrix_finalreport in acrometrix_finalreport_list:
	print "- %s" % acrometrix_finalreport.split('/')[-1]
	acrometrix_workbook = openpyxl.load_workbook(acrometrix_finalreport)
	acrometrix_sheet = acrometrix_workbook['Annotation']

	variants_found = 0
	true_variants = 0
	false_variants = 0
	already_seen = []
	missing_true_variants = acro_gene_cpos[:]
	for row_idx in range(2, acrometrix_sheet.max_row-2):
		gene = acrometrix_sheet.cell(row=row_idx,column=4).value
		exon = acrometrix_sheet.cell(row=row_idx,column=5).value
		cpos = acrometrix_sheet.cell(row=row_idx,column=6).value
		if exon == None:
			exon = ''
		targeted_gene_exon = '%s_%s' % (gene,exon)
		gene_cpos = '%s_%s' % (gene,cpos)
		if targeted_gene_exon not in targeted_gene_exon_list:
			continue
		if not gene_cpos in already_seen:
			variants_found += 1
			already_seen.append(gene_cpos)
		else:
			continue
		if gene_cpos in acro_gene_cpos:
			true_variants += 1
			missing_true_variants.remove(gene_cpos)
		#elif cpos in nomen_error_list:
			#true_variants += 1
		else:
			print "\t -" + targeted_gene_exon + ":" + cpos
			false_variants += 1
			bla = targeted_gene_exon + ":" + cpos
			if bla not in error_list:
				error_list.append(bla)
	
	print "\t - Variants found :  %s" % variants_found
	print "\t - True variants :  %s" % true_variants
	missing_true_variants_without_exception = [var for var in missing_true_variants if var not in true_variants_exceptions]
	print "\t - Variants missing (without exceptions) : %s " % missing_true_variants_without_exception #missing_true_variants
	print "\t - False variants %s" % false_variants
				
	sensitivity = float(true_variants)/float(len(acro_gene_cpos))
	sensitivity_without_exception = float(true_variants)/float(len(acro_gene_cpos)-len(true_variants_exceptions))
	sensitivities.append(sensitivity)
	sensitivities_without_exception.append(sensitivity_without_exception)
	print "\t - Sensitivity =  %s" % sensitivity
	print "\t - Sensitivity (without exceptions) =  %s" % sensitivity_without_exception
				
	specificity = float(true_variants)/float(variants_found)
	specificities.append(specificity)
	print "\t - Specificity =  %s" % specificity

print "---- RESULTS ----"
mean_sensi = 0.0
for i in range(len(sensitivities)):
	mean_sensi = mean_sensi + sensitivities[i]
mean_sensi = mean_sensi / float(len(sensitivities))
print "MEAN SENSITIVITY = %.2f" % mean_sensi
mean_sensi_exception = 0.0
for i in range(len(sensitivities_without_exception)):
	mean_sensi_exception = mean_sensi_exception + sensitivities_without_exception[i]
mean_sensi_exception = mean_sensi_exception / float(len(sensitivities_without_exception))
print "MEAN SENSITIVITY (without exceptions) = %.2f" % mean_sensi_exception
mean_speci = 0.0
for i in range(len(specificities)):
	mean_speci = mean_speci + specificities[i]
mean_speci = mean_speci / float(len(specificities))
print "MEAN SPECIFICITY = %.2f" % mean_speci


#for res in final_results:
	#print res
