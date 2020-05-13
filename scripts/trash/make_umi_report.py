#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import sys
import csv
import openpyxl

def representsInt(s): # pour eviter avertissement "nombre ecrit en texte" sous excel
	try: 
		s = int(s)
		return s
	except ValueError:
		return s
		
def cell_format(cell, font=None, alignment=None, color=None, format=None, border=None):
	if font == 'bold':
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, bold=True)
	elif font == 'red':
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, color='f44242')
	else:
		cell.font = openpyxl.styles.Font(name='Calibri', size=11)
	if alignment == 'center':
		cell.alignment = openpyxl.styles.Alignment(horizontal='center')
	elif alignment == 'vertical-center':
		cell.alignment = openpyxl.styles.Alignment(vertical='center')
	elif alignment == 'left':
		cell.alignment = openpyxl.styles.Alignment(horizontal='left')
	if color == 'LightGreen':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='D8E4BC') # EBF1DE
	elif color == 'LightRed':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='d28e8e') #F2DCDB
	elif color == 'LightOrange':	
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='ffe7e0') # ffd8ad
	elif color == 'LightBlue':	
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='add8e6')
	elif color == 'LightPink':	
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='ffc9f7')
	elif color == 'Yellow':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='feffa3')
	elif color == 'Blue':
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, color='004c99')
	else:
		cell.fill = openpyxl.styles.PatternFill(fill_type=None,start_color='FFFFFF')
	if border == 'thin':
		cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'),bottom=openpyxl.styles.Side(style='thin'))
	if border == 'top_medium':
		cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='medium'))
	if format == 'Percent':
		cell.number_format = '0.0%'


#########################
###### MAIN SCRIPT ######
#########################

column_reorder = { # csvreader -> openpyxl (base 1)
6:1,0:2,17:3,16:4,21:5,22:6,23:7,12:8,11:9,10:10,9:11,7:12,14:13,15:14,1:15,3:16,4:17,2:18,5:19,8:20,13:21,18:22,19:23,20:24,24:25,25:26,26:27,27:28,28:29,29:30,30:31
}

smcounter_anno_file = open(sys.argv[1],'r')
smcounter_anno_reader = csv.reader(smcounter_anno_file,delimiter='\t')

umi_finalreport = openpyxl.Workbook()
annotationSheet = umi_finalreport.create_sheet(title='smCounter.anno')
try:
	del umi_finalreport['Sheet']
except:
	pass

header = smcounter_anno_reader.next()
pos_sorted_reader = sorted(smcounter_anno_reader, key=lambda x: int(x[1]))
chr_sorted_reader = sorted(pos_sorted_reader, key=lambda x: int(x[0].replace('chrX','chr24').replace('chr','')))

for i in range(len(header)):
	annotationSheet.cell(row=1,column=column_reorder[i]).value = header[i]
	cell_format(annotationSheet.cell(row=1,column=column_reorder[i]),font='bold',border='thin')

l = 2
for line in chr_sorted_reader:
	for i in range(len(line)):
		annotationSheet.cell(row=l,column=column_reorder[i]).value = representsInt(line[i])
	if line[15] == 'HIGH' or line[15] == 'MODERATE' or 'splice_region_variant' in line[14] :
		for i in range(len(line)):
			cell_format(annotationSheet.cell(row=l,column=i+1),color='LightGreen')
	else:
		for i in range(len(line)):
			cell_format(annotationSheet.cell(row=l,column=i+1))		
	if ('+' in line[22] or '-' in line[22] or '*' in line[22]) or line[15] == 'MODIFIER':
		cell_format(annotationSheet.cell(row=l,column=4),color='Blue')
		cell_format(annotationSheet.cell(row=l,column=5),color='Blue')
		cell_format(annotationSheet.cell(row=l,column=7),color='Blue')
		cell_format(annotationSheet.cell(row=l,column=8),color='Blue')

	l = l +1

############
## LAYOUT ##
############
	
#### AUTO-SIZE COLUMNS ####
sns = ['smCounter.anno']
for sn in sns:
	maxsize = 20
	ws = umi_finalreport[sn]
	dims = {}
	for row in ws.rows:
		for cell in row:
			if cell.value:
				try:
					text_size = len(str(cell.value)) + 2
					dims[cell.column] = min(max(dims.get(cell.column, 0), text_size),maxsize)
				except:
					pass
	for col, value in dims.items():
		ws.column_dimensions[col].width = value


###########################
###### SAVE WORKBOOK ######
###########################

umi_finalreport.save(os.path.dirname(sys.argv[1])+'/%s.smCounter.finalReport.xlsx' % sys.argv[1].split('/')[-1].split('.smCounter')[0])

