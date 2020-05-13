#!/usr/bin/env python
# -*- coding: utf-8 -*-
import pycurl, cStringIO, hashlib, json
import openpyxl
import sys
import datetime
import unicodedata 
#import re

def new_record():
	return  {
	'record_id': '',
	'diamic_id': '',
	'last_name': '',
	'first_name': '',
	'dna_id': '',
	'pathology': '', # lung / colon / melanoma / other
	'pathology_other' : '',
	
	'cell_tum_percent': '',
	'conc_dna_stock': '',
	'extract_date': '',
	'origin': '',
	
	'barcode': '',
	'run': '',
	'lib': '',
	'chip': '',
	'run_date': '',
	'techs': '',

	'egfr_ex18___normal': '0',
	'egfr_ex18___mutated': '0',
	'egfr_ex18___amplified': '0',
	'egfr_ex18___verified': '0',
	'egfr_ex18___uninterpretable': '0',
	'egfr_ex18___unconclusive': '0',		
	'egfr_ex18___unrealised': '0',		
	
	'egfr_ex19___normal': '0',
	'egfr_ex19___mutated': '0',
	'egfr_ex19___amplified': '0',
	'egfr_ex19___verified': '0',
	'egfr_ex19___uninterpretable': '0',
	'egfr_ex19___unconclusive': '0',		
	'egfr_ex19___unrealised': '0',	

	'egfr_ex20___normal': '0',
	'egfr_ex20___mutated': '0',
	'egfr_ex20___amplified': '0',
	'egfr_ex20___verified': '0',
	'egfr_ex20___uninterpretable': '0',
	'egfr_ex20___unconclusive': '0',		
	'egfr_ex20___unrealised': '0',	

	'egfr_ex21___normal': '0',
	'egfr_ex21___mutated': '0',
	'egfr_ex21___amplified': '0',
	'egfr_ex21___verified': '0',
	'egfr_ex21___uninterpretable': '0',
	'egfr_ex21___unconclusive': '0',		
	'egfr_ex21___unrealised': '0',	

 	'braf_ex11___mutated': '0',
 	'braf_ex11___normal': '0',	
 	'braf_ex11___uninterpretable': '0',	
 	'braf_ex11___unconclusive': '0',	
 	'braf_ex11___verified': '0',	
 	'braf_ex11___amplified': '0',	
	'braf_ex11___unrealised': '0',	
										
	'braf_ex15___mutated': '0',
	'braf_ex15___normal': '0',
	'braf_ex15___uninterpretable': '0',
	'braf_ex15___unconclusive': '0',
	'braf_ex15___verified': '0',
	'braf_ex15___amplified': '0',
	'braf_ex15___unrealised': '0',

	'kras_ex2___mutated': '0',
	'kras_ex2___normal': '0',
	'kras_ex2___uninterpretable': '0',
	'kras_ex2___unconclusive': '0',
	'kras_ex2___verified': '0',
	'kras_ex2___amplified': '0',
	'kras_ex2___unrealised': '0',

	'kras_ex3___mutated': '0',
	'kras_ex3___normal': '0',
	'kras_ex3___uninterpretable': '0',
	'kras_ex3___unconclusive': '0',
	'kras_ex3___verified': '0',
	'kras_ex3___amplified': '0',
	'kras_ex3___unrealised': '0',

	'kras_ex4___mutated': '0',
	'kras_ex4___normal': '0', 
	'kras_ex4___uninterpretable': '0',
	'kras_ex4___unconclusive': '0',
	'kras_ex4___verified': '0',
 	'kras_ex4___amplified': '0',
	'kras_ex4___unrealised': '0',

	'erbb2_ex20___mutated': '0',
	'erbb2_ex20___normal': '0',
	'erbb2_ex20___uninterpretable': '0',
	'erbb2_ex20___unconclusive': '0',	
	'erbb2_ex20___verified': '0',	
	'erbb2_ex20___amplified': '0',
	'erbb2_ex20___unrealised': '0',
	
	'pik3ca_ex10___mutated': '0',
	'pik3ca_ex10___normal': '0',
	'pik3ca_ex10___uninterpretable': '0',
	'pik3ca_ex10___unconclusive': '0',
	'pik3ca_ex10___verified': '0',
	'pik3ca_ex10___amplified': '0',
	'pik3ca_ex10___unrealised': '0',

	'pik3ca_ex21___mutated': '0',	
	'pik3ca_ex21___normal': '0',	
	'pik3ca_ex21___uninterpretable': '0',	
	'pik3ca_ex21___unconclusive': '0',	
	'pik3ca_ex21___verified': '0',	
	'pik3ca_ex21___amplified': '0',	
	'pik3ca_ex21___unrealised': '0',

	'met___mutated': '0',	
	'met___normal': '0',	
	'met___uninterpretable': '0',	
	'met___unconclusive': '0',	
	'met___amplified': '0',	
	'met___verified': '0',
	'met___unrealised': '0',

 	'stk11___mutated': '0',
 	'stk11___normal': '0',
	'stk11___uninterpretable': '0',
	'stk11___unconclusive': '0',
 	'stk11___verified': '0',
 	'stk11___amplified': '0',
	'stk11___unrealised': '0',
	
	'nras_ex2___normal': '0',
	'nras_ex2___mutated': '0',
	'nras_ex2___amplified': '0',
	'nras_ex2___verified': '0',
	'nras_ex2___uninterpretable': '0',
	'nras_ex2___unconclusive': '0',
	'nras_ex2___unrealised': '0',
	
	'nras_ex3___normal': '0',
	'nras_ex3___mutated': '0',
	'nras_ex3___amplified': '0',
	'nras_ex3___verified': '0',
	'nras_ex3___uninterpretable': '0',
	'nras_ex3___unconclusive': '0',
	'nras_ex3___unrealised': '0',
	
	'nras_ex4___normal': '0',
	'nras_ex4___mutated': '0',
	'nras_ex4___amplified': '0',
	'nras_ex4___verified': '0',
	'nras_ex4___uninterpretable': '0',
	'nras_ex4___unconclusive': '0',
	'nras_ex4___unrealised': '0',

	'kit_ex8___normal': '0',
	'kit_ex8___mutated': '0',
	'kit_ex8___amplified': '0',
	'kit_ex8___verified': '0',
	'kit_ex8___uninterpretable': '0',
	'kit_ex8___unconclusive': '0',
	'kit_ex8___unrealised': '0',

	'kit_ex9___normal': '0',
	'kit_ex9___mutated': '0',
	'kit_ex9___amplified': '0',
	'kit_ex9___verified': '0',
	'kit_ex9___uninterpretable': '0',
	'kit_ex9___unconclusive': '0',
	'kit_ex9___unrealised': '0',
	
	'kit_ex11___normal': '0',
	'kit_ex11___mutated': '0',
	'kit_ex11___amplified': '0',
	'kit_ex11___verified': '0',
	'kit_ex11___uninterpretable': '0',
	'kit_ex11___unconclusive': '0',
	'kit_ex11___unrealised': '0',
	
	'kit_ex13___normal': '0',
	'kit_ex13___mutated': '0',
	'kit_ex13___amplified': '0',
	'kit_ex13___verified': '0',
	'kit_ex13___uninterpretable': '0',
	'kit_ex13___unconclusive': '0',
	'kit_ex13___unrealised': '0',
	
	'alk_ex2___normal': '0',
	'alk_ex2___mutated': '0',
	'alk_ex2___amplified': '0',
	'alk_ex2___verified': '0',
	'alk_ex2___uninterpretable': '0',
	'alk_ex2___unconclusive': '0',
	'alk_ex2___unrealised': '0',
	
	'alk_ex3___normal': '0',
	'alk_ex3___mutated': '0',
	'alk_ex3___amplified': '0',
	'alk_ex3___verified': '0',
	'alk_ex3___uninterpretable': '0',
	'alk_ex3___unconclusive': '0',
	'alk_ex3___unrealised': '0',
	
	'alk_ex4___normal': '0',
	'alk_ex4___mutated': '0',
	'alk_ex4___amplified': '0',
	'alk_ex4___verified': '0',
	'alk_ex4___uninterpretable': '0',
	'alk_ex4___unconclusive': '0',
	'alk_ex4___unrealised': '0',
	
	'hras_ex2___normal': '0',
	'hras_ex2___mutated': '0',
	'hras_ex2___amplified': '0',
	'hras_ex2___verified': '0',
	'hras_ex2___uninterpretable': '0',
	'hras_ex2___unconclusive': '0',
	'hras_ex2___unrealised': '0',
	
	'hras_ex3___normal': '0',
	'hras_ex3___mutated': '0',
	'hras_ex3___amplified': '0',
	'hras_ex3___verified': '0',
	'hras_ex3___uninterpretable': '0',
	'hras_ex3___unconclusive': '0',
	'hras_ex3___unrealised': '0',
	
	'hras_ex4___normal': '0',
	'hras_ex4___mutated': '0',
	'hras_ex4___amplified': '0',
	'hras_ex4___verified': '0',
	'hras_ex4___uninterpretable': '0',
	'hras_ex4___unconclusive': '0',
	'hras_ex4___unrealised': '0',
	
	'map2k1_ex2___normal': '0',
	'map2k1_ex2___mutated': '0',
	'map2k1_ex2___amplified': '0',
	'map2k1_ex2___verified': '0',
	'map2k1_ex2___uninterpretable': '0',
	'map2k1_ex2___unconclusive': '0',
	'map2k1_ex2___unrealised': '0',
	
	'fgfr3_ex7___normal': '0',
	'fgfr3_ex7___mutated': '0',
	'fgfr3_ex7___amplified': '0',
	'fgfr3_ex7___verified': '0',
	'fgfr3_ex7___uninterpretable': '0',
	'fgfr3_ex7___unconclusive': '0',
	'fgfr3_ex7___unrealised': '0',
	
	'fgfr3_ex10___normal': '0',
	'fgfr3_ex10___mutated': '0',
	'fgfr3_ex10___amplified': '0',
	'fgfr3_ex10___verified': '0',
	'fgfr3_ex10___uninterpretable': '0',
	'fgfr3_ex10___unconclusive': '0',
	'fgfr3_ex10___unrealised': '0',
	
	'fgfr3_ex15___normal': '0',
	'fgfr3_ex15___mutated': '0',
	'fgfr3_ex15___amplified': '0',
	'fgfr3_ex15___verified': '0',
	'fgfr3_ex15___uninterpretable': '0',
	'fgfr3_ex15___unconclusive': '0',
	'fgfr3_ex15___unrealised': '0',
	
	'egfr_ex18_variants': '',
	'egfr_ex19_variants': '',
	'egfr_ex20_variants': '',
	'egfr_ex21_variants': '',
 	'braf_ex11_variants': '',
	'braf_ex15_variants': '',
	'kras_ex2_variants': '',
	'kras_ex3_variants': '',
	'kras_ex4_variants': '',	
	'erbb2_ex20_variants': '',
	'pik3ca_ex10_variants': '',
	'pik3ca_ex21_variants': '',
	'met_variants': '',	
	'stk11_variants': '',
	'nras_ex2_variants': '',
	'nras_ex3_variants': '',
	'nras_ex4_variants': '',
	'kit_ex8_variants': '',
	'kit_ex9_variants': '',
	'kit_ex11_variants': '',
	'kit_ex13_variants': '',
	'kit_ex17_variants': '',
	'kit_ex18_variants': '',
	'alk_ex2_variants': '',
	'alk_ex3_variants': '',
	'alk_ex4_variants': '',
	'hras_ex2_variants': '',
	'hras_ex3_variants': '',
	'hras_ex4_variants': '',
	'map2k1_ex2_variants': '',
	'fgfr3_ex7_variants': '',
	'fgfr3_ex10_variants': '',
	'fgfr3_ex15_variants': '',
	
	'comm': '',

	'bilan_complete': '2' # 0 : incomplete, 1 : unverified, 2 : complete. '' = 0
	}

def get_check_cell_data(sheet,r,c):
	cell_data = unicode(sheet.cell(row=r,column=c).value)
	colname = unicode(sheet.cell(row=1,column=c).value)
	request_input = False
	if cell_data == None:
		request_input = True
	elif u'\n' in cell_data:
		request_input = True
	elif u'_' in cell_data:
		request_input = True
	elif u' ' in cell_data:
		request_input = True
	if not (c == 2 or c == 3): # Nom et prenom
		if u'-' in cell_data:
			request_input = True
	
	if request_input:
		colname = unicodedata.normalize('NFKD', colname).encode('ascii','ignore')
		cell_data = unicodedata.normalize('NFKD', cell_data).encode('ascii','ignore')
		ui = raw_input("\t- [%s : %s] -> %s ? :" % (colname,r,cell_data))
		if ui != '':
			cell_data = ui
			
	return cell_data

######################################
# IMPORTING DATA THROUGHT REDCAP API #
######################################

bilan_mutation_path = '/media/stuff/redcap_bilan_mutation/Bilan_mutation_2019.xlsx'
bilan_mutation = openpyxl.load_workbook(bilan_mutation_path)

sheetnames = bilan_mutation.sheetnames
for sheetname in sheetnames:
	print "* %s sheet" % sheetname
	sheet = bilan_mutation[sheetname]
	
	if sheetname in ['NGS Poumon']:
		continue
		pathology = 'lung'
	elif sheetname in ['NGS Colon']:
		continue
		pathology = 'colon'
	elif sheetname in [u'NGS Mélanome']:
		#continue
		pathology = 'melanoma'
	else:
		#continue
		pathology = 'other'
	
	# column name to index
	column2index = {}
	for j in range(1,sheet.max_column+1):
		if not sheet.cell(row=1,column=j).value == None:
			column2index[sheet.cell(row=1,column=j).value.lower()] = j
		
	# parsing sheet
	start = 2
	stop = sheet.max_row+1
	for i in range(start,stop):
		print "- row %s" % i
		
		diamic_id = sheet.cell(row=i,column=1).value
		# si diamic vide, passe.
		if sheet.cell(row=i,column=1).value == None : # avoid empty lines
			print "\t * blank line"
			continue
		diamic_id = get_check_cell_data(sheet,i,1)
		if diamic_id == '*':
			print "\t * ignoring line"
			continue
		
		dna_id = get_check_cell_data(sheet,i,4)
		if dna_id == '*':
			print "\t * ignoring line"
			continue
		
		record_id = diamic_id + '-' + dna_id
		
		last_name = get_check_cell_data(sheet,i,2)
		if last_name == '*':
			print "\t * ignoring line"
			continue
		
		first_name = get_check_cell_data(sheet,i,3)
		if first_name == '*':
			print "\t * ignoring line"
			continue
		
		pathology_other = False
		if pathology == 'other':
			pathology_other = sheet.cell(row=i,column=6).value

		cell_tum_percent = sheet.cell(row=i,column=5).value
		conc_dna_stock = sheet.cell(row=i,column=7).value
		extract_date = sheet.cell(row=i,column=8).value
		# decalage + 1 column avec bilan mut 2019
		## VERSION 2017 2018
		#origin = sheet.cell(row=i,column=column2index['chu / ch / ville']).value
		#barcode = sheet.cell(row=i,column=9).value
		#chip = sheet.cell(row=i,column=10).value
		#run_date = sheet.cell(row=i,column=12).value
		#techs = sheet.cell(row=i,column=13).value
		#lib_run = str(sheet.cell(row=i,column=11).value)
		## VERSION 2019
		origin = sheet.cell(row=i,column=9).value
		barcode = sheet.cell(row=i,column=10).value
		chip = sheet.cell(row=i,column=11).value
		run_date = sheet.cell(row=i,column=13).value
		techs = sheet.cell(row=i,column=14).value
		lib_run = str(sheet.cell(row=i,column=12).value)
		if '_' in lib_run:
			lib_run = lib_run.split('_')
			lib = lib_run[0]
			run = lib_run[1]
		else:
			lib = ''
			run = lib_run
		
		# NEW RECORD #
		record = new_record()
		record['diamic_id'] = diamic_id
		record['last_name'] = last_name	
		record['first_name'] = first_name
		record['dna_id'] = dna_id
		record['record_id'] = record_id
		record['pathology'] = pathology
		if pathology_other:
			record['pathology_other'] = pathology_other
		record['cell_tum_percent'] = cell_tum_percent
		record['conc_dna_stock'] = conc_dna_stock
		if isinstance(extract_date, datetime.datetime):
			extract_date = extract_date.strftime("%d/%m/%Y")
		record['extract_date'] = extract_date
		record['origin'] = origin
		record['barcode'] = barcode
		record['run'] = run
		record['lib'] = lib
		record['chip'] = chip
		if isinstance(run_date, datetime.datetime):
			run_date = run_date.strftime("%d/%m/%Y")
		record['run_date'] = run_date
		record['techs'] = techs
		
		record['comm'] = sheet.cell(row=i,column=column2index['commentaire']).value

		# GET VARIANT DATA AND STUFF

		if pathology == 'lung':
			loci = ['egfr_ex18','egfr_ex19','egfr_ex20','egfr_ex21','braf_ex11','braf_ex15','kras_ex2','kras_ex3','kras_ex4','erbb2_ex20','pik3ca_ex10','pik3ca_ex21','met'] # 'stk11'
		elif pathology == 'colon':
			loci = ['braf_ex11','braf_ex15','kras_ex2','kras_ex3','kras_ex4','nras_ex2','nras_ex3','nras_ex4','pik3ca_ex10','pik3ca_ex21']
		elif pathology == 'melanoma':
			loci = ['braf_ex11','braf_ex15','kit_ex8','kit_ex9','kit_ex11','kit_ex13','kit_ex17','kit_ex18','nras_ex2','nras_ex3','nras_ex4']
		elif pathology == 'other':
			loci = ['MET (exons 2 et 14 a 20)','BRAF (exons 11 et 15)','ALK (exons 22 a 25)','KIT (exons 8, 9, 11, 13, 17 et 18)','KRAS (exons 2, 3, 4)','HRAS (exons 2, 3, 4)','NRAS (exons 2, 3, 4)','MAP2K1 (exon 2)','FGFR3 (exons 7, 10 et 15)','EGFR']
			
		for locus in loci:
			try:
				l = locus.replace('_',' ')
				l = l.lower()
				locus_data = sheet.cell(row=i,column=column2index[l]).value
			except Exception as e:
				print "**WARNING [%s] pathology locus not found : %s on line %s : %s" % (sys.argv[0].split('/')[-1],type(e).__name__,sys.exc_info()[-1].tb_lineno,e)
			if locus_data == None:
				continue
			locus_data_splitted = locus_data.split('\n')
			if not pathology == 'other' :
				if 'N' in locus_data_splitted:
					record['%s___normal' % locus] = '1'
				if 'NI' in locus_data_splitted:
					record['%s___uninterpretable' % locus] = '1'
				if 'NC' in locus_data_splitted:
					record['%s___unconclusive' % locus] = '1'
				if 'NR' in locus_data_splitted:
					record['%s___unrealised' % locus] = '1'
				for ver in [u'vérif','verif']:
					if ver in locus_data.lower():
						record['%s___verified' % locus] = '1'
						break
				for amp in ['amplification','ampli']:
					if amp in locus_data.lower():
						record['%s___amplified' % locus] = '1'
						break
				for mutdesc in ['c.','p.']:	
					if mutdesc in locus_data.lower():
						record['%s___mutated' % locus] = '1'
						record['%s_variants' % locus] = locus_data						
			else: # if pathology == 'other':
				other_loci = []
				if 'MET' in locus:
					other_loci.append('met')
				elif 'BRAF' in locus:
					if 'exon 11' in locus_data.lower():
						other_loci.append('braf_ex11')
					if 'exon 15' in locus_data.lower():
						other_loci.append('braf_ex15')
					if not 'exon' in locus_data.lower():
						other_loci = ['braf_ex11','braf_ex15']
				elif 'ALK' in locus :
					other_loci = ['alk_ex2','alk_ex3','alk_ex4']
				elif 'KIT' in locus:
					if 'exon 8' in locus_data.lower():
						other_loci.append('kit_ex8')
					if 'exon 9' in locus_data.lower():
						other_loci.append('kit_ex9')
					if 'exon 11' in locus_data.lower():
						other_loci.append('kit_ex11')
					if 'exon 13' in locus_data.lower():
						other_loci.append('kit_ex13')
					if 'exon 17' in locus_data.lower():
						other_loci.append('kit_ex17')
					if 'exon 18' in locus_data.lower():
						other_loci.append('kit_ex18')
					if not 'exon' in locus_data.lower():
						other_loci = ['kit_ex8','kit_ex9','kit_ex11','kit_ex13','kit_ex17','kit_ex18']
				elif 'KRAS' in locus:
					if 'exon 2' in locus_data.lower():
						other_loci.append('kras_ex2')
					if 'exon 3' in locus_data.lower():
						other_loci.append('kras_ex3')
					if 'exon 3' in locus_data.lower():
						other_loci.append('kras_ex3')
					if not 'exon' in locus_data.lower():
						other_loci = ['kras_ex2','kras_ex3','kras_ex4']
				elif 'HRAS' in locus:
					if 'exon 2' in locus_data.lower():
						other_loci.append('hras_ex2')
					if 'exon 3' in locus_data.lower():
						other_loci.append('hras_ex3')
					if 'exon 4' in locus_data.lower():
						other_loci.append('hras_ex4')
					if not 'exon' in locus_data.lower():
						other_loci = ['hras_ex2','hras_ex3','hras_ex4']
				elif 'NRAS' in locus:
					if 'exon 2' in locus_data.lower():
						other_loci.append('nras_ex2')
					if 'exon 3' in locus_data.lower():
						other_loci.append('nras_ex3')
					if 'exon 4' in locus_data.lower():
						other_loci.append('nras_ex4')
					if not 'exon' in locus_data.lower():
						other_loci = ['nras_ex2','nras_ex3','nras_ex4']
				elif 'MAP2K1' in locus:
					other_loci.append('map2k1_ex2')
				elif 'FGFR3' in locus:
					if 'exon 7' in locus_data.lower():
						other_loci.append('fgfr3_ex7')
					if 'exon 10' in locus_data.lower():
						other_loci.append('fgfr3_ex10')
					if 'exon 15' in locus_data.lower():
						other_loci.append('fgfr3_ex15')
					if not 'exon' in locus_data.lower():
						other_loci = ['fgfr3_ex7','fgfr3_ex10','fgfr3_ex15']
				elif 'EGFR' in locus:
					other_loci = ['egfr_ex18','egfr_ex19','egfr_ex20','egfr_ex21']
				####
				for other_locus in other_loci:	
					if 'N' in locus_data_splitted:
						record['%s___normal' % other_locus] = '1'
					if 'NI' in locus_data_splitted:
						record['%s___uninterpretable' % other_locus] = '1'
					if 'NC' in locus_data_splitted:
						record['%s___unconclusive' % other_locus] = '1'
					if 'NR' in locus_data_splitted:
						record['%s___unrealised' % other_locus] = '1'
					for ver in [u'vérif','verif']:
						if ver in locus_data.lower():
							record['%s___verified' % other_locus] = '1'
							break
					for amp in ['amplification','ampli',u'amplifié']:
						if amp in locus_data.lower():
							record['%s___amplified' % other_locus] = '1'
							break
					for mutdesc in ['c.','p.']:	
						if mutdesc in locus_data.lower():
							record['%s___mutated' % other_locus] = '1'
							record['%s_variants' % other_locus] = locus_data
				
		######################################
		# IMPORTING DATA THROUGHT REDCAP API #
		######################################

		record_data = json.dumps([record],default=str)

		buf = cStringIO.StringIO()
		token = '262A3D5B836F770939C9C7CBF8844D2D'
		redcap_ip = '10.67.1.22'

		data = {
			'token': 'B7135B24799ECCCF7B80903AA3D7202A',
			'content': 'record',
			'format': 'json',
			'type': 'flat',
			'overwriteBehavior': 'normal',
			'forceAutoNumber': 'false',
			'data': record_data,
			'dateFormat': 'DMY',
			'returnContent': 'count',
			'returnFormat': 'json',
			'record_id': record_id
		}

		ch = pycurl.Curl()
		ch.setopt(ch.URL, 'http://%s/redcap/api/' % redcap_ip)
		ch.setopt(ch.HTTPPOST, data.items())
		ch.setopt(ch.WRITEFUNCTION, buf.write)
		ch.perform()
		ch.close()

		print buf.getvalue()
		buf.close()
