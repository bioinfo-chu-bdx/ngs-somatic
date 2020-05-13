#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import json
import openpyxl
import copy
import glob
import time
from optparse import OptionParser

def cell_format(cell, font=None, alignment=None, color=None, format=None, border=None):
	if font == 'bold':
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, bold=True)
	elif font == 'red':
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, color='f44242')
	else:
		cell.font = openpyxl.styles.Font(name='Calibri', size=11)
	if alignment == 'center':
		cell.alignment = openpyxl.styles.Alignment(horizontal='center')
	elif alignment == 'vertical-center':
		cell.alignment = openpyxl.styles.Alignment(vertical='center')
	elif alignment == 'left':
		cell.alignment = openpyxl.styles.Alignment(horizontal='left')
	if color == 'LightGreen':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='D8E4BC') # EBF1DE
	elif color == 'LightRed':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='d28e8e') #F2DCDB
	elif color == 'LightOrange':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='ffe7e0') # ffd8ad
	elif color == 'LightBlue':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='add8e6')
	elif color == 'LightPink':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='ffc9f7')
	elif color == 'Yellow':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='feffa3')
	elif color == 'Blue':
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, color='004c99')
	else:
		cell.fill = openpyxl.styles.PatternFill(fill_type=None,start_color='FFFFFF')
	if border == 'thin':
		cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'),bottom=openpyxl.styles.Side(style='thin'))
	if border == 'top_medium':
		cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='medium'))
	if format == 'Percent':
		cell.number_format = '0.0%'

#########################################################################################

parser = OptionParser()
parser.add_option('-b', '--run-folder', 		help="run_folder", 	dest='run_folder')
(options, args) = parser.parse_args()

run_folder = options.run_folder
bamlist = glob.glob(run_folder+'/*/*.bam')
bamlist = [item for item in bamlist if not 'processed' in item]

with open('/DATA/work/global_parameters.json', 'r') as g:
	global_param = json.load(g)	
with open(run_folder+'/barcodes.json', 'r') as g:
	barcodes_json = json.load(g)

for bamfile in sorted(bamlist) :	
	sample = bamfile.split('/')[-1].split('_IonXpress')[0]
	barcode = 'IonXpress_' + bamfile.split('IonXpress_')[-1].split('.bam')[0]
	sample_folder = os.path.dirname(bamfile)

	bed_name = barcodes_json[barcode]['target_region_filepath'].split('/unmerged/detail/')[-1]
	run_type = False
	for _run_type in global_param['run_type']:
		if global_param['run_type'][_run_type]['target_bed'].split('/')[-1] == bed_name:
			run_type = _run_type
			break
	
	finalReport_path = sample_folder+'/'+sample+'_'+barcode+'.finalReport.xlsx'
	finalReport = openpyxl.load_workbook(finalReport_path)
	cnvSheet = finalReport['CNV']

	#########
	## CNV ##
	#########

	cna = '%s/_CNA/%s/CNV_finalReport.xlsx' % (run_folder,run_type)

	print " [%s] Replacing CNV sheet in progress ..." % time.strftime("%H:%M:%S")
	inBook = openpyxl.load_workbook(cna)
	inSheet = inBook['copy number analysis']
		
	for row_idx in range(1, inSheet.max_row+1):
		for col_idx in range(1, inSheet.max_column+1):
			read_cell = inSheet.cell(row = row_idx, column = col_idx)
			cnvSheet.cell(row=row_idx,column=col_idx).value = read_cell.value
			if read_cell.has_style:
				cnvSheet.cell(row=row_idx,column=col_idx).font = copy.copy(read_cell.font)
				cnvSheet.cell(row=row_idx,column=col_idx).border = copy.copy(read_cell.border)
				cnvSheet.cell(row=row_idx,column=col_idx).fill = copy.copy(read_cell.fill)
				cnvSheet.cell(row=row_idx,column=col_idx).alignment = copy.copy(read_cell.alignment)

	for row_idx in range(1, inSheet.max_row+1):
		c = cnvSheet.cell(row=row_idx,column=1)
		if c.value == sample:
			cell_format(c,color='Yellow')
			break

	finalReport.save(finalReport_path)
