#!/usr/bin/env python
import sys
import openpyxl

def representsInt(s): # pour eviter avertissement "nombre ecrit en texte" sous excel
	try: 
		s = int(s)
		return s
	except ValueError:
		return s
		
def cell_format(cell, font=None, alignment=None, color=None, format=None, border=None):
	if font == 'bold':
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, bold=True)
	else:
		cell.font = openpyxl.styles.Font(name='Calibri', size=11)
	if alignment == 'center':
		cell.alignment = openpyxl.styles.Alignment(horizontal='center')
	elif alignment == 'left':
		cell.alignment = openpyxl.styles.Alignment(horizontal='left')
	if color == 'LightGreen':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='D8E4BC') # EBF1DE
	elif color == 'LightRed':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='d28e8e') #F2DCDB
	elif color == 'Yellow':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='feffa3')
	elif color == 'Blue':
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, color='004c99')
	else:
		cell.fill = openpyxl.styles.PatternFill(fill_type=None,start_color='FFFFFF')
	if border:
		cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'),bottom=openpyxl.styles.Side(style='thin'))
	if format == 'Percent':
		cell.number_format = '0.0%'

###############################################################################

suivi_met_path = "/media/n06lbth/sauvegardes_pgm/SBT/Mutations_intron_13-14_MET.xlsx"
finalreport_path = sys.argv[1]
sample = sys.argv[2]
run_name = sys.argv[3]

###############################################################################

fp = openpyxl.load_workbook(finalreport_path)
annotation_sheet = fp.get_sheet_by_name('Annotation')
annotation_rows = tuple(annotation_sheet.rows)

suivi_met = openpyxl.load_workbook(suivi_met_path)
suivi_sheet = suivi_met.get_sheet_by_name('MET_intron_13-14')
suivi_rows = tuple(suivi_sheet.rows)

line2write = len(suivi_rows)+1

# variants lines
for i in range(len(annotation_rows[0])):
	if annotation_rows[0][i].value == 'Gene':
		gene_index = i
	if annotation_rows[0][i].value == 'Region':
		region_index = i
	if annotation_rows[0][i].value == 'Start.Pos':
		pos_index = i

for i in range(1,len(annotation_rows)):
	if annotation_rows[i][gene_index].value == 'MET' and annotation_rows[i][region_index].value == 'intronic' and ((116411808 <= int(annotation_rows[i][pos_index].value) <= 116411902) or (116412044 <= int(annotation_rows[i][pos_index].value) <= 116412087)):
		line2write = line2write + 1
		suivi_sheet.cell(row=line2write,column=1).value = run_name
		suivi_sheet.cell(row=line2write,column=2).value = sample
		suivi_sheet.cell(row=line2write,column=1).font = openpyxl.styles.Font(name='Calibri', size=11)
		suivi_sheet.cell(row=line2write,column=2).font = openpyxl.styles.Font(name='Calibri', size=11)
		for j in range(1,len(annotation_rows[i])):
			suivi_sheet.cell(row=line2write,column=j+2).value = annotation_rows[i][j].value
			suivi_sheet.cell(row=line2write,column=j+2).font = openpyxl.styles.Font(name='Calibri', size=11)
		
suivi_met.save(suivi_met_path)
