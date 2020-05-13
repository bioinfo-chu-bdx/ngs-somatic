#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import json
import xlrd
import glob
import time
import openpyxl
import subprocess
from optparse import OptionParser

def cell2str(data):
	if data == None or data == '':
		return '.'
	elif isinstance(data, (int, long, float)) :
		return str(data)
	else:
		return data
    
def getValueWithMergeLookup(sheet, row, column):
    cellval = sheet.cell(row=row,column=column).value
    while cellval == None:
		row = row - 1
		cellval = sheet.cell(row=row,column=column).value
    return cellval

### GATHERING PARAMETERS ############################################################

parser = OptionParser()
parser.add_option('-r', '--run-folder', help="full run path",				dest='run_folder')
parser.add_option('-w', '--sbt-wait', 	help="(optional) if run SBT, exit",	dest='sbt_wait', action='store_true')
(options, args) = parser.parse_args()

with open('/DATA/work/global_parameters.json', 'r') as g:
	global_param = json.load(g)
	
with open(options.run_folder+'/barcodes.json', 'r') as g:
	barcodes_json = json.load(g)

bilan_mut = xlrd.open_workbook(global_param['bilan_mut_path'])

header = ['Run_name','Sample','Chr','Transcript','Gene','Exon','Intron','c.','p.','Start.Pos','Ref.Seq','Var.Seq','Class.','Var.Freq.',
'Var.Cov.','Pos.Cov.','Region','Type','Consequence','Impact','COSMIC','dbSNP','InterVar','ClinVar.significance','ClinVar.disease',
'NCI60','ESP','1000G_ALL','1000G_EUR','1000G_AMR','1000G_AFR','1000G_EAS','1000G_SAS','SIFT','POLYPHEN2_HVAR','Comm.Bio.']

#####################################################################################

finalreports = []
if options.run_folder:
	if os.path.isdir(options.run_folder+'/_2018_reanalysis'):
		finalreports = glob.glob(options.run_folder+'/_2018_reanalysis/*/*finalReport.xlsx')
	else:
		finalreports = glob.glob(options.run_folder+'/*/*finalReport.xlsx')
	finalreports = [item for item in finalreports if not '~' in item]
	path_split = options.run_folder.split('/')
	if path_split[-1] == '':
		run_name = path_split[-2]
	else:
		run_name = path_split[-1]
	print "- Processing run %s ..." % run_name
else:
	print "Error - No run specified"
	exit()

sample2variantlist = {}
sample2finalreport = {}
for report in finalreports:
	sample = report.split('/')[-1].split('_IonXpress')[0]
	sample2finalreport[sample] = report
	barcode = 'IonXpress_' + report.split('IonXpress_')[-1].split('.finalReport')[0]
	target = barcodes_json[barcode]['target_region_filepath'].split('/')[-1]
	for _run_type in global_param['run_type']:
		if global_param['run_type'][_run_type]['target_bed'].split('/')[-1] == target:
			sample2variantlist[sample] = '/DATA/work/scripts/tests/unique_variantList_%s' % run_name
			break

variants_seen = []
for variantlist in list(set(sample2variantlist.values())):	
	unique_VariantList = open(variantlist,'w')

	header_string = '\t'.join(header)
	unique_VariantList.write(header_string+'\n')
	
	sampleset = [key for key in sample2variantlist if sample2variantlist[key] == variantlist]
	for sample in sampleset:
		print "\t - %s" % sample
		finalreport = openpyxl.load_workbook(sample2finalreport[sample])
		annotation_sheet = finalreport['Annotation']
		if 'VEP' in finalreport.sheetnames:
			vep_sheet = finalreport['VEP']

		# column name to index
		column2index = {}
		for j in range(1,annotation_sheet.max_column+1):
			column2index[annotation_sheet.cell(row=1,column=j).value] = j

		for i in range(2,annotation_sheet.max_row+1):
			if annotation_sheet.cell(row=i,column=column2index['Chr']).value == None: # avoid empty line and "Amplicons < 300X: " line
				continue
			
			Transcript = cell2str(annotation_sheet.cell(row=i,column=column2index['Transcript']).value)
			c_nomen = cell2str(annotation_sheet.cell(row=i,column=column2index['c.']).value)
			Region = cell2str(annotation_sheet.cell(row=i,column=column2index['Region']).value)
			Type = cell2str(annotation_sheet.cell(row=i,column=column2index['Type']).value)
			
			if (Transcript,c_nomen) not in variants_seen:
				variants_seen.append((Transcript,c_nomen))
			else:
				continue
				
			if ((Region in ['?','intronic','UTR3','UTR5','ncRNA_intronic']) or (Type == 'synonymous')):
				continue
			
			Comm = cell2str(annotation_sheet.cell(row=i,column=column2index['Comm.Bio.']).value)
			Comm = Comm.replace(u'é','e').replace(u'è','e').replace(u'ê','e').replace(u'à','a')
			Chr = cell2str(annotation_sheet.cell(row=i,column=column2index['Chr']).value)
			
			Gene = cell2str(annotation_sheet.cell(row=i,column=column2index['Gene']).value)
			Exon = cell2str(annotation_sheet.cell(row=i,column=column2index['Exon']).value)
			
			p_nomen = cell2str(annotation_sheet.cell(row=i,column=column2index['p.']).value)
			Start_Pos = cell2str(annotation_sheet.cell(row=i,column=column2index['Start.Pos']).value)
			Ref_Seq = cell2str(annotation_sheet.cell(row=i,column=column2index['Ref.Seq']).value)
			Var_Seq = cell2str(annotation_sheet.cell(row=i,column=column2index['Var.Seq']).value)
			Class = cell2str(annotation_sheet.cell(row=i,column=column2index['Class.']).value)
			Var_Freq = cell2str(annotation_sheet.cell(row=i,column=column2index['Var.Freq.']).value)
			Var_Cov = cell2str(annotation_sheet.cell(row=i,column=column2index['Var.Cov.']).value)
			Pos_Cov = cell2str(annotation_sheet.cell(row=i,column=column2index['Pos.Cov.']).value)
			COSMIC = cell2str(annotation_sheet.cell(row=i,column=column2index['COSMIC']).value)
			dbSNP = cell2str(annotation_sheet.cell(row=i,column=column2index['dbSNP']).value)
			InterVar = cell2str(annotation_sheet.cell(row=i,column=column2index['InterVar']).value)
			ClinVar_significance = cell2str(annotation_sheet.cell(row=i,column=column2index['ClinVar.significance']).value)
			ClinVar_disease = cell2str(annotation_sheet.cell(row=i,column=column2index['ClinVar.disease']).value)
			NCI60 = cell2str(annotation_sheet.cell(row=i,column=column2index['NCI60']).value)
			ESP = cell2str(annotation_sheet.cell(row=i,column=column2index['ESP']).value)
			_1000G_ALL = cell2str(annotation_sheet.cell(row=i,column=column2index['1000G_ALL']).value)
			_1000G_EUR = cell2str(annotation_sheet.cell(row=i,column=column2index['1000G_EUR']).value)
			_1000G_AMR = cell2str(annotation_sheet.cell(row=i,column=column2index['1000G_AMR']).value)
			_1000G_AFR = cell2str(annotation_sheet.cell(row=i,column=column2index['1000G_AFR']).value)
			_1000G_EAS = cell2str(annotation_sheet.cell(row=i,column=column2index['1000G_EAS']).value)
			_1000G_SAS = cell2str(annotation_sheet.cell(row=i,column=column2index['1000G_SAS']).value)
			SIFT = cell2str(annotation_sheet.cell(row=i,column=column2index['SIFT']).value)
			POLYPHEN2_HVAR = cell2str(annotation_sheet.cell(row=i,column=column2index['POLYPHEN2_HVAR']).value)
			
			vep_intron = '.'
			vep_consequence = '.'
			vep_impact = '.'
			if 'VEP' in finalreport.sheetnames:
				for v in range(2,vep_sheet.max_row+1):
					vep_transcript = cell2str(vep_sheet.cell(row=v,column=3).value.split('.')[0])
					vep_pos = cell2str(getValueWithMergeLookup(vep_sheet,v,15))
					vep_ref = cell2str(getValueWithMergeLookup(vep_sheet,v,16))
					vep_alt = cell2str(getValueWithMergeLookup(vep_sheet,v,17))
					if (vep_transcript,vep_pos,vep_ref,vep_alt) == (Transcript,Start_Pos,Ref_Seq,Var_Seq):
						vep_intron = cell2str(vep_sheet.cell(row=v,column=6).value)
						vep_consequence = cell2str(vep_sheet.cell(row=v,column=9).value)
						vep_impact = cell2str(vep_sheet.cell(row=v,column=10).value)
						break
			
			variant = [run_name,sample,Chr,Transcript,Gene,Exon,vep_intron,c_nomen,p_nomen,Start_Pos,Ref_Seq,Var_Seq,Class,Var_Freq,Var_Cov,Pos_Cov,Region,Type,vep_consequence,vep_impact,COSMIC,dbSNP,InterVar,ClinVar_significance,ClinVar_disease,NCI60,ESP,_1000G_ALL,_1000G_EUR,_1000G_AMR,_1000G_AFR,_1000G_EAS,_1000G_SAS,SIFT,POLYPHEN2_HVAR,Comm]
			
			variant_string = '\t'.join(variant)
			unique_VariantList.write(variant_string+'\n')

	unique_VariantList.close()
