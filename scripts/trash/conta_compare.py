#!/usr/bin/env python
import os
import openpyxl
import copy

contareport_paths = [
'/media/n06lbth/sauvegardes_pgm/SBT/Run_300-399/Auto_user_S5-0198-185-Run319_320_35pM_Chef_SBT-colon-lung_v8_530_354/_Check-contamination/Check-contamination_H2O_lib1_NGS319-.xlsx',
'/media/n06lbth/sauvegardes_pgm/SBT/Run_300-399/Auto_user_S5-0198-185-Run319_320_35pM_Chef_SBT-colon-lung_v8_530_354/_Check-contamination/Check-contamination_H2O_lib2_NGS320-.xlsx',
'/media/n06lbth/sauvegardes_pgm/SBT/Run_300-399/Auto_user_S5-0198-185-Run319_320_35pM_Chef_SBT-colon-lung_v8_530_354/_Check-contamination/Check-contamination_NTC_lib1_NGS319-.xlsx',
'/media/n06lbth/sauvegardes_pgm/SBT/Run_300-399/Auto_user_S5-0198-185-Run319_320_35pM_Chef_SBT-colon-lung_v8_530_354/_Check-contamination/Check-contamination_NTC_lib2_NGS320-.xlsx',
'/media/n06lbth/sauvegardes_pgm/SBT/Run_300-399/Auto_user_S5-0198-186-NGS321_322_35pM_Chef_SBT-colon-lung_v8_530_356/_Check-contamination/Check-contamination_H2O_lib1_NGS321-.xlsx',
'/media/n06lbth/sauvegardes_pgm/SBT/Run_300-399/Auto_user_S5-0198-186-NGS321_322_35pM_Chef_SBT-colon-lung_v8_530_356/_Check-contamination/Check-contamination_H2O_lib2_NGS322-.xlsx',
'/media/n06lbth/sauvegardes_pgm/SBT/Run_300-399/Auto_user_S5-0198-186-NGS321_322_35pM_Chef_SBT-colon-lung_v8_530_356/_Check-contamination/Check-contamination_NTC_lib1_NGS321-.xlsx',
'/media/n06lbth/sauvegardes_pgm/SBT/Run_300-399/Auto_user_S5-0198-186-NGS321_322_35pM_Chef_SBT-colon-lung_v8_530_356/_Check-contamination/Check-contamination_NTC_lib2_NGS322-.xlsx',
'/media/n06lbth/sauvegardes_pgm/SBT/Auto_user_S5-0198-188-run323_324_35pM_Chef_SBT-colon-lung_v8_530_359/_Check-contamination/Check-contamination_H2O_lib1_NGS323-.xlsx',
'/media/n06lbth/sauvegardes_pgm/SBT/Auto_user_S5-0198-188-run323_324_35pM_Chef_SBT-colon-lung_v8_530_359/_Check-contamination/Check-contamination_H2O_lib2_NGS324-.xlsx',
'/media/n06lbth/sauvegardes_pgm/SBT/Auto_user_S5-0198-188-run323_324_35pM_Chef_SBT-colon-lung_v8_530_359/_Check-contamination/Check-contamination_NTC_lib1_NGS323-.xlsx',
'/media/n06lbth/sauvegardes_pgm/SBT/Auto_user_S5-0198-188-run323_324_35pM_Chef_SBT-colon-lung_v8_530_359/_Check-contamination/Check-contamination_NTC_lib2_NGS324-.xlsx',
'/media/n06lbth/sauvegardes_pgm/SBT/Auto_user_S5-0198-192-Run324bis_35pM_Chef_SBT-colon-lung_v8_530_364/_Check-contamination/Check-contamination_H2O_lib2_NGS324bis-.xlsx',
'/media/n06lbth/sauvegardes_pgm/SBT/Auto_user_S5-0198-192-Run324bis_35pM_Chef_SBT-colon-lung_v8_530_364/_Check-contamination/Check-contamination_NTC_lib2_NGS324bis-.xlsx',
'/media/n06lbth/sauvegardes_pgm/SBT/Auto_user_S5-0198-193-Run325_326_35pM_Chef_SBT-colon-lung_v8_530_365/_Check-contamination/Check-contamination_H2O_lib1_NGS325-.xlsx',
'/media/n06lbth/sauvegardes_pgm/SBT/Auto_user_S5-0198-193-Run325_326_35pM_Chef_SBT-colon-lung_v8_530_365/_Check-contamination/Check-contamination_H2O_lib2_NGS326-.xlsx',
'/media/n06lbth/sauvegardes_pgm/SBT/Auto_user_S5-0198-193-Run325_326_35pM_Chef_SBT-colon-lung_v8_530_365/_Check-contamination/Check-contamination_NTC_lib1_NGS325-.xlsx',
'/media/n06lbth/sauvegardes_pgm/SBT/Auto_user_S5-0198-193-Run325_326_35pM_Chef_SBT-colon-lung_v8_530_365/_Check-contamination/Check-contamination_NTC_lib2_NGS326-.xlsx'
]

variants_seen = {}
results = openpyxl.Workbook()
res_sheet = results.get_sheet_by_name('Sheet')

for contareport_path in contareport_paths:
	control_name = contareport_path.split('Check-contamination_')[-1].split('-.xlsx')[0]
	contareport = openpyxl.load_workbook(contareport_path)
	annoSheet = contareport.get_sheet_by_name('Annotation')
	annoSheet_rows = tuple(annoSheet.rows)
	for i in range(1,len(annoSheet_rows)):
		nm = annoSheet_rows[i][3].value
		c = annoSheet_rows[i][6].value
		if not nm:
			continue
		if (nm,c) in variants_seen:
			variants_seen[(nm,c)][0] = variants_seen[(nm,c)][0] + ',' + control_name # add name
			variants_seen[(nm,c)][8] = variants_seen[(nm,c)][8] + ',' + str(annoSheet_rows[i][8].value) # freq
			variants_seen[(nm,c)][10] = variants_seen[(nm,c)][10] + ',' + str(annoSheet_rows[i][10].value) # var cov
			variants_seen[(nm,c)][11] = variants_seen[(nm,c)][11] + ',' + str(annoSheet_rows[i][11].value) # pos cov
		else:
			#print annoSheet_rows[i][8].value
			variants_seen[(nm,c)] = [control_name,annoSheet_rows[i][1].value,annoSheet_rows[i][2].value,annoSheet_rows[i][3].value,annoSheet_rows[i][4].value,annoSheet_rows[i][5].value,annoSheet_rows[i][6].value,annoSheet_rows[i][7].value,str(annoSheet_rows[i][8].value),annoSheet_rows[i][9].value,str(annoSheet_rows[i][10].value),str(annoSheet_rows[i][11].value),annoSheet_rows[i][12].value,annoSheet_rows[i][13].value,annoSheet_rows[i][14].value,annoSheet_rows[i][15].value,annoSheet_rows[i][16].value,annoSheet_rows[i][17].value,annoSheet_rows[i][18].value,annoSheet_rows[i][19].value,annoSheet_rows[i][20].value,annoSheet_rows[i][21].value]

row_num = 0			
for variant in variants_seen:
	row_num = row_num + 1
	for i in range(len(variants_seen[variant])):
		res_sheet.cell(row=row_num,column=i+1).value = variants_seen[variant][i]
		
results.save('/media/stuff/conta_compare.xlsx')
