#!/usr/bin/python
import sys
import os
import glob
import json
import openpyxl
import csv
from optparse import OptionParser

### GATHERING PARAMETERS ############################################################

parser = OptionParser()
parser.add_option('-b', '--bam',		help="Input bam file for SINGLE BAM ANALYSIS (must contain flow signal from TSS)", dest='bam')
parser.add_option('-r', '--full-run',	help="Run folder path  for FULL RUN ANALYSIS", dest='full_run') 
parser.add_option('-t', '--run-type',	help="Run type : SBT, LAM, TP53-HEMATO... (OPTIONAL, NOT NEEDED if barcodes.json is present in run folder)", dest='run_type') 
(options, args) = parser.parse_args()

if options.bam and options.full_run:
	sys.stderr.write("[run_analysis.py] Error: <--bam> and <--full-run> are not compatibles\n")
	sys.exit()
if options.full_run:
	bamlist = glob.glob(options.full_run+'/*/*.bam')
	bamlist = [item for item in bamlist if not 'processed' in item]
elif options.bam:
	bamlist = [options.bam]
else:
	sys.stderr.write("[run_analysis.py] Error: no <--bam> or <--full-run> specified\n")
	sys.exit()
	
with open('/DATA/work/global_parameters.json', 'r') as g:
	global_param = json.load(g)

temoin_variants = {}
temoinCNV_file = open('/DATA/work/scripts/temoinCNV_variants_lymphomeT.tsv','r')
temoinCNV_reader = csv.reader(temoinCNV_file,delimiter = '\t')
for line in temoinCNV_reader:
	temoin_variants[(line[0],line[1])] = line[2]

### RUN ANALYSIS ######################################################################

bam_data = {}
	
for bamfile in sorted(bamlist) :
	sample = bamfile.split('/')[-1].split('_IonXpress')[0]
	print "- %s" % sample
	barcode = 'IonXpress_' + bamfile.split('IonXpress_')[-1].split('.bam')[0]
	sample_folder = os.path.dirname(bamfile)
	run_folder = os.path.dirname(sample_folder)
	
	finalreport_path = sample_folder + '/' + sample+'_'+barcode+'.finalReport.xlsx'
	finalreport = openpyxl.load_workbook(finalreport_path)
	annoSheet = finalreport.get_sheet_by_name('Annotation')
	
	for i in range(2,annoSheet.max_row+1):
		if annoSheet.cell(row=i,column=2).value == None: # avoid empty line and "Amplicons < 300X: " line
			continue
		nm = annoSheet.cell(row=i,column=3).value
		c = annoSheet.cell(row=i,column=6).value
		if (nm,c) in temoin_variants:
			if annoSheet.cell(row=i,column=1).value:
				annoSheet.cell(row=i,column=1).value = 'Found in Control ' + temoin_variants[(nm,c)] + ',' + annoSheet.cell(row=i,column=1).value
			else:
				annoSheet.cell(row=i,column=1).value = 'Found in Control ' + temoin_variants[(nm,c)]
		#print annoSheet.cell(row=i,column=1).value

	finalreport.save(finalreport_path)
