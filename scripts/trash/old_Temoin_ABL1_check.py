#!/usr/bin/env python
import sys
import os
import openpyxl
import subprocess

def representsInt(s): # pour eviter avertissement "nombre ecrit en texte" sous excel
	try: 
		s = int(s)
		return s
	except ValueError:
		return s
		
def cell_format(cell, font=None, alignment=None, color=None, format=None, border=None,exterior_border=None):
	if font == 'bold':
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, bold=True)
	else:
		cell.font = openpyxl.styles.Font(name='Calibri', size=11)
	if alignment == 'center':
		cell.alignment = openpyxl.styles.Alignment(horizontal='center',wrap_text=True)
	elif alignment == 'left':
		cell.alignment = openpyxl.styles.Alignment(horizontal='left',wrap_text=True)
	if color == 'LightGreen':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='D8E4BC') # EBF1DE
	elif color == 'LightRed':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='d28e8e') #F2DCDB
	elif color == 'LightBlue':	
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='add8e6')
	elif color == 'Yellow':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='feffa3')
	elif color == 'Blue':
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, color='004c99')
	else:
		cell.fill = openpyxl.styles.PatternFill(fill_type=None,start_color='FFFFFF')
	if border:
		cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'),bottom=openpyxl.styles.Side(style='thin'))
	if exterior_border:
		cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='thin'),bottom=openpyxl.styles.Side(style='thin'))
	if format == 'Percent':
		cell.number_format = '0.0%'

###############################################################################

suivi_abl1_path = "/media/n06lbth/sauvegardes_pgm/ABL1/EN_LAB_19_2333_Suivi_temoin_ABL1.xlsx"
temoin_abl1_finalreport_path = sys.argv[1]
sample = sys.argv[2]
run_name = sys.argv[3]
run_name = run_name.replace('Auto_user_S5-0198','S5')


###############################################################################

# i/o
fp = openpyxl.load_workbook(temoin_abl1_finalreport_path)
annotation_sheet = fp.get_sheet_by_name('Annotation')
annotation_rows = tuple(annotation_sheet.rows)

suivi_abl1 = openpyxl.load_workbook(suivi_abl1_path)
suivi_sheet = suivi_abl1.get_sheet_by_name('Temoin ABL1')

img = openpyxl.drawing.image.Image('/DATA/work/scripts/ChuBordeaux_small.png')
suivi_sheet.add_image(img,'A1')

row2write = suivi_sheet.max_row + 1
variant2check = (suivi_sheet.cell(row=10,column=3).value,suivi_sheet.cell(row=10,column=7).value) # NM, c.

# run et sample
suivi_sheet.cell(row=row2write,column=1).value = sample+'_'+run_name
#suivi_sheet.merge_cells(start_row=row2write,start_column=1,end_row=row2write,end_column=3)
cell_format(suivi_sheet.cell(row=row2write,column=1),color='LightBlue',font='bold',exterior_border=True)
cell_format(suivi_sheet.cell(row=row2write,column=2),color='LightBlue',font='bold',exterior_border=True)
cell_format(suivi_sheet.cell(row=row2write,column=3),color='LightBlue',font='bold',exterior_border=True)

#suivi_sheet.merge_cells(start_row=row2write,start_column=7,end_row=row2write,end_column=8)
cell_format(suivi_sheet.cell(row=row2write,column=7),border=True)

# variants lines
for i in range(len(annotation_rows[0])):
	if annotation_rows[0][i].value == 'Transcript':
		nm_index = i
	if annotation_rows[0][i].value == 'c.':
		c_index = i
	if annotation_rows[0][i].value == 'c.(annovar)':
		annovar_index = i
	if annotation_rows[0][i].value == 'Var.Freq.' or annotation_rows[0][i].value == 'Freq':
		freq_index = i
	if annotation_rows[0][i].value == 'Var.Cov.':
		var_cov_index = i	
	if annotation_rows[0][i].value == 'Pos.Cov.' or annotation_rows[0][i].value == 'Depth':
		pos_cov_index = i	

list_not_found = []
for j in range(1,len(annotation_rows)):
	variant = (annotation_rows[j][nm_index].value.split('.')[0],annotation_rows[j][c_index].value)
	variant_annovar = (annotation_rows[j][nm_index].value.split('.')[0],annotation_rows[j][annovar_index].value)
	variant_freq = '?'
	if (variant2check == variant) or (variant2check == variant_annovar):
		variant_freq = annotation_rows[j][freq_index].value
		break
if variant_freq == '?': # not found!
	suivi_sheet.cell(row=row2write,column=4).value = '?'
	suivi_sheet.cell(row=row2write,column=5).value = '?'
	suivi_sheet.cell(row=row2write,column=6).value = '?'
	cell_format(suivi_sheet.cell(row=row2write,column=4),font='bold',color='LightRed',border=True)
	cell_format(suivi_sheet.cell(row=row2write,column=5),font='bold',color='LightRed',border=True)
	cell_format(suivi_sheet.cell(row=row2write,column=6),font='bold',color='LightRed',border=True)
	list_not_found.append(variant2check)
else:
	suivi_sheet.cell(row=row2write,column=4).value = representsInt(variant_freq)
	suivi_sheet.cell(row=row2write,column=5).value = representsInt(annotation_rows[j][var_cov_index].value)
	suivi_sheet.cell(row=row2write,column=6).value = representsInt(annotation_rows[j][pos_cov_index].value)
	cell_format(suivi_sheet.cell(row=row2write,column=4),color='LightGreen',border=True)
	cell_format(suivi_sheet.cell(row=row2write,column=5),color='LightGreen',border=True)
	cell_format(suivi_sheet.cell(row=row2write,column=6),color='LightGreen',border=True)

suivi_abl1.save(suivi_abl1_path)
