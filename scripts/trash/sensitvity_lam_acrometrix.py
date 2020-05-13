#!/usr/bin/env python
import sys
import os
import openpyxl
import subprocess

def representsInt(s): # pour eviter avertissement "nombre ecrit en texte" sous excel
	try: 
		s = int(s)
		return s
	except ValueError:
		return s
		
def cell_format(cell, font=None, alignment=None, color=None, format=None, border=None):
	if font == 'bold':
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, bold=True)
	else:
		cell.font = openpyxl.styles.Font(name='Calibri', size=11)
	if alignment == 'center':
		cell.alignment = openpyxl.styles.Alignment(horizontal='center')
	elif alignment == 'left':
		cell.alignment = openpyxl.styles.Alignment(horizontal='left')
	if color == 'LightGreen':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='D8E4BC') # EBF1DE
	elif color == 'LightRed':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='d28e8e') #F2DCDB
	elif color == 'Yellow':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='feffa3')
	elif color == 'Blue':
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, color='004c99')
	else:
		cell.fill = openpyxl.styles.PatternFill(fill_type=None,start_color='FFFFFF')
	if border:
		cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'),bottom=openpyxl.styles.Side(style='thin'))
	if format == 'Percent':
		cell.number_format = '0.0%'

###############################################################################

suivi_acro_path = "/media/n06lbth/sauvegardes_pgm/tests_qualification_pipeline/Specificity/HEMATO/Suivi_temoin_Acrometrix.xlsx"
acrometrix_finalreport_path = sys.argv[1]
sample = sys.argv[2]
run_name = sys.argv[3]

###############################################################################

# i/o
fp = openpyxl.load_workbook(acrometrix_finalreport_path)
annotation_sheet = fp.get_sheet_by_name('Annotation')
annotation_rows = tuple(annotation_sheet.rows)

suivi_acro = openpyxl.load_workbook(suivi_acro_path)
suivi_sheet = suivi_acro.get_sheet_by_name('Acrometrix')
suivi_rows = tuple(suivi_sheet.rows)

column2write = len(suivi_rows[0])+1

# header 1
suivi_sheet.cell(row=1,column=column2write).value = sample+'_'+run_name
cell_format(suivi_sheet.cell(row=1,column=column2write),font='bold',border=True)

# header 2
suivi_sheet.cell(row=2,column=column2write).value = 'Var.freq'
cell_format(suivi_sheet.cell(row=2,column=column2write),border=True)

# variants lines
for i in range(len(annotation_rows[0])):
	if annotation_rows[0][i].value == 'Transcript':
		nm_index = i
	if annotation_rows[0][i].value == 'Gene':
		gene_index = i
	if annotation_rows[0][i].value == 'c.':
		c_index = i
	if annotation_rows[0][i].value == 'c.(annovar)':
		annovar_index = i
	if annotation_rows[0][i].value == 'Var.Freq.':
		freq_index = i	

list_not_found = []
for i in range(2,len(suivi_rows)):
	variant2check = (suivi_rows[i][2].value,suivi_rows[i][5].value) # GENE, c.
	for j in range(1,len(annotation_rows)):
		variant = (annotation_rows[j][gene_index].value,annotation_rows[j][c_index].value)
		variant_annovar = (annotation_rows[j][gene_index].value,annotation_rows[j][annovar_index].value)
		variant_freq = '?'
		if (variant2check == variant) or (variant2check == variant_annovar):
			variant_freq = annotation_rows[j][freq_index].value
			break
	if variant_freq == '?': # not found!
		suivi_sheet.cell(row=i+1,column=column2write).value = representsInt(variant_freq)
		cell_format(suivi_sheet.cell(row=i+1,column=column2write),font='bold',color='LightRed',border=True)
		list_not_found.append(variant2check)
	else:
		suivi_sheet.cell(row=i+1,column=column2write).value = representsInt(variant_freq)
		cell_format(suivi_sheet.cell(row=i+1,column=column2write),border=True)

suivi_acro.save(suivi_acro_path)

# TODO : lancer un checkMut en automatique pour les variants non trouves?
#if list_not_found:
	#run_path = os.path.dirname(os.path.dirname(acrometrix_finalreport_path))
	#if not os.path.isdir(run_path+'/_checkMut'):
		#subprocess.call(['mkdir', run_path+'/_checkMut'])
	#for variant in set(list_not_found):
		#subprocess.call(['python','/DATA/work/scripts/checkMut.py','--run-folder',run_path,'--nm',variant[0],'cpos',variant[1])
		
		## reverse ou pas
		## dossier acrometrix serait mieux
