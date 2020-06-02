##!/usr/bin/env python
import subprocess
import openpyxl
import json
import sys

def json_loads_byteified(json_text):
    return _byteify(
        json.loads(json_text, object_hook=_byteify),
        ignore_dicts=True
    )

def _byteify(data, ignore_dicts = False):
    # if this is a unicode string, return its string representation
    if isinstance(data, unicode):
        return data.encode('utf-8')
    # if this is a list of values, return list of byteified values
    if isinstance(data, list):
        return [ _byteify(item, ignore_dicts=True) for item in data ]
    # if this is a dictionary, return dictionary of byteified keys and values
    # but only if we haven't already byteified it
    if isinstance(data, dict) and not ignore_dicts:
        return {
            _byteify(key, ignore_dicts=True): _byteify(value, ignore_dicts=True)
            for key, value in data.iteritems()
        }
    # if it's anything else, return it in its original form
    return data
    
def cell_format(cell, font=None, alignment=None, color=None, format=None, border=None, percent=False, alert=False):
	cell.font = openpyxl.styles.Font(name='Calibri', size=11)
	if alert==True:
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, color='FF0000')
	if alignment == 'center':
		cell.alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrap_text=True)
	if color == 'LightGreen':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='EBF1DE')
	elif color == 'Green':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='D8E4BC')
	elif color == 'LightOrange':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='FDE9D9')
	elif color == 'Orange':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='FCD5B4')
	elif color == 'Blue':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='DCE6F1')
	else:
		cell.fill = openpyxl.styles.PatternFill(fill_type=None,start_color='FFFFFF')
	#if border:
		#cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'),bottom=openpyxl.styles.Side(style='thin'))
	if percent:
		cell.number_format = '0.0%'


# i/o
path = '/media/n06lbth/sauvegardes_pgm/FusionPlex_CTL/Run_Statistics.xlsx'
wb = openpyxl.load_workbook(path)
sheet = wb.get_sheet_by_name('all_run_statistics')
rows = tuple(sheet.rows)
line2write = len(rows)+1

runs = []
for row in rows:
	if row[0].value not in runs:
		runs.append(row[0].value)

r = subprocess.check_output('curl -X GET "http://10.67.63.93/rest_api/jobs/" -H "accept: application/json" -H "authorization: Basic c2J0QGNodS1ib3JkZWF1eC5mcjpzYnQxMjNTQlQh" -H "X-CSRFToken: pBJkg5aAfS0LAb33LLQBwbVlnyBfIM0yzKsCLxw2xdVbzQaEGcFXNR31qdO00SIA"',shell=True)
j = json_loads_byteified(r)
for result in j['results']:
	job_name = result['name']
	if job_name in runs:
		continue
	job_id = result['job_id']
	r2 = subprocess.check_output('curl -X GET "http://10.67.63.93/rest_api/jobs/%s/" -H "accept: application/json" -H "authorization: Basic c2J0QGNodS1ib3JkZWF1eC5mcjpzYnQxMjNTQlQh" -H "X-CSRFToken: pBJkg5aAfS0LAb33LLQBwbVlnyBfIM0yzKsCLxw2xdVbzQaEGcFXNR31qdO00SIA"' % job_id,shell=True)
	j2 = json_loads_byteified(r2)
	start_merge = line2write
	for sample in j2['samples']:
		sample_name = sample['name'].replace('.unaligned','')
		sample_id = sample['id']
		r3 = subprocess.check_output('curl -X GET "http://10.67.63.93/rest_api/jobs/%s/samples/%s/results/read-stats" -H "accept: application/json" -H "authorization: Basic c2J0QGNodS1ib3JkZWF1eC5mcjpzYnQxMjNTQlQh" -H "X-CSRFToken: pBJkg5aAfS0LAb33LLQBwbVlnyBfIM0yzKsCLxw2xdVbzQaEGcFXNR31qdO00SIA"' % (job_id,sample_id),shell=True)
		j3 = json_loads_byteified(r3)
		total_fragments = j3['molbar_stats']['total_molbar_reads']
		for rstat in j3['read_stat_types']:
			if rstat['read_type'] == 'All Fragments':
				mapped_reads = rstat['mapped_num']
				break
		rna_median_fragment_length = j3['fragment_means_and_medians']['rna_median_length']
		for stat in j3['total_stats']:
			if stat['frag_type'] == 'All Fragments':
				dna_percent = stat['dna_reads_percent']
				dna_reads = stat['dna_reads']
				rna_percent = stat['rna_reads_percent']
				rna_reads = stat['rna_reads']
				break
				
		sheet.cell(row=line2write,column=1).value = job_name
		sheet.cell(row=line2write,column=2).value = sample_name
		sheet.cell(row=line2write,column=3).value = total_fragments
		cell_format(sheet.cell(row=line2write,column=3),color='LightGreen')
		
		if 'PATIENT-INCONNU' in sample_name:
			cell_format(sheet.cell(row=line2write,column=2),alert=True)
			cell_format(sheet.cell(row=line2write,column=3),color='LightGreen',alert=True)
		
		sheet.cell(row=line2write,column=4).value = mapped_reads
		cell_format(sheet.cell(row=line2write,column=4),color='Green')
		sheet.cell(row=line2write,column=5).value = dna_reads
		cell_format(sheet.cell(row=line2write,column=5),color='Blue')
		sheet.cell(row=line2write,column=6).value = dna_percent
		cell_format(sheet.cell(row=line2write,column=6),color='Blue')
		sheet.cell(row=line2write,column=7).value = rna_reads
		cell_format(sheet.cell(row=line2write,column=7),color='LightOrange')
		sheet.cell(row=line2write,column=8).value = rna_percent
		cell_format(sheet.cell(row=line2write,column=8),color='LightOrange')
		sheet.cell(row=line2write,column=9).value = rna_median_fragment_length
		cell_format(sheet.cell(row=line2write,column=9),color='Orange')
		
		line2write = line2write+1
		
		
	# merge
	sheet.merge_cells(start_row=start_merge, start_column=1, end_row=line2write-1, end_column=1)
	cell_format(sheet.cell(row=start_merge,column=1),alignment='center')

wb.save(path)
