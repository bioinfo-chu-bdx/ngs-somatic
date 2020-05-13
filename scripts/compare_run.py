#!/usr/bin/env python
import sys
import os
import glob
import openpyxl
import datetime

def cell_format(cell, font=None, alignment=None, color=None, format=None, border=None):
	if font == 'bold':
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, bold=True)
	elif font == 'red':
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, color='f44242')
	else:
		cell.font = openpyxl.styles.Font(name='Calibri', size=11)
	if alignment == 'center':
		cell.alignment = openpyxl.styles.Alignment(horizontal='center')
	elif alignment == 'vertical-center':
		cell.alignment = openpyxl.styles.Alignment(vertical='center')
	elif alignment == 'left':
		cell.alignment = openpyxl.styles.Alignment(horizontal='left')
	if color == 'LightGreen':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='D8E4BC') # EBF1DE
	elif color == 'LightRed':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='d28e8e') #F2DCDB
	elif color == 'LightOrange':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='ffe7e0') # ffd8ad
	elif color == 'LightBlue':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='add8e6')
	elif color == 'LightPink':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='ffc9f7')
	elif color == 'LightPurple':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='E0DBE9')
	elif color == 'Yellow':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='feffa3')
	elif color == 'Blue':
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, color='004c99')
	else:
		cell.fill = openpyxl.styles.PatternFill(fill_type=None,start_color='FFFFFF')
	if border == 'thin':
		cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'),bottom=openpyxl.styles.Side(style='thin'))
	if border == 'top_medium':
		cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='medium'))
	if format == 'Percent':
		cell.number_format = '0.0%'


control_names = ['H2O','H20','NTC'] # liste des noms possibles pour les temoins negatifs
runs = []
if len(sys.argv) < 3:
	print "error : need at least 2 runs to compare"
	sys.exit()
for i in range(1,len(sys.argv)):
	print "* run %s : %s" % (i,sys.argv[i])
	runs.append(sys.argv[i])

#### LES RUNS DOIVENT AVOIR DES NOMS DIFFERENTS

compare_green_only = True
date = datetime.datetime.now().strftime("%y-%m-%d-%H-%M")
output_name = 'compare_run_%s' % date

output = '/media/stuff/%s.xlsx' % output_name
workbook = openpyxl.Workbook()
sheet = workbook.create_sheet(title=output_name)
try:
	del workbook['Sheet']
except:
	pass

data = {}
run_names = []

for i in range(2):
	run = runs[i]
	runname = run.split('/')[-1]
	if runname == '':
		runname = run.split('/')[-2]
	if not runname in run_names:
		run_names.append(runname)
	else: # deja utilise
		for i in range(2,10):
			new_runname = runname + '_' + str(i)
			if new_runname not in run_names:
				runname = new_runname
				run_names.append(runname)
				break
			else:
				continue
	finalreportlist = glob.glob(run+'/*/*.finalReport.xlsx')
	if i == 1:
		finalreportlist = glob.glob(run+'/*/*.finalReport.old.xlsx')
	finalreportlist = [item for item in finalreportlist if not '~' in item]
	
	for control_name in control_names:
		finalreportlist = [item for item in finalreportlist if not control_name in item.upper()]
	
	for finalreport in finalreportlist:
		print finalreport
		sample = finalreport.split('/')[-1].split('_IonXpress')[0]
		print sample
		barcode = 'IonXpress_' + finalreport.split('IonXpress_')[-1].split('.bam')[0]
		sample_folder = os.path.dirname(finalreport)
		#finalreport = sample_folder + '/' + sample+'_'+barcode+'.finalReport.xlsx'
		finalreport_workbook = openpyxl.load_workbook(finalreport)
		annosheet = finalreport_workbook['Annotation']
		
		# col index of the finalReport
		col2index = {}
		for j in range(1,annosheet.max_column+1):
			col2index[annosheet.cell(row=1,column=j).value] = j

		if not sample in data:
			data[sample] = {}
		
		for row_idx in range(2, annosheet.max_row+1):
			gene = annosheet.cell(row=row_idx,column=col2index['Gene']).value
			cpos = annosheet.cell(row=row_idx,column=col2index['c.']).value
			region = annosheet.cell(row=row_idx,column=col2index['Region']).value
			try:
				_type = annosheet.cell(row=row_idx,column=col2index['Type']).value
			except : 
				_type = annosheet.cell(row=row_idx,column=col2index['Consequence']).value
			if cpos == None:
				continue
			if compare_green_only :
				if ((region not in ['intronic','UTR3','UTR5','ncRNA_intronic']) and (_type != 'synonymous')):
					pass
				else:
					continue
			freq = annosheet.cell(row=row_idx,column=col2index['Freq']).value
			try:
				poscov = annosheet.cell(row=row_idx,column=col2index['Pos.Cov.']).value
			except:
				poscov = annosheet.cell(row=row_idx,column=col2index['Depth']).value
			if not((gene,cpos,region,_type) in data[sample]):
				data[sample][(gene,cpos,region,_type)] = {}
			data[sample][(gene,cpos,region,_type)][run] = (freq,poscov)

# writing header
c = 4
for run in runs:
	#runname = run.split('/')[-1]
	#if runname == '':
		#runname = run.split('/')[-2]
	sheet.cell(row=1,column=c).value = run#runname
	cell_format(sheet.cell(row=1,column=c),font='bold',border='thin')
	sheet.merge_cells(start_row=1,start_column=c,end_row=1,end_column=c+1)
	c = c + 2
	
header = ['PATIENT','GENE','MUTATION','FREQ','POSCOV','FREQ','POSCOV']
for i in range(len(header)):
	sheet.cell(row=2,column=i+1).value = header[i]
	cell_format(sheet.cell(row=2,column=i+1),font='bold',border='thin')

r = 3
for sample in sorted(data.keys()):
	for variant in sorted(data[sample].keys()):
		#var = data[sample][variant]
		gene = variant[0]
		cpos = variant[1]
		region = variant[2]
		_type = variant[3]
		sheet.cell(row=r,column=1).value = sample
		sheet.cell(row=r,column=2).value = gene
		sheet.cell(row=r,column=3).value = cpos
		if ((region not in ['intronic','UTR3','UTR5','ncRNA_intronic']) and (_type != 'synonymous')):
			for c in range(2,4+(2*len(runs))):
				cell_format(sheet.cell(row=r,column=c),color='LightGreen')
		if ('+' in cpos or '-' in cpos or '*' in cpos) and (region != 'exonic'):
			cell_format(sheet.cell(row=r,column=3),color='Blue')

		c = 4
		for run in runs :
			#runname = run.split('/')[-1]
			#if runname == '':
				#runname = run.split('/')[-2]
			if not(run in data[sample][variant]):
				freq = '?'
				poscov = '?'
				cell_format(sheet.cell(row=r,column=c),color='LightRed')
				sheet.merge_cells(start_row=r,start_column=c,end_row=r,end_column=c+1)
			else:
				freq = data[sample][variant][run][0]
				poscov = data[sample][variant][run][1]
			sheet.cell(row=r,column=c).value = freq
			sheet.cell(row=r,column=c+1).value = poscov
			c = c + 2
		r = r + 1

workbook.save(output)
