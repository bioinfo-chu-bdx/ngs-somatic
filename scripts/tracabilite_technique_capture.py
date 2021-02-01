#!/usr/bin/env python
# -*- coding: utf-8 -*-
import sys
import json
import os
import math
import openpyxl
import urllib2
import zipfile
import xmltodict


def cell_format(cell, font=None, alignment=None, color=None, format=None, border=None):
	if font == 'bold':
		cell.font = openpyxl.styles.Font(name='Calibri', size=11, bold=True)
	else:
		cell.font = openpyxl.styles.Font(name='Calibri', size=11)
	if alignment == 'center':
		cell.alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrap_text=True)
	elif alignment == 'left':
		cell.alignment = openpyxl.styles.Alignment(horizontal='left')
	if color == 'LightGreen':
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='D8E4BC') # EBF1DE
	elif color == 'Orange':	
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='ffc000')		
	elif color == 'LightOrange':	
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='ffe7e0') # ffd8ad
	elif color == 'LightBlue1':	
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='c5d9f1')
	elif color == 'LightBlue2':	
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='add8e6')
	elif color == 'Cyan':	
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='ccffff')	
	elif color == 'LightPink':	
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='ffc9f7')
	elif color == 'LightPurple':	
		cell.fill = openpyxl.styles.PatternFill(fill_type='solid',start_color='E0DBE9')
	else:
		cell.fill = openpyxl.styles.PatternFill(fill_type=None,start_color='FFFFFF')
	if border == 'thin':
		cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'),bottom=openpyxl.styles.Side(style='thin'))
	if border == 'top_medium':
		cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='medium'))
	if format == 'Percent':
		cell.number_format = '0.0%'

def representsInt(s): # pour eviter avertissement "nombre ecrit en texte" sous excel
	try: 
		s = int(s)
		return s
	except ValueError:
		return s

pipeline_folder = os.environ['NGS_PIPELINE_BX_DIR']
illumina_folder = sys.argv[1]
tracabilite_xlsx_path = sys.argv[2]
run_name = sys.argv[3]

# PARSE XLM RUN PARAMETERS
with open('%s/RunParameters.xml' % illumina_folder,'r') as run_parameters_xml:
	xmldict = xmltodict.parse(run_parameters_xml)

# i/o
workbook = openpyxl.load_workbook(tracabilite_xlsx_path)
sheet = workbook.get_sheet_by_name(workbook.sheetnames[0])
col2write = sheet.max_column+1

img = openpyxl.drawing.image.Image('%s/scripts/ChuBordeaux.png' % pipeline_folder)
# img = openpyxl.drawing.image.Image('%s/scripts/ChuBordeaux_small.png' % pipeline_folder)
sheet.add_image(img,'A1')

# write run name
sheet.cell(row=7,column=col2write).value = run_name
cell_format(sheet.cell(row=7,column=col2write),font='bold',alignment='center',border='thin')
# write data
# Flow Cell - Serial Number
sheet.cell(row=24,column=col2write).value = xmldict['RunParameters']['FlowCellRfidTag']['SerialNumber']
cell_format(sheet.cell(row=24,column=col2write),alignment='center',border='thin',color='LightBlue1')
# Flow Cell - Part number
sheet.cell(row=25,column=col2write).value = representsInt(xmldict['RunParameters']['FlowCellRfidTag']['PartNumber'])
cell_format(sheet.cell(row=25,column=col2write),alignment='center',border='thin',color='LightBlue1')
# Flow Cell - Lot number
sheet.cell(row=26,column=col2write).value = representsInt(xmldict['RunParameters']['FlowCellRfidTag']['LotNumber'])
cell_format(sheet.cell(row=26,column=col2write),alignment='center',border='thin',color='LightBlue1')
# Flow Cell - Exp. date
sheet.cell(row=27,column=col2write).value = xmldict['RunParameters']['FlowCellRfidTag']['ExpirationDate']
cell_format(sheet.cell(row=27,column=col2write),alignment='center',border='thin',color='LightBlue1')
# Buffer - Serial Number
sheet.cell(row=28,column=col2write).value = xmldict['RunParameters']['PR2BottleRfidTag']['SerialNumber']
cell_format(sheet.cell(row=28,column=col2write),alignment='center',border='thin',color='LightPurple')
# Buffer - Part number
sheet.cell(row=29,column=col2write).value = representsInt(xmldict['RunParameters']['PR2BottleRfidTag']['LotNumber'])
cell_format(sheet.cell(row=29,column=col2write),alignment='center',border='thin',color='LightPurple')
# Buffer - Lot number
sheet.cell(row=30,column=col2write).value = xmldict['RunParameters']['PR2BottleRfidTag']['SerialNumber']
cell_format(sheet.cell(row=30,column=col2write),alignment='center',border='thin',color='LightPurple')
# Buffer - Exp. Date
sheet.cell(row=31,column=col2write).value = xmldict['RunParameters']['PR2BottleRfidTag']['ExpirationDate']
cell_format(sheet.cell(row=31,column=col2write),alignment='center',border='thin',color='LightPurple')
# Reagent Kit - Serial Number
sheet.cell(row=32,column=col2write).value = xmldict['RunParameters']['ReagentKitRfidTag']['SerialNumber']
cell_format(sheet.cell(row=32,column=col2write),alignment='center',border='thin',color='Cyan')
# Reagent Kit - Part number
sheet.cell(row=33,column=col2write).value = representsInt(xmldict['RunParameters']['ReagentKitRfidTag']['PartNumber'])
cell_format(sheet.cell(row=33,column=col2write),alignment='center',border='thin',color='Cyan')
# Reagent Kit - Lot number
sheet.cell(row=34,column=col2write).value = representsInt(xmldict['RunParameters']['ReagentKitRfidTag']['LotNumber'])
cell_format(sheet.cell(row=34,column=col2write),alignment='center',border='thin',color='Cyan')
# Reagent Kit  - Exp. Date
sheet.cell(row=35,column=col2write).value = xmldict['RunParameters']['ReagentKitRfidTag']['ExpirationDate']
cell_format(sheet.cell(row=35,column=col2write),alignment='center',border='thin',color='Cyan')

workbook.save(tracabilite_xlsx_path)
