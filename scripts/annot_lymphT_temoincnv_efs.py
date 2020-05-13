#!/usr/bin/python
import sys
import os
import glob
import json
import openpyxl
import csv
from optparse import OptionParser

def get_column2index(sheet): # renvoie dict col2index
	col2index = {}
	for i in range(1,sheet.max_column+1):
		col2index[sheet.cell(row=1,column=i).value] = i
	return col2index

### GATHERING PARAMETERS ############################################################

parser = OptionParser()
parser.add_option('-b', '--bam',		help="Input bam file for SINGLE BAM ANALYSIS (must contain flow signal from TSS)", dest='bam')
parser.add_option('-r', '--full-run',	help="Run folder path  for FULL RUN ANALYSIS", dest='full_run') 
parser.add_option('-t', '--run-type',	help="Run type : SBT, LAM, TP53-HEMATO... (OPTIONAL, NOT NEEDED if barcodes.json is present in run folder)", dest='run_type') 
(options, args) = parser.parse_args()

if options.bam and options.full_run:
	sys.stderr.write("[run_analysis.py] Error: <--bam> and <--full-run> are not compatibles\n")
	sys.exit()
if options.full_run:
	bamlist = glob.glob(options.full_run+'/*/*.bam')
	bamlist = [item for item in bamlist if not 'processed' in item]
elif options.bam:
	bamlist = [options.bam]
else:
	sys.stderr.write("[run_analysis.py] Error: no <--bam> or <--full-run> specified\n")
	sys.exit()

pipeline_folder = os.environ['NGS_PIPELINE_BX_DIR']
with open('%s/global_parameters.json' % pipeline_folder, 'r') as g:
	global_param = json.loads(g.read().replace('$NGS_PIPELINE_BX_DIR',os.environ['NGS_PIPELINE_BX_DIR']))

temoin_variants = {}
temoinCNV_file = open('%s/scripts/temoinCNV_variants_lymphomeT.tsv' % pipeline_folder,'r')
temoinCNV_reader = csv.reader(temoinCNV_file,delimiter = '\t')
for line in temoinCNV_reader:
	temoin_variants[(line[0],line[1])] = line[2]

### RUN ANALYSIS ######################################################################

bam_data = {}
	
for bamfile in sorted(bamlist) :
	sample = bamfile.split('/')[-1].split('_IonXpress')[0]
	print "- %s" % sample
	barcode = 'IonXpress_' + bamfile.split('IonXpress_')[-1].split('.bam')[0]
	sample_folder = os.path.dirname(bamfile)
	run_folder = os.path.dirname(sample_folder)
	
	finalreport_path = sample_folder + '/' + sample+'_'+barcode+'.finalReport.xlsx'
	finalreport = openpyxl.load_workbook(finalreport_path)
	annoSheet = finalreport.get_sheet_by_name('Annotation')
	anncol = get_column2index(annoSheet)
	
	for i in range(2,annoSheet.max_row+1):
		if annoSheet.cell(row=i,column=anncol['Chr']).value == None: # avoid empty line and "Amplicons < 300X: " line
			continue
		nm = annoSheet.cell(row=i,column=anncol['Transcript']).value
		nm = nm.split('.')[0]
		c = annoSheet.cell(row=i,column=anncol['c.']).value
		if (nm,c) in temoin_variants:
			if annoSheet.cell(row=i,column=anncol['Commentaire']).value:
				annoSheet.cell(row=i,column=1).value = 'Found in Control ' + temoin_variants[(nm,c)] + ',' + annoSheet.cell(row=i,column=anncol['Commentaire']).value
			else:
				annoSheet.cell(row=i,column=1).value = 'Found in Control ' + temoin_variants[(nm,c)]
		#print annoSheet.cell(row=i,column=1).value

	finalreport.save(finalreport_path)
